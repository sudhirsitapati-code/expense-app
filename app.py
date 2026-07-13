"""
app.py
Flask app — all routes: WhatsApp webhook, HTML screens, JSON APIs.
"""

import json
import os
import threading
from datetime import datetime

from apscheduler.schedulers.background import BackgroundScheduler
from dotenv import load_dotenv
from flask import Flask, request, render_template, jsonify, session, redirect, url_for, send_from_directory

from src.approval_engine import ApprovalEngine, ExpenseRequest
from src.reconcile import run_reconciliation
from src.acc27_writer import sync_approved_to_history, export_monthly_excel
from src.icici_statement_parser import fetch_and_parse_statements
from src.master_ledger import (
    load_ledger, get_uncertain, update_transaction,
    sync_from_gmail as ledger_sync_gmail,
    get_cc_balance, reconcile_with_approvals,
    import_from_icici_transactions,
    repair_pdf_descriptions, deduplicate_ledger,
    LEDGER_PATH, _load_json as _ml_load_json, _save_json as _ml_save_json,
    _parse_date as _ml_parse_date,
)
from src.whatsapp_handler import (
    build_twiml_reply, parse_incoming,
    send_approval_request, send_approval_result,
    send_auto_approval_notice, send_clarification_request,
    send_to_sudhir,
    SUDHIR, HOUSEHOLD_MEMBERS,
)
from src import db

# ── FY26 static data (embedded so Railway deploy doesn't need data/ dir) ─────
_FY26_MONTHLY = {"Misc": {"Apr": 16685, "May": 836, "Jun": 16239, "Jul": 6208, "Aug": 59760, "Sep": 625, "Oct": 34168, "Nov": 16938, "Dec": 2676, "Jan": 6562, "Feb": 16537, "Mar": 11382}, "Clothes": {"Apr": 0, "May": 25990, "Jun": 54219, "Jul": 66865, "Aug": 64337, "Sep": 0, "Oct": 57982, "Nov": 38164, "Dec": 19461, "Jan": 8290, "Feb": 47896, "Mar": 7652}, "Gifts": {"Apr": 6500, "May": 3950, "Jun": 0, "Jul": 0, "Aug": 8825, "Sep": 0, "Oct": 40000, "Nov": 33099, "Dec": 0, "Jan": 3540, "Feb": 1571, "Mar": 2863}, "Cash": {"Apr": 10000, "May": 0, "Jun": 20000, "Jul": 25000, "Aug": 50000, "Sep": 10000, "Oct": 10000, "Nov": 80000, "Dec": 0, "Jan": 11600, "Feb": 31000, "Mar": 10000}, "Maintenance Expense": {"Apr": 17440, "May": 9814, "Jun": 12775, "Jul": 2411, "Aug": 3029, "Sep": 750, "Oct": 22850, "Nov": 5887, "Dec": 0, "Jan": 2500, "Feb": 6243, "Mar": 4404}, "Malhar": {"Apr": 40800, "May": 53930, "Jun": 93923, "Jul": 89728, "Aug": 51314, "Sep": 91350, "Oct": 177919, "Nov": 165493, "Dec": 0, "Jan": 81581, "Feb": 88880, "Mar": 30382}, "Home office": {"Apr": 88350, "May": 25461, "Jun": 10450, "Jul": 23057, "Aug": 24447, "Sep": 11933, "Oct": 9455, "Nov": 24783, "Dec": 0, "Jan": 7430, "Feb": 3096, "Mar": 5320}, "Electricity & Gas": {"Apr": 50780, "May": 52723, "Jun": 52068, "Jul": 31128, "Aug": 25843, "Sep": 24863, "Oct": 47174, "Nov": 38834, "Dec": 0, "Jan": 37325, "Feb": 24186, "Mar": 7616}, "Alcohol": {"Apr": 14900, "May": 0, "Jun": 0, "Jul": 9400, "Aug": 1450, "Sep": 0, "Oct": 0, "Nov": 48700, "Dec": 0, "Jan": 0, "Feb": 0, "Mar": 500}, "Medical": {"Apr": 14143, "May": 96808, "Jun": 27382, "Jul": 77477, "Aug": 56988, "Sep": 77126, "Oct": 72143, "Nov": 45345, "Dec": 1210, "Jan": 57526, "Feb": 20200, "Mar": 45100}, "Holiday": {"Apr": 17031, "May": 486489, "Jun": 132378, "Jul": 952200, "Aug": 73464, "Sep": 4218, "Oct": 93531, "Nov": 149259, "Dec": 0, "Jan": 8128, "Feb": 56366, "Mar": 215510}, "Groceries": {"Apr": 89175, "May": 120368, "Jun": 90011, "Jul": 106208, "Aug": 135331, "Sep": 172407, "Oct": 155555, "Nov": 253751, "Dec": 5404, "Jan": 220193, "Feb": 140067, "Mar": 200420}, "Eating Out": {"Apr": 46057, "May": 27258, "Jun": 25723, "Jul": 48862, "Aug": 49917, "Sep": 51657, "Oct": 69146, "Nov": 48007, "Dec": 20773, "Jan": 53623, "Feb": 36546, "Mar": 42347}, "Amma": {"Apr": 27679, "May": 16458, "Jun": 7276, "Jul": 10068, "Aug": 12826, "Sep": 52229, "Oct": 1375, "Nov": -1356, "Dec": 0, "Jan": 15595, "Feb": -26776, "Mar": -1823}, "Wellness": {"Apr": 63157, "May": 22672, "Jun": 38199, "Jul": 46159, "Aug": 13292, "Sep": 59132, "Oct": 52176, "Nov": 61175, "Dec": 0, "Jan": 7745, "Feb": 13545, "Mar": 10649}, "Ketki": {"Apr": 426656, "May": 115151, "Jun": -15927, "Jul": 78940, "Aug": 126146, "Sep": 50147, "Oct": 326622, "Nov": 203580, "Dec": 65053, "Jan": 40084, "Feb": 27653, "Mar": 101024}, "Staff Salary": {"Apr": 118522, "May": 258292, "Jun": 269854, "Jul": 195408, "Aug": 178854, "Sep": 176859, "Oct": 188398, "Nov": 419898, "Dec": 99438, "Jan": 279669, "Feb": 224518, "Mar": 221018}, "Financial Expense / OD Interest": {"Apr": 0, "May": 0, "Jun": 0, "Jul": 0, "Aug": 0, "Sep": 0, "Oct": 0, "Nov": 0, "Dec": 0, "Jan": 0, "Feb": 11799, "Mar": -3245}, "Entertainment": {"Apr": 1148, "May": 5278, "Jun": 68503, "Jul": 40408, "Aug": 14280, "Sep": 21692, "Oct": 19600, "Nov": 15839, "Dec": 9485, "Jan": 4214, "Feb": 649, "Mar": 25143}, "One Time Charge": {"Apr": 45325, "May": 12530, "Jun": 4750, "Jul": 79984, "Aug": 76574, "Sep": 149914, "Oct": 64441, "Nov": 199349, "Dec": 71035, "Jan": 24050, "Feb": 31250, "Mar": 11850}, "Children Education": {"Apr": 981577, "May": 68350, "Jun": 550856, "Jul": 101149, "Aug": 124383, "Sep": 54100, "Oct": 38700, "Nov": 514160, "Dec": 0, "Jan": 29700, "Feb": 45500, "Mar": 524550}, "Kalpataru Maintenance": {"Apr": 35997, "May": 35933, "Jun": 35879, "Jul": -7041, "Aug": 197231, "Sep": 36587, "Oct": 35879, "Nov": 71168, "Dec": 0, "Jan": 35879, "Feb": 35879, "Mar": 35933}, "Charity": {"Apr": 20000, "May": 0, "Jun": 0, "Jul": 167800, "Aug": 105725, "Sep": 24048, "Oct": 300000, "Nov": 115000, "Dec": 150000, "Jan": 625100, "Feb": 0, "Mar": 18000}, "Uspaar": {"Apr": 133783, "May": 140978, "Jun": 214372, "Jul": 97910, "Aug": 204600, "Sep": 72828, "Oct": 38984, "Nov": 170003, "Dec": 0, "Jan": 103660, "Feb": 40350, "Mar": 86325}, "Insurance": {"Apr": 118000, "May": 0, "Jun": 0, "Jul": 0, "Aug": 0, "Sep": 0, "Oct": 102820, "Nov": 0, "Dec": 142800, "Jan": 0, "Feb": 0, "Mar": 0}, "Home Loan": {"Apr": 903259, "May": 1146721, "Jun": 916551, "Jul": 1201557, "Aug": 554547, "Sep": 992227, "Oct": 963511, "Nov": 1007410, "Dec": 1777170, "Jan": 1389117, "Feb": 928825, "Mar": 1285719}, "Tax": {"Apr": 0, "May": 0, "Jun": 0, "Jul": 13519347, "Aug": 0, "Sep": 3630000, "Oct": 61499379, "Nov": 0, "Dec": 0, "Jan": 0, "Feb": 3848950, "Mar": 0}}
_FY26_ACTUALS  = {"Misc": 189000, "Cash": 258000, "Electricity & Gas": 393000, "Groceries": 1689000, "Staff Salary": 2631000, "Alcohol": 75000, "Wellness": 388000, "Clothes": 391000, "Gifts": 100000, "Medical": 591000, "Amma": 114000, "Ketki": 1545000, "Children Education": 3033000, "Charity": 1526000, "Uspaar": 1304000, "Holiday": 2189000, "Eating Out": 520000, "Entertainment": 226000, "Malhar": 965000, "Maintenance Expense": 88000, "Home office": 234000, "One Time Charge": 771000, "Kalpataru Maintenance": 549000, "Financial Expense / OD Interest": 9000, "Insurance": 364000, "Home Loan": 13067000, "Tax": 45100000}

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "change-me")

engine = ApprovalEngine()

BASE_DIR = os.path.dirname(__file__)
DATA_DIR = os.path.join(BASE_DIR, "data")
CONFIG_DIR = os.path.join(BASE_DIR, "config")

# Legacy file paths kept for local dev fallback; production uses db module
APPROVAL_LOG = os.path.join(DATA_DIR, "approval_log.json")
RECONCILE_LOG = os.path.join(DATA_DIR, "reconcile_log.json")
TRANSACTIONS_PATH = os.path.join(DATA_DIR, "icici_transactions.json")

db.init_db()

PENDING_CLARIFICATION: dict = {}
NUMBER_TO_NAME = {v: k for k, v in HOUSEHOLD_MEMBERS.items() if v}


PASSWORDS = {
    "vincent": os.getenv("VINCENT_PASSWORD", "vincent123"),
    "sudhir":  os.getenv("SUDHIR_PASSWORD",  "sudhir123"),
    "vivek":   os.getenv("VIVEK_PASSWORD",   "Vivek123"),
}


def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


# ── Helpers ──────────────────────────────────────────────────────────────────

_PATH_TO_KEY = {
    APPROVAL_LOG:      "approval_log",
    RECONCILE_LOG:     "reconcile_log",
    TRANSACTIONS_PATH: "icici_transactions",
    LEDGER_PATH:       "master_ledger",
}

def _load_json(path):
    key = _PATH_TO_KEY.get(path)
    if key:
        return db.load(key)
    if not os.path.exists(path):
        return []
    with open(path) as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            return []


def _save_json(path, data):
    key = _PATH_TO_KEY.get(path)
    if key:
        db.save(key, data)
        return
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w") as f:
        json.dump(data, f, indent=2)


def _update_log_entry(request_id: str, updates: dict):
    log = db.load("approval_log")
    for entry in log:
        if entry.get("request_id") == request_id:
            entry.update(updates)
            break
    db.save("approval_log", log)


# ── AUTH ─────────────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        user = request.form.get("user", "").lower()
        password = request.form.get("password", "")
        if PASSWORDS.get(user) == password:
            session["user"] = user
            return redirect(url_for("index"))
        return render_template("login.html", error="Incorrect name or password.")
    return render_template("login.html", error=None)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ── HTML SCREENS ─────────────────────────────────────────────────────────────

@app.route("/")
@app.route("/home")
@login_required
def home():
    return render_template("home.html", user=session["user"])


@app.route("/expenses")
@login_required
def index():
    return render_template("index.html", user=session["user"])


@app.route("/tax")
@login_required
def tax():
    return render_template("tax.html", user=session["user"])

@app.route("/api/tax/ledger-income/<fy>")
@login_required
def tax_ledger_income(fy):
    """Return credit entries from the master ledger for the given tax FY.
    Tax FY26 (Apr 2025–Mar 2026) = ledger fy_year 2025.
    Tax FY27 (Apr 2026–Mar 2027) = ledger fy_year 2026.
    """
    fy_map = {"FY25": 2024, "FY26": 2025, "FY27": 2026}
    ledger_year = fy_map.get(fy)
    if not ledger_year:
        return jsonify({"error": "Unknown FY"}), 400

    ledger = db.load("master_ledger") or []
    entries = [t for t in ledger if t.get("fy_year") == ledger_year]

    from collections import defaultdict
    credits = defaultdict(float)
    samples = defaultdict(list)
    for t in entries:
        cr = float(t.get("credit") or 0)
        if cr < 10:
            continue
        heading = t.get("heading") or ""
        credits[heading] += cr
        if len(samples[heading]) < 5:
            samples[heading].append({
                "date": t.get("date",""),
                "paid_to": t.get("paid_to",""),
                "amount": cr,
                "account": t.get("account",""),
            })

    result = []
    for heading, total in sorted(credits.items(), key=lambda x: -x[1]):
        result.append({
            "heading": heading or "(no heading)",
            "total": total,
            "samples": samples[heading],
        })
    return jsonify({"fy": fy, "ledger_year": ledger_year, "credits": result})

@app.route("/financial-statements")
@login_required
def financial_statements():
    return render_template("financial_statements.html", user=session["user"])


# Keep old routes redirecting to /expenses
@app.route("/entry")
@app.route("/dashboard")
@app.route("/report")
def redirect_old():
    return redirect(url_for("index"))


# ── JSON APIs ─────────────────────────────────────────────────────────────────

@app.route("/api/expenses", methods=["GET"])
@login_required
def api_expenses():
    return jsonify(_load_json(APPROVAL_LOG))


@app.route("/api/transactions", methods=["GET"])
def api_transactions():
    return jsonify(_load_json(TRANSACTIONS_PATH))


@app.route("/api/reconcile-log", methods=["GET"])
def api_reconcile_log():
    return jsonify(_load_json(RECONCILE_LOG))


# ── MIS monthly data — module-level so both api_mis and api_financial_statements can use them ──
MIS_FY26_MONTHLY = {
    "Misc":           {"Apr":16685,"May":836,"Jun":16239,"Jul":6208,"Aug":59760,"Sep":625,"Oct":34168,"Nov":16938,"Dec":2676,"Jan":6562,"Feb":16537,"Mar":11382},
    "Clothes":        {"Apr":0,"May":25990,"Jun":54219,"Jul":66865,"Aug":64337,"Sep":0,"Oct":57982,"Nov":38164,"Dec":19461,"Jan":8290,"Feb":47896,"Mar":7652},
    "Gifts":          {"Apr":6500,"May":3950,"Jun":0,"Jul":0,"Aug":8825,"Sep":0,"Oct":40000,"Nov":33099,"Dec":0,"Jan":3540,"Feb":1571,"Mar":2863},
    "Cash":           {"Apr":10000,"May":0,"Jun":20000,"Jul":25000,"Aug":50000,"Sep":10000,"Oct":10000,"Nov":80000,"Dec":0,"Jan":11600,"Feb":31000,"Mar":10000},
    "Maintenance Expense": {"Apr":17440,"May":9814,"Jun":12775,"Jul":2411,"Aug":3029,"Sep":750,"Oct":22850,"Nov":5887,"Dec":0,"Jan":2500,"Feb":6243,"Mar":4404},
    "Malhar":         {"Apr":40800,"May":53930,"Jun":93923,"Jul":89728,"Aug":51314,"Sep":91350,"Oct":177919,"Nov":165493,"Dec":0,"Jan":81581,"Feb":88880,"Mar":30382},
    "Home office":    {"Apr":88350,"May":25461,"Jun":10450,"Jul":23057,"Aug":24447,"Sep":11933,"Oct":9455,"Nov":24783,"Dec":0,"Jan":7430,"Feb":3096,"Mar":5320},
    "Electricity & Gas": {"Apr":50780,"May":52723,"Jun":52068,"Jul":31128,"Aug":25843,"Sep":24863,"Oct":47174,"Nov":38834,"Dec":0,"Jan":37325,"Feb":24186,"Mar":7616},
    "Alcohol":        {"Apr":14900,"May":0,"Jun":0,"Jul":9400,"Aug":1450,"Sep":0,"Oct":0,"Nov":48700,"Dec":0,"Jan":0,"Feb":0,"Mar":500},
    "Medical":        {"Apr":14143,"May":96808,"Jun":27382,"Jul":77477,"Aug":56988,"Sep":77126,"Oct":72143,"Nov":45345,"Dec":1210,"Jan":57526,"Feb":20200,"Mar":45100},
    "Holiday":        {"Apr":17031,"May":486489,"Jun":132378,"Jul":952200,"Aug":73464,"Sep":4218,"Oct":93531,"Nov":149259,"Dec":0,"Jan":8128,"Feb":56366,"Mar":215510},
    "Groceries":      {"Apr":89175,"May":120368,"Jun":90011,"Jul":106208,"Aug":135331,"Sep":172407,"Oct":155555,"Nov":253751,"Dec":5404,"Jan":220193,"Feb":140067,"Mar":200420},
    "Eating Out":     {"Apr":46057,"May":27258,"Jun":25723,"Jul":48862,"Aug":49917,"Sep":51657,"Oct":69146,"Nov":48007,"Dec":20773,"Jan":53623,"Feb":36546,"Mar":42347},
    "Amma":           {"Apr":27679,"May":16458,"Jun":7276,"Jul":10068,"Aug":12826,"Sep":52229,"Oct":1375,"Nov":-1356,"Dec":0,"Jan":15595,"Feb":-26776,"Mar":-1823},
    "Wellness":       {"Apr":63157,"May":22672,"Jun":38199,"Jul":46159,"Aug":13292,"Sep":59132,"Oct":52176,"Nov":61175,"Dec":0,"Jan":7745,"Feb":13545,"Mar":10649},
    "Ketki":          {"Apr":426656,"May":115151,"Jun":-15927,"Jul":78940,"Aug":126146,"Sep":50147,"Oct":326622,"Nov":203580,"Dec":65053,"Jan":40084,"Feb":27653,"Mar":101024},
    "Staff Salary":   {"Apr":118522,"May":258292,"Jun":269854,"Jul":195408,"Aug":178854,"Sep":176859,"Oct":188398,"Nov":419898,"Dec":99438,"Jan":279669,"Feb":224518,"Mar":221018},
    "Financial Expense / OD Interest": {"Apr":0,"May":0,"Jun":0,"Jul":0,"Aug":0,"Sep":0,"Oct":0,"Nov":0,"Dec":0,"Jan":0,"Feb":0,"Mar":360000},
    "Entertainment":  {"Apr":1148,"May":5278,"Jun":68503,"Jul":40408,"Aug":14280,"Sep":21692,"Oct":19600,"Nov":15839,"Dec":9485,"Jan":4214,"Feb":649,"Mar":25143},
    "One Time Charge":{"Apr":45325,"May":12530,"Jun":4750,"Jul":79984,"Aug":76574,"Sep":149914,"Oct":64441,"Nov":199349,"Dec":71035,"Jan":24050,"Feb":31250,"Mar":11850},
    "Children Education": {"Apr":981577,"May":68350,"Jun":550856,"Jul":101149,"Aug":124383,"Sep":54100,"Oct":38700,"Nov":514160,"Dec":0,"Jan":29700,"Feb":45500,"Mar":524550},
    "Kalpataru Maintenance": {"Apr":35997,"May":35933,"Jun":35879,"Jul":-7041,"Aug":197231,"Sep":36587,"Oct":35879,"Nov":71168,"Dec":0,"Jan":35879,"Feb":35879,"Mar":35933},
    "Charity":        {"Apr":20000,"May":0,"Jun":0,"Jul":167800,"Aug":105725,"Sep":24048,"Oct":300000,"Nov":115000,"Dec":150000,"Jan":625100,"Feb":0,"Mar":18000},
    "Uspaar":         {"Apr":133783,"May":140978,"Jun":214372,"Jul":97910,"Aug":204600,"Sep":72828,"Oct":38984,"Nov":170003,"Dec":0,"Jan":103660,"Feb":40350,"Mar":86325},
    "Insurance":      {"Apr":118000,"May":0,"Jun":0,"Jul":0,"Aug":0,"Sep":0,"Oct":102820,"Nov":0,"Dec":142800,"Jan":0,"Feb":0,"Mar":0},
    "Home Loan":      {"Apr":903259,"May":1146721,"Jun":916551,"Jul":1201557,"Aug":554547,"Sep":992227,"Oct":963511,"Nov":1007410,"Dec":1777170,"Jan":1389117,"Feb":928825,"Mar":1285719},
    "Tax":            {"Apr":0,"May":0,"Jun":0,"Jul":13519347,"Aug":0,"Sep":3630000,"Oct":61499379,"Nov":0,"Dec":0,"Jan":0,"Feb":3848950,"Mar":0},
}
MIS_FY25_MONTHLY = {
    "Misc":           {"Apr":6088,"May":8858,"Jun":20244,"Jul":32654,"Aug":20773,"Sep":24987,"Oct":2704,"Nov":5462,"Dec":20402,"Jan":23428,"Feb":6027,"Mar":4321},
    "Clothes":        {"Apr":106470,"May":30095,"Jun":30780,"Jul":0,"Aug":9899,"Sep":8590,"Oct":0,"Nov":20798,"Dec":48088,"Jan":2490,"Feb":22291,"Mar":14990},
    "Gifts":          {"Apr":3599,"May":0,"Jun":14285,"Jul":15857,"Aug":5847,"Sep":12462,"Oct":25000,"Nov":44705,"Dec":11270,"Jan":990,"Feb":0,"Mar":0},
    "Cash":           {"Apr":0,"May":55000,"Jun":10000,"Jul":20000,"Aug":30000,"Sep":40000,"Oct":30000,"Nov":60000,"Dec":30000,"Jan":345500,"Feb":20500,"Mar":20000},
    "Maintenance Expense": {"Apr":0,"May":5500,"Jun":0,"Jul":0,"Aug":0,"Sep":1902,"Oct":13383,"Nov":4253,"Dec":3124,"Jan":2900,"Feb":0,"Mar":449},
    "Malhar":         {"Apr":38269,"May":53199,"Jun":91025,"Jul":98388,"Aug":35080,"Sep":74930,"Oct":22000,"Nov":41900,"Dec":41836,"Jan":56530,"Feb":42330,"Mar":83090},
    "Home office":    {"Apr":30702,"May":13150,"Jun":7000,"Jul":8500,"Aug":9200,"Sep":26914,"Oct":13500,"Nov":13359,"Dec":10000,"Jan":1200,"Feb":1000,"Mar":51000},
    "Electricity & Gas": {"Apr":18233,"May":50902,"Jun":40791,"Jul":31210,"Aug":27031,"Sep":37216,"Oct":32991,"Nov":34039,"Dec":18101,"Jan":19564,"Feb":22573,"Mar":28734},
    "Alcohol":        {"Apr":5500,"May":1500,"Jun":0,"Jul":32235,"Aug":0,"Sep":0,"Oct":17089,"Nov":0,"Dec":30235,"Jan":22691,"Feb":0,"Mar":0},
    "Medical":        {"Apr":25104,"May":17747,"Jun":16695,"Jul":15049,"Aug":34811,"Sep":28366,"Oct":7032,"Nov":6215,"Dec":95127,"Jan":61258,"Feb":260944,"Mar":65715},
    "Holiday":        {"Apr":331046,"May":388673,"Jun":527589,"Jul":484268,"Aug":82855,"Sep":221069,"Oct":92797,"Nov":0,"Dec":42609,"Jan":0,"Feb":37368,"Mar":66627},
    "Groceries":      {"Apr":114509,"May":117604,"Jun":108667,"Jul":98176,"Aug":108702,"Sep":81999,"Oct":113779,"Nov":80055,"Dec":85734,"Jan":95234,"Feb":90302,"Mar":67905},
    "Eating Out":     {"Apr":25839,"May":165938,"Jun":34440,"Jul":119314,"Aug":11086,"Sep":44803,"Oct":5921,"Nov":115949,"Dec":17293,"Jan":73701,"Feb":5160,"Mar":9266},
    "Amma":           {"Apr":17294,"May":2434,"Jun":7480,"Jul":13736,"Aug":11310,"Sep":20583,"Oct":15262,"Nov":4273,"Dec":6909,"Jan":1939,"Feb":45101,"Mar":588},
    "Ketki":          {"Apr":74654,"May":60324,"Jun":118994,"Jul":66982,"Aug":162796,"Sep":92541,"Oct":117532,"Nov":84241,"Dec":134245,"Jan":63858,"Feb":73205,"Mar":101126},
    "Wellness":       {"Apr":4100,"May":1750,"Jun":19762,"Jul":23432,"Aug":1800,"Sep":0,"Oct":15878,"Nov":26884,"Dec":33216,"Jan":17370,"Feb":36550,"Mar":81309},
    "One Time Charge":{"Apr":2880,"May":7500,"Jun":2850,"Jul":2950,"Aug":7660,"Sep":6800,"Oct":26980,"Nov":20638,"Dec":13460,"Jan":4100,"Feb":13990,"Mar":20650},
    "Entertainment":  {"Apr":67241,"May":27665,"Jun":7098,"Jul":33295,"Aug":189,"Sep":13232,"Oct":1779,"Nov":7175,"Dec":7009,"Jan":499,"Feb":13645,"Mar":6690},
    "Staff Salary":   {"Apr":236310,"May":197770,"Jun":175870,"Jul":176370,"Aug":220870,"Sep":209570,"Oct":244730,"Nov":230870,"Dec":385283,"Jan":303020,"Feb":19020,"Mar":175720},
    "Financial Expense / OD Interest": {"Apr":38529,"May":20146,"Jun":0,"Jul":0,"Aug":0,"Sep":74667,"Oct":0,"Nov":2058164,"Dec":0,"Jan":234937,"Feb":29529,"Mar":0},
    "Children Education": {"Apr":856197,"May":188432,"Jun":22700,"Jul":78855,"Aug":110550,"Sep":127800,"Oct":56000,"Nov":956923,"Dec":28700,"Jan":28000,"Feb":40400,"Mar":33350},
    "Kalpataru Maintenance": {"Apr":29421,"May":29421,"Jun":29421,"Jul":29421,"Aug":29421,"Sep":29421,"Oct":158721,"Nov":35879,"Dec":35879,"Jan":31902,"Feb":35879,"Mar":35879},
    "Charity":        {"Apr":583500,"May":300000,"Jun":13650,"Jul":818250,"Aug":550000,"Sep":118000,"Oct":245247,"Nov":36898,"Dec":118882,"Jan":0,"Feb":100000,"Mar":415585},
    "Uspaar":         {"Apr":17580,"May":16360,"Jun":157855,"Jul":30000,"Aug":87410,"Sep":90000,"Oct":108790,"Nov":80000,"Dec":111799,"Jan":115080,"Feb":143200,"Mar":151260},
    "Insurance":      {"Apr":0,"May":0,"Jun":0,"Jul":0,"Aug":0,"Sep":0,"Oct":0,"Nov":119558,"Dec":168504,"Jan":0,"Feb":0,"Mar":0},
    "Home Loan":      {"Apr":587781,"May":417556,"Jun":352171,"Jul":250748,"Aug":283686,"Sep":175803,"Oct":316059,"Nov":333238,"Dec":640435,"Jan":210055,"Feb":361851,"Mar":429844},
    "Tax":            {"Apr":1620,"May":0,"Jun":23038641,"Jul":385000,"Aug":0,"Sep":0,"Oct":83751,"Nov":0,"Dec":0,"Jan":0,"Feb":0,"Mar":0},
}
MIS_FY24_MONTHLY = {
    "Misc":           {"Apr":10612,"May":7412,"Jun":5604,"Jul":3779,"Aug":3872,"Sep":1709,"Oct":690,"Nov":0,"Dec":0,"Jan":23492,"Feb":5290,"Mar":7080},
    "Clothes":        {"Apr":8912,"May":41188,"Jun":1999,"Jul":13493,"Aug":0,"Sep":0,"Oct":6689,"Nov":4998,"Dec":1778,"Jan":0,"Feb":0,"Mar":41753},
    "Gifts":          {"Apr":798,"May":6017,"Jun":1499,"Jul":95214,"Aug":0,"Sep":28468,"Oct":0,"Nov":81000,"Dec":2847,"Jan":3448,"Feb":0,"Mar":0},
    "Cash":           {"Apr":20000,"May":0,"Jun":118169,"Jul":6000,"Aug":13500,"Sep":25046,"Oct":5000,"Nov":0,"Dec":10000,"Jan":0,"Feb":10000,"Mar":0},
    "Maintenance Expense": {"Apr":121359,"May":178721,"Jun":222206,"Jul":193185,"Aug":241423,"Sep":186081,"Oct":251831,"Nov":351996,"Dec":177411,"Jan":219494,"Feb":99367,"Mar":182662},
    "Malhar":         {"Apr":24679,"May":25616,"Jun":41806,"Jul":89882,"Aug":37696,"Sep":27621,"Oct":40210,"Nov":0,"Dec":22173,"Jan":26590,"Feb":27870,"Mar":49750},
    "Kashid":         {"Apr":22038,"May":0,"Jun":0,"Jul":47220,"Aug":18972,"Sep":8000,"Oct":9600,"Nov":18800,"Dec":20670,"Jan":58880,"Feb":0,"Mar":0},
    "Rent":           {"Apr":0,"May":141487,"Jun":0,"Jul":327673,"Aug":29000,"Sep":0,"Oct":0,"Nov":0,"Dec":0,"Jan":0,"Feb":0,"Mar":0},
    "Personal Loans": {"Apr":0,"May":0,"Jun":0,"Jul":0,"Aug":0,"Sep":192436,"Oct":74250,"Nov":0,"Dec":117000,"Jan":2067919,"Feb":0,"Mar":0},
    "Home office":    {"Apr":13782,"May":53133,"Jun":12792,"Jul":30412,"Aug":11812,"Sep":31167,"Oct":6206,"Nov":33010,"Dec":13500,"Jan":7000,"Feb":6000,"Mar":5000},
    "Electricity & Gas": {"Apr":15383,"May":28199,"Jun":20582,"Jul":14461,"Aug":12563,"Sep":18634,"Oct":14635,"Nov":23979,"Dec":14912,"Jan":14781,"Feb":12491,"Mar":16975},
    "Alcohol":        {"Apr":2800,"May":0,"Jun":19700,"Jul":19598,"Aug":18314,"Sep":0,"Oct":0,"Nov":0,"Dec":11610,"Jan":0,"Feb":0,"Mar":0},
    "Medical":        {"Apr":0,"May":27400,"Jun":4000,"Jul":34500,"Aug":15297,"Sep":90264,"Oct":11746,"Nov":2313,"Dec":20138,"Jan":0,"Feb":1500,"Mar":74140},
    "Holiday":        {"Apr":81805,"May":299243,"Jun":232239,"Jul":46421,"Aug":9250,"Sep":6800,"Oct":0,"Nov":49154,"Dec":164904,"Jan":27688,"Feb":296465,"Mar":73302},
    "Groceries":      {"Apr":114558,"May":86085,"Jun":96964,"Jul":157379,"Aug":67436,"Sep":113728,"Oct":89000,"Nov":93120,"Dec":98000,"Jan":111205,"Feb":70000,"Mar":85000},
    "Eating Out":     {"Apr":22416,"May":20065,"Jun":166419,"Jul":30346,"Aug":24036,"Sep":34605,"Oct":6549,"Nov":88444,"Dec":1980,"Jan":6779,"Feb":60212,"Mar":16879},
    "Amma":           {"Apr":0,"May":0,"Jun":0,"Jul":0,"Aug":0,"Sep":0,"Oct":21000,"Nov":37486,"Dec":3198,"Jan":9826,"Feb":10999,"Mar":12901},
    "Ketki":          {"Apr":0,"May":0,"Jun":0,"Jul":0,"Aug":0,"Sep":0,"Oct":0,"Nov":0,"Dec":145000,"Jan":1500,"Feb":0,"Mar":0},
    "Wellness":       {"Apr":0,"May":0,"Jun":0,"Jul":0,"Aug":0,"Sep":0,"Oct":0,"Nov":0,"Dec":0,"Jan":0,"Feb":0,"Mar":0},
    "One Time Charge":{"Apr":0,"May":0,"Jun":0,"Jul":0,"Aug":0,"Sep":0,"Oct":0,"Nov":0,"Dec":0,"Jan":0,"Feb":0,"Mar":0},
    "Entertainment":  {"Apr":0,"May":0,"Jun":0,"Jul":0,"Aug":0,"Sep":0,"Oct":0,"Nov":0,"Dec":0,"Jan":0,"Feb":0,"Mar":0},
    "Staff Salary":   {"Apr":222000,"May":94500,"Jun":125000,"Jul":142000,"Aug":128000,"Sep":152000,"Oct":153100,"Nov":149000,"Dec":133000,"Jan":127000,"Feb":151000,"Mar":117100},
    "Financial Expense / OD Interest": {"Apr":9000,"May":0,"Jun":0,"Jul":0,"Aug":50000,"Sep":0,"Oct":0,"Nov":0,"Dec":0,"Jan":1525,"Feb":0,"Mar":0},
    "Children Education": {"Apr":0,"May":756000,"Jun":21990,"Jul":196068,"Aug":73518,"Sep":12780,"Oct":2800,"Nov":7200,"Dec":770655,"Jan":209300,"Feb":5349,"Mar":16099},
    "Kalpataru Maintenance": {"Apr":0,"May":0,"Jun":0,"Jul":0,"Aug":0,"Sep":26256,"Oct":28193,"Nov":28225,"Dec":28349,"Jan":28225,"Feb":28225,"Mar":151067},
    "Charity":        {"Apr":0,"May":50000,"Jun":40000,"Jul":21040,"Aug":0,"Sep":0,"Oct":0,"Nov":0,"Dec":50000,"Jan":100000,"Feb":100000,"Mar":0},
    "Uspaar":         {"Apr":0,"May":0,"Jun":0,"Jul":0,"Aug":0,"Sep":0,"Oct":0,"Nov":0,"Dec":0,"Jan":0,"Feb":0,"Mar":0},
    "Insurance":      {"Apr":0,"May":0,"Jun":0,"Jul":342056,"Aug":0,"Sep":0,"Oct":148374,"Nov":22189,"Dec":168504,"Jan":0,"Feb":0,"Mar":138429},
    "Home Loan":      {"Apr":598843,"May":430433,"Jun":474414,"Jul":426159,"Aug":136367,"Sep":323197,"Oct":0,"Nov":593591,"Dec":411966,"Jan":590471,"Feb":600440,"Mar":575024},
    "Club":           {"Apr":0,"May":0,"Jun":0,"Jul":0,"Aug":3405002,"Sep":0,"Oct":0,"Nov":0,"Dec":0,"Jan":0,"Feb":0,"Mar":0},
    "Tax":            {"Apr":1620,"May":2112347,"Jun":4598,"Jul":7103283,"Aug":0,"Sep":0,"Oct":61394,"Nov":-87840,"Dec":100,"Jan":800,"Feb":0,"Mar":0},
}


@app.route("/api/mis", methods=["GET"])
@login_required
def api_mis():
    """Return MIS data grouped by super-category.
    Query param: period = month | quarter | ytd  (default: month)

    FY26 column always shows the full-year FY26 actual (from fy26_actuals.json).
    FY27 Budget:
      - month/quarter: scaled to period months
      - ytd: full-year annual budget
    FY27 Actual: what's been approved in the period (month/quarter/ytd).
    """
    period = request.args.get("period", "month")
    fy     = request.args.get("fy", "FY27")   # FY26 | FY27

    with open(os.path.join(CONFIG_DIR, "budget_fy27.json")) as f:
        _bfile = json.load(f)
    budget_annual = _bfile["annual"]   # FY27 annual by ACC26 heading (Blueprint)
    budget_monthly_app = _bfile["monthly"]  # monthly by app category (approval engine)

    # ── FY monthly data — from module-level constants ───────────────────────────
    FY26_MONTHLY = MIS_FY26_MONTHLY
    FY25_MONTHLY = MIS_FY25_MONTHLY
    FY24_MONTHLY = MIS_FY24_MONTHLY

    # Calendar month number → FY month abbreviation
    CAL_TO_FY_MON = {4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec",1:"Jan",2:"Feb",3:"Mar"}

    log = _load_json(APPROVAL_LOG)

    # ── Period months ─────────────────────────────────────────────────────────
    now = datetime.now()
    cal_month = now.month
    fy_start_year = now.year if cal_month >= 4 else now.year - 1

    def _period_months(p):
        if p == "month":
            return [now.strftime("%Y-%m")]
        elif p == "quarter":
            fy_month = (cal_month - 4) % 12 + 1
            q_start_fy = ((fy_month - 1) // 3) * 3 + 1
            months = []
            for offset in range(3):
                cm_idx = (q_start_fy - 1 + offset) % 12
                cal = (cm_idx + 4 - 1) % 12 + 1
                yr = fy_start_year if cal >= 4 else fy_start_year + 1
                months.append(f"{yr}-{cal:02d}")
            return months
        else:  # ytd
            months, cm, yr = [], 4, fy_start_year
            while True:
                months.append(f"{yr}-{cm:02d}")
                if yr == now.year and cm == cal_month:
                    break
                cm += 1
                if cm == 13:
                    cm, yr = 1, yr + 1
            return months

    period_months = _period_months(period)
    n_months = len(period_months)

    # FY month names covered by this period (for FY26 lookup)
    fy_mon_names = [CAL_TO_FY_MON[int(m.split("-")[1])] for m in period_months]

    # ── Heading → super-category ──────────────────────────────────────────────
    HEADING_SUPER = {
        "Groceries":"Household","Staff Salary":"Household","Electricity & Gas":"Household",
        "Misc":"Household","Cash":"Household",
        "Alcohol":"Personal","Wellness":"Personal","Clothes":"Personal",
        "Gifts":"Family","Medical":"Family","Amma":"Family","Ketki":"Family",
        "Children Education":"Family",
        "Charity":"Giving","Uspaar":"Giving",
        "Holiday":"Lifestyle","Eating Out":"Lifestyle","Entertainment":"Lifestyle","Club":"Lifestyle",
        "Malhar":"Property","Maintenance Expense":"Property","Home office":"Property",
        "One Time Charge":"Property","Kalpataru Maintenance":"Property",
        "Kashid":"Property","Rent":"Property",
        "Financial Expense / OD Interest":"Financial","Financial Expense":"Financial",
        "Insurance":"Financial","Home Loan":"Financial","Tax":"Financial",
        "Personal Loans":"Financial",
    }
    SUPER_ORDER = ["Household","Personal","Family","Giving","Lifestyle","Property","Financial"]

    # App category → ACC26 heading (for FY27 actual from approval log)
    APP_TO_HEADING = {
        "groceries":"Groceries","staff":"Staff Salary","utilities":"Electricity & Gas",
        "miscellaneous":"Misc","personal_care":"Wellness","clothing":"Clothes",
        "gifts":"Gifts","medical":"Medical","education":"Children Education",
        "dining":"Eating Out","entertainment":"Entertainment","transport":"Holiday",
        "maintenance":"Maintenance Expense","home_repair":"One Time Charge",
    }

    # ── FY26 for period: sum actual monthly data for the matched FY months ────
    def _fy26_period(heading):
        monthly = FY26_MONTHLY.get(heading, {})
        return max(0, sum(monthly.get(m, 0) for m in fy_mon_names))

    def _fy26_full_year(heading):
        monthly = FY26_MONTHLY.get(heading, {})
        return max(0, sum(monthly.values()))

    # ── FY27 budget scaled to period ─────────────────────────────────────────
    def _budget(heading):
        annual = budget_annual.get(heading, 0)
        return annual if period == "ytd" else round(annual / 12 * n_months)

    # ── Heading normalisation: non-canonical → canonical ─────────────────────
    # Headings that are investments/transfers — exclude from budget tracker
    _INVESTMENT_HEADINGS = {"Art","Investment","Short Term Advance","Credit Card Loan"}
    # Canonical set (only these appear as rows)
    _CANONICAL = set(budget_annual.keys()) | set(FY26_MONTHLY.keys())
    # Case-insensitive + common variant → canonical
    _HEADING_NORM = {h.lower(): h for h in _CANONICAL}
    _HEADING_NORM.update({
        "malhar renovation": "Malhar",
        "interbank": "Misc",
        "offical": "Misc", "official": "Misc",
        "eating out": "Eating Out",
        "children education": "Children Education",
        "staff salary": "Staff Salary",
        "electricity & gas": "Electricity & Gas",
        "financial expense": "Financial Expense / OD Interest",
        "education": "Children Education",
        "gift": "Gifts",
        "dining": "Eating Out",
        "unknown": "Misc",
        "salary": "Staff Salary",
        "electric": "Electricity & Gas",
        "maintenance": "Maintenance Expense",
        "grocery": "Groceries",
        "none": "None",
    })

    def _norm_heading(h):
        if not h: return "Misc"
        if h in _INVESTMENT_HEADINGS: return None          # exclude
        if h in _CANONICAL: return h
        return _HEADING_NORM.get(h.lower().strip(), "Misc")  # roll unknown → Misc

    # ── FY24 mode: actuals come from hardcoded FY24_MONTHLY; no live ledger needed ──
    if fy == "FY24":
        all_headings = set(budget_annual.keys()) | set(FY24_MONTHLY.keys())
        by_super: dict = {s: [] for s in SUPER_ORDER}
        for heading in sorted(all_headings):
            super_cat = HEADING_SUPER.get(heading, "Household")
            monthly = FY24_MONTHLY.get(heading, {})
            fy24_total = round(max(0, sum(monthly.values())))
            budget = round(budget_annual.get(heading, 0))
            row = {
                "category": heading,
                "fy26_actual": fy24_total,
                "fy26_full_year": fy24_total,
                "fy27_budget": budget,
                "fy27_actual": fy24_total,
                "pct": round(fy24_total / budget * 100) if budget else 0,
            }
            by_super.setdefault(super_cat, []).append(row)

        groups = []
        grand = {"fy26": 0, "fy26_full_year": 0, "budget": 0, "actual": 0}
        for super_cat in SUPER_ORDER:
            rows = by_super.get(super_cat, [])
            if not rows:
                continue
            sub = {
                "fy26":          sum(r["fy26_actual"] for r in rows),
                "fy26_full_year": sum(r["fy26_full_year"] for r in rows),
                "budget":        sum(r["fy27_budget"] for r in rows),
                "actual":        sum(r["fy27_actual"] for r in rows),
            }
            sub["pct"] = round(sub["actual"] / sub["budget"] * 100) if sub["budget"] else 0
            grand["fy26"]          += sub["fy26"]
            grand["fy26_full_year"] += sub["fy26_full_year"]
            grand["budget"]        += sub["budget"]
            grand["actual"]        += sub["actual"]
            groups.append({"super_category": super_cat, "rows": rows, "subtotal": sub})
        grand["pct"] = round(grand["actual"] / grand["budget"] * 100) if grand["budget"] else 0
        return jsonify({"period": "ytd", "fy": "FY24",
                        "period_months": list(FY24_MONTHLY.get("Groceries", {}).keys()),
                        "groups": groups, "grand": grand})

    # ── FY26 mode: actuals come from hardcoded FY26_MONTHLY; no live ledger needed ──
    elif fy == "FY26":
        all_headings = _CANONICAL
        by_super: dict = {s: [] for s in SUPER_ORDER}
        for heading in sorted(all_headings):
            super_cat = HEADING_SUPER.get(heading, "Household")
            fy26_total = round(_fy26_full_year(heading))
            fy26_period_val = round(_fy26_period(heading))
            budget = round(budget_annual.get(heading, 0))
            row = {
                "category": heading,
                "fy26_actual": fy26_period_val,
                "fy26_full_year": fy26_total,
                "fy27_budget": budget,
                "fy27_actual": fy26_total,   # "actual" col = FY26 full-year spend
                "pct": round(fy26_total / budget * 100) if budget else 0,
            }
            by_super.setdefault(super_cat, []).append(row)

        groups = []
        grand = {"fy26": 0, "fy26_full_year": 0, "budget": 0, "actual": 0}
        for super_cat in SUPER_ORDER:
            rows = by_super.get(super_cat, [])
            if not rows:
                continue
            sub = {
                "fy26":          sum(r["fy26_actual"] for r in rows),
                "fy26_full_year": sum(r["fy26_full_year"] for r in rows),
                "budget":        sum(r["fy27_budget"] for r in rows),
                "actual":        sum(r["fy27_actual"] for r in rows),
            }
            sub["pct"] = round(sub["actual"] / sub["budget"] * 100) if sub["budget"] else 0
            grand["fy26"]          += sub["fy26"]
            grand["fy26_full_year"] += sub["fy26_full_year"]
            grand["budget"]        += sub["budget"]
            grand["actual"]        += sub["actual"]
            groups.append({"super_category": super_cat, "rows": rows, "subtotal": sub})
        grand["pct"] = round(grand["actual"] / grand["budget"] * 100) if grand["budget"] else 0
        return jsonify({"period": "ytd", "fy": "FY26",
                        "period_months": list(FY26_MONTHLY.get("Groceries", {}).keys()),
                        "groups": groups, "grand": grand})

    # ── FY25 mode: actuals come from hardcoded FY25_MONTHLY; no live ledger needed ──
    elif fy == "FY25":
        all_headings = set(budget_annual.keys()) | set(FY25_MONTHLY.keys())
        by_super: dict = {s: [] for s in SUPER_ORDER}
        for heading in sorted(all_headings):
            super_cat = HEADING_SUPER.get(heading, "Household")
            monthly = FY25_MONTHLY.get(heading, {})
            fy25_total = round(max(0, sum(monthly.values())))
            budget = round(budget_annual.get(heading, 0))
            row = {
                "category": heading,
                "fy26_actual": fy25_total,
                "fy26_full_year": fy25_total,
                "fy27_budget": budget,
                "fy27_actual": fy25_total,
                "pct": round(fy25_total / budget * 100) if budget else 0,
            }
            by_super.setdefault(super_cat, []).append(row)

        groups = []
        grand = {"fy26": 0, "fy26_full_year": 0, "budget": 0, "actual": 0}
        for super_cat in SUPER_ORDER:
            rows = by_super.get(super_cat, [])
            if not rows:
                continue
            sub = {
                "fy26":          sum(r["fy26_actual"] for r in rows),
                "fy26_full_year": sum(r["fy26_full_year"] for r in rows),
                "budget":        sum(r["fy27_budget"] for r in rows),
                "actual":        sum(r["fy27_actual"] for r in rows),
            }
            sub["pct"] = round(sub["actual"] / sub["budget"] * 100) if sub["budget"] else 0
            grand["fy26"]          += sub["fy26"]
            grand["fy26_full_year"] += sub["fy26_full_year"]
            grand["budget"]        += sub["budget"]
            grand["actual"]        += sub["actual"]
            groups.append({"super_category": super_cat, "rows": rows, "subtotal": sub})
        grand["pct"] = round(grand["actual"] / grand["budget"] * 100) if grand["budget"] else 0
        return jsonify({"period": "ytd", "fy": "FY25",
                        "period_months": list(FY25_MONTHLY.get("Groceries", {}).keys()),
                        "groups": groups, "grand": grand})

    # ── monthly_full: one column per FY27 month + FY26 avg/mo ────────────────
    elif period == "monthly_full":
        from src.master_ledger import _parse_date as _ml_parse_date
        FY27_MONTHS = ["Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb","Mar"]
        MON_TO_YM = {
            "Apr":"2026-04","May":"2026-05","Jun":"2026-06","Jul":"2026-07",
            "Aug":"2026-08","Sep":"2026-09","Oct":"2026-10","Nov":"2026-11",
            "Dec":"2026-12","Jan":"2027-01","Feb":"2027-02","Mar":"2027-03",
        }
        _fy27_start = datetime(2026, 4, 1)
        fy27_by_ym: dict = {}
        for txn in load_ledger():
            if (txn.get("type") or "").lower() != "expense":
                continue
            if txn.get("uncertain"):
                continue
            dt = _ml_parse_date(txn.get("date", ""))
            if not dt or dt < _fy27_start:
                continue
            ym = dt.strftime("%Y-%m")
            if ym not in MON_TO_YM.values():
                continue
            heading = _norm_heading(txn.get("heading", "") or "")
            if heading is None:
                continue
            net = float(txn.get("debit", 0) or 0) - float(txn.get("credit", 0) or 0)
            fy27_by_ym.setdefault(heading, {})
            fy27_by_ym[heading][ym] = fy27_by_ym[heading].get(ym, 0) + net


        by_super: dict = {s: [] for s in SUPER_ORDER}
        for heading in sorted(_CANONICAL):
            super_cat = HEADING_SUPER.get(heading, "Household")
            fy26_avg = round(_fy26_full_year(heading) / 12)
            monthly = {mon: round(fy27_by_ym.get(heading, {}).get(MON_TO_YM[mon], 0)) for mon in FY27_MONTHS}
            by_super.setdefault(super_cat, []).append({"category": heading, "fy26_avg": fy26_avg, "monthly": monthly})

        groups = []
        grand_monthly = {mon: 0 for mon in FY27_MONTHS}
        grand_avg = 0
        for super_cat in SUPER_ORDER:
            rows = by_super.get(super_cat, [])
            if not rows:
                continue
            sub_monthly = {mon: sum(r["monthly"].get(mon, 0) for r in rows) for mon in FY27_MONTHS}
            sub_avg = sum(r["fy26_avg"] for r in rows)
            for mon in FY27_MONTHS:
                grand_monthly[mon] += sub_monthly[mon]
            grand_avg += sub_avg
            groups.append({"super_category": super_cat, "rows": rows,
                           "subtotal": {"fy26_avg": sub_avg, "monthly": sub_monthly}})
        grand = {"fy26_avg": grand_avg, "monthly": grand_monthly}
        return jsonify({"period": "monthly_full", "months": FY27_MONTHS, "groups": groups, "grand": grand})

    # ── quarterly_full: Q1-Q4 FY26 + Q1-Q4 FY27 ─────────────────────────────
    elif period == "quarterly_full":
        from src.master_ledger import _parse_date as _ml_parse_date
        FY26_Q = {
            "Q1": ["Apr","May","Jun"], "Q2": ["Jul","Aug","Sep"],
            "Q3": ["Oct","Nov","Dec"], "Q4": ["Jan","Feb","Mar"],
        }
        FY27_Q_YMS = {
            "Q1": ["2026-04","2026-05","2026-06"],
            "Q2": ["2026-07","2026-08","2026-09"],
            "Q3": ["2026-10","2026-11","2026-12"],
            "Q4": ["2027-01","2027-02","2027-03"],
        }
        _fy27_start = datetime(2026, 4, 1)
        fy27_by_ym: dict = {}
        for txn in load_ledger():
            if (txn.get("type") or "").lower() != "expense":
                continue
            if txn.get("uncertain"):
                continue
            dt = _ml_parse_date(txn.get("date", ""))
            if not dt or dt < _fy27_start:
                continue
            heading = _norm_heading(txn.get("heading", "") or "")
            if heading is None:
                continue
            net = float(txn.get("debit", 0) or 0) - float(txn.get("credit", 0) or 0)
            ym = dt.strftime("%Y-%m")
            fy27_by_ym.setdefault(heading, {})
            fy27_by_ym[heading][ym] = fy27_by_ym[heading].get(ym, 0) + net


        def _fy26_q(heading, mons):
            return round(sum(FY26_MONTHLY.get(heading, {}).get(m, 0) for m in mons))

        def _fy27_q(heading, yms):
            return round(sum(fy27_by_ym.get(heading, {}).get(ym, 0) for ym in yms))

        QCOLS = ["fy26_q1","fy26_q2","fy26_q3","fy26_q4","fy27_q1","fy27_q2","fy27_q3","fy27_q4"]
        by_super: dict = {s: [] for s in SUPER_ORDER}
        for heading in sorted(_CANONICAL):
            super_cat = HEADING_SUPER.get(heading, "Household")
            row = {
                "category": heading,
                "fy26_q1": _fy26_q(heading, FY26_Q["Q1"]),
                "fy26_q2": _fy26_q(heading, FY26_Q["Q2"]),
                "fy26_q3": _fy26_q(heading, FY26_Q["Q3"]),
                "fy26_q4": _fy26_q(heading, FY26_Q["Q4"]),
                "fy27_q1": _fy27_q(heading, FY27_Q_YMS["Q1"]),
                "fy27_q2": _fy27_q(heading, FY27_Q_YMS["Q2"]),
                "fy27_q3": 0, "fy27_q4": 0,
            }
            by_super.setdefault(super_cat, []).append(row)

        groups = []
        grand = {k: 0 for k in QCOLS}
        for super_cat in SUPER_ORDER:
            rows = by_super.get(super_cat, [])
            if not rows:
                continue
            sub = {k: sum(r[k] for r in rows) for k in QCOLS}
            for k in QCOLS:
                grand[k] += sub[k]
            groups.append({"super_category": super_cat, "rows": rows, "subtotal": sub})
        return jsonify({"period": "quarterly_full", "groups": groups, "grand": grand})

    # ── FY27 actual from master ledger ────────────────────────────────────────
    from src.master_ledger import _parse_date as _ml_parse_date
    fy27_actual: dict = {}
    _fy27_start = datetime(2026, 4, 1)
    for txn in load_ledger():
        if (txn.get("type") or "").lower() != "expense":
            continue
        if txn.get("uncertain"):
            continue
        dt = _ml_parse_date(txn.get("date", ""))
        if not dt or dt < _fy27_start:
            continue
        ym = dt.strftime("%Y-%m")
        if ym not in period_months:
            continue
        raw_heading = txn.get("heading", "") or ""
        heading = _norm_heading(raw_heading)
        if heading is None:
            continue   # investment heading — skip
        net = float(txn.get("debit", 0) or 0) - float(txn.get("credit", 0) or 0)
        fy27_actual[heading] = fy27_actual.get(heading, 0) + net


    # ── Build grouped rows — only canonical headings ──────────────────────────
    all_headings = _CANONICAL
    by_super: dict = {s: [] for s in SUPER_ORDER}

    for heading in sorted(all_headings):
        super_cat = HEADING_SUPER.get(heading, "Household")
        fy26 = round(_fy26_period(heading))
        budget = _budget(heading)
        actual = round(fy27_actual.get(heading, 0))
        pct = round(actual / budget * 100) if budget else 0
        row = {
            "category": heading,
            "fy26_actual": fy26,
            "fy27_budget": budget,
            "fy27_actual": actual,
            "pct": pct,
        }
        if period == "ytd":
            row["fy26_full_year"] = round(_fy26_full_year(heading))
            row["fy25_actual"] = round(sum(FY25_MONTHLY.get(heading, {}).values()))
            row["fy24_actual"] = round(sum(FY24_MONTHLY.get(heading, {}).values()))
        by_super.setdefault(super_cat, []).append(row)

    groups = []
    grand = {"fy26": 0, "fy26_full_year": 0, "fy25_actual": 0, "fy24_actual": 0, "budget": 0, "actual": 0}
    for super_cat in SUPER_ORDER:
        rows = by_super.get(super_cat, [])
        if not rows:
            continue
        sub = {
            "fy26":   sum(r["fy26_actual"] for r in rows),
            "budget": sum(r["fy27_budget"] for r in rows),
            "actual": sum(r["fy27_actual"] for r in rows),
        }
        if period == "ytd":
            sub["fy26_full_year"] = sum(r.get("fy26_full_year", 0) for r in rows)
            sub["fy25_actual"] = sum(r.get("fy25_actual", 0) for r in rows)
            sub["fy24_actual"] = sum(r.get("fy24_actual", 0) for r in rows)
            grand["fy26_full_year"] += sub["fy26_full_year"]
            grand["fy25_actual"] += sub["fy25_actual"]
            grand["fy24_actual"] += sub["fy24_actual"]
        sub["pct"] = round(sub["actual"] / sub["budget"] * 100) if sub["budget"] else 0
        grand["fy26"]   += sub["fy26"]
        grand["budget"] += sub["budget"]
        grand["actual"] += sub["actual"]
        groups.append({"super_category": super_cat, "rows": rows, "subtotal": sub})

    grand["pct"] = round(grand["actual"] / grand["budget"] * 100) if grand["budget"] else 0

    return jsonify({
        "period": period,
        "period_months": period_months,
        "groups": groups,
        "grand": grand,
    })


@app.route("/api/financial-statements", methods=["GET"])
@login_required
def api_financial_statements():
    """Return P&L, Balance Sheet, and Asset detail for FY24/FY25/FY26."""

    HEADING_SUPER = {
        "Groceries":"Household","Staff Salary":"Household","Electricity & Gas":"Household",
        "Misc":"Household","Cash":"Household",
        "Alcohol":"Personal","Wellness":"Personal","Clothes":"Personal",
        "Gifts":"Family","Medical":"Family","Amma":"Family","Ketki":"Family",
        "Children Education":"Family",
        "Charity":"Giving","Uspaar":"Giving",
        "Holiday":"Lifestyle","Eating Out":"Lifestyle","Entertainment":"Lifestyle","Club":"Lifestyle",
        "Malhar":"Property","Maintenance Expense":"Property","Home office":"Property",
        "One Time Charge":"Property","Kalpataru Maintenance":"Property",
        "Kashid":"Property","Rent":"Property",
        "Financial Expense / OD Interest":"Financial","Financial Expense":"Financial",
        "Insurance":"Financial","Home Loan":"Financial","Tax":"Financial",
        "Personal Loans":"Financial",
    }
    SUPER_ORDER = ["Household","Personal","Family","Giving","Lifestyle","Property","Financial"]
    NON_FIN = ["Household","Personal","Family","Giving","Lifestyle","Property"]

    # ── Expenses from MIS monthly dicts (same source as budget tracker) ────────
    FIN_HEADINGS = ["Home Loan", "Insurance", "Financial Expense / OD Interest",
                    "Financial Expense", "Tax", "Personal Loans"]

    def mis_expenses(monthly_dict):
        """Return super-category totals, per-heading detail, AND individual Financial heading totals."""
        super_totals = {cat: 0 for cat in SUPER_ORDER}
        heading_totals = {}
        fin_detail = {h: 0 for h in FIN_HEADINGS}
        for heading, months in monthly_dict.items():
            super_cat = HEADING_SUPER.get(heading, "Household")
            val = sum(v for v in months.values() if v and v > 0)
            super_totals[super_cat] += val
            heading_totals[heading] = round(val / 100000, 2)
            if heading in fin_detail:
                fin_detail[heading] += val
        # group headings by super-category for drilldown
        detail_by_super = {cat: {} for cat in SUPER_ORDER}
        for heading, val in heading_totals.items():
            super_cat = HEADING_SUPER.get(heading, "Household")
            detail_by_super[super_cat][heading] = val
        return (
            {k: round(v / 100000, 2) for k, v in super_totals.items()},
            {k: round(v / 100000, 2) for k, v in fin_detail.items()},
            detail_by_super,
        )

    exp_fy24, fin_fy24, det_fy24 = mis_expenses(MIS_FY24_MONTHLY)
    exp_fy25, fin_fy25, det_fy25 = mis_expenses(MIS_FY25_MONTHLY)
    exp_fy26, fin_fy26, det_fy26 = mis_expenses(MIS_FY26_MONTHLY)

    # ── Hardcoded income / tax from tax files ────────────────────────────────
    income = {
        # FY24 from ITR AY2024-25: salary 1002.8L, ESOP perq 867.7L, divs 11.4L, interest 6.9L, foreign profit 2.1L
        # Capital gains: net losses (STCG -12.8L, LTCG -6.1L) carried forward to AY2025-26
        "FY24": {"salary": 1002, "esop": 870, "dividends": 11, "interest": 7, "capital_gains": 0, "other": 2},
        "FY25": {"salary": 1873, "esop": 1604, "dividends": 53, "interest": 0, "capital_gains": 256, "other": 14},
        # FY26 dividends: 4 GCPL interim dividends from ICICI Demat statement (59.43L)
        # FY26 ESOP: batch7 29417sh perq 346.6L + batch8 141732sh perq 1573.3L = 1920L (from ESGS Perquisite Tax Details)
        # FY26 capital_gains: FIFO over full holding (incl pledged) —
        #   May-25: 60364sh (Nov-22 @ 815.45) + 4295sh (Apr-23 @ 962.38) = 300L LTCG
        #   Feb-26: 1333sh (Apr-23) + 4598sh + 12017sh (Aug-23) + 9052sh (Oct-23) = 52.2L LTCG → total 352L
        "FY26": {"salary": 1122, "esop": 1920, "dividends": 59, "interest": 12, "capital_gains": 352, "other": 7},
    }
    # Total tax = Tax DAS (employer TDS on salary) + Tax Paid (advance in ledger for ESOP/CG/other)
    # FY26 Tax DAS: salary 1122L × 39% effective (30% slab + 25% surcharge + 4% cess) = 438L
    # FY26 Advance: 825L (ledger) covers ESOP perquisite tax + capital gains + dividends
    # UPDATE FY26 with Form 16 actual once available
    # FY24: GCPL TDS 716.7L (from ITR) + ledger advance 92L = 809L total
    # tax_das computed as tax_total - tax_paid(ledger); 809 - 92 = 717L matches ITR GCPL TDS
    tax_total = {"FY24": 809, "FY25": 1634, "FY26": 1263}

    # ── Balance sheet (net worth) ────────────────────────────────────────────
    balance_sheet = {
        "FY23": {"company_shares":550,"property":1990,"equity":930,"gold_bond":100,"nps":310,"pf":75,"gratuity":0,"private_eq":289,"art_jewellery":92,"total_assets":4336,"home_loans":0,"od":640,"loan_shares":0,"total_liabilities":640,"net_worth":3696},
        "FY24": {"company_shares":1800,"property":2140,"equity":1277,"gold_bond":135,"nps":370,"pf":150,"gratuity":0,"private_eq":289,"art_jewellery":92,"total_assets":6253,"home_loans":200,"od":495,"loan_shares":440,"total_liabilities":1135,"net_worth":5133},
        "FY25": {"company_shares":3400,"property":2740,"equity":1380,"gold_bond":300,"nps":470,"pf":234,"gratuity":0,"private_eq":346,"art_jewellery":115,"total_assets":8985,"home_loans":200,"od":538,"loan_shares":840,"total_liabilities":1578,"net_worth":7445},
        # FY26 equity (domestic): Solidarity 671+PPFAS 165+SBI ETF 90+Latent 187=1113
        # international_equity: Marcellus IB 342 (USD 360K×95)+StanChart equity 320 (USD 336K×95, Mar31 portfolio)=662
        # bank: StanChart cash 150 (USD 158K×95, Mar31 incl TDs)+ICICI 1331 15+SBI TBD≈5=170
        # gold_bond: SGB 314+Nippon Gold ETF 68=382  art_jewellery: 115+Mizugami 10+Cartier 4.51=130
        # home_loans: HL1 C51 440+HL2 C34 342+ICICI 9175 HomeSaver 571=1353  od: ICICI 7281 ~5
        # total: 3500+4080+1113+662+382+478+306+346+130+170=11167  net_worth: 11167-1358=9809
        "FY26": {"company_shares":3500,"property":4080,"equity":1113,"gold_bond":382,"nps":478,"pf":306,"private_eq":346,"art_jewellery":130,"international_equity":662,"bank":170,"total_assets":11167,"home_loans":1353,"od":5,"loan_shares":0,"total_liabilities":1358,"net_worth":9809},
    }

    # ── Asset detail ─────────────────────────────────────────────────────────
    assets_detail = {
        "FY24": [
            {"name":"GCPL Shares","type":"Listed Equity","fy24":2000,"fy25":None,"fy26":None},
            {"name":"Solidarity PMS","type":"Listed Equity","fy24":644,"fy25":None,"fy26":None},
            {"name":"PPFAS MF","type":"Mutual Fund","fy24":150,"fy25":None,"fy26":None},
            {"name":"Marcellus","type":"Mutual Fund","fy24":241,"fy25":None,"fy26":None},
            {"name":"Gold Bonds (SGB)","type":"Gold","fy24":135,"fy25":None,"fy26":None},
            {"name":"C34 Kalpataru","type":"Property","fy24":1200,"fy25":None,"fy26":None},
            {"name":"C51 Kalpataru","type":"Property","fy24":200,"fy25":None,"fy26":None},
            {"name":"Malhar Land","type":"Property","fy24":400,"fy25":None,"fy26":None},
            {"name":"Uspaar / Sarve","type":"Property","fy24":340,"fy25":None,"fy26":None},
            {"name":"NPS","type":"Retirement","fy24":370,"fy25":None,"fy26":None},
            {"name":"EPF","type":"Retirement","fy24":150,"fy25":None,"fy26":None},
            {"name":"Private Equity (LO Foods, LocalBuy, Licious etc.)","type":"Private Eq.","fy24":182,"fy25":None,"fy26":None},
            {"name":"Art & Jewellery","type":"Alternative","fy24":93,"fy25":None,"fy26":None},
        ]
    }

    # Build unified asset detail list across all years
    assets = [
        {"name":"GCPL Shares","type":"Listed Equity","fy24":2000,"fy25":3400,"fy26":3500},
        {"name":"Solidarity PMS","type":"Listed Equity","fy24":644,"fy25":724,"fy26":671},
        {"name":"PPFAS MF","type":"Mutual Fund","fy24":150,"fy25":164,"fy26":165},
        {"name":"SBI ETF Nifty Next 50","type":"Mutual Fund","fy24":None,"fy25":107,"fy26":90},
        {"name":"Latent AIF","type":"AIF","fy24":None,"fy25":107,"fy26":187},
        {"name":"Marcellus IB (USD)","type":"International Equity","fy24":241,"fy25":302,"fy26":342},
        {"name":"StanChart Equity Portfolio","type":"International Equity","fy24":None,"fy25":None,"fy26":320},
        {"name":"StanChart Cash / Savings","type":"Bank Account","fy24":None,"fy25":None,"fy26":150},
        {"name":"ICICI 1331","type":"Bank Account","fy24":None,"fy25":None,"fy26":15},
        {"name":"Nippon Gold ETF (GOLDBEES)","type":"Gold","fy24":None,"fy25":None,"fy26":68},
        {"name":"Gold Bonds (SGB)","type":"Gold","fy24":135,"fy25":301,"fy26":314},
        {"name":"Godrej Pet (GPA)","type":"Private Eq.","fy24":None,"fy25":164,"fy26":164},
        {"name":"LO Foods","type":"Private Eq.","fy24":20,"fy25":20,"fy26":20},
        {"name":"LocalBuy (Superk)","type":"Private Eq.","fy24":100,"fy25":100,"fy26":100},
        {"name":"Licious","type":"Private Eq.","fy24":52,"fy25":52,"fy26":52},
        {"name":"X to 10X","type":"Private Eq.","fy24":10,"fy25":10,"fy26":10},
        {"name":"C34 Kalpataru","type":"Property","fy24":1200,"fy25":1200,"fy26":1300},
        {"name":"C51 Kalpataru","type":"Property","fy24":200,"fy25":800,"fy26":880},
        {"name":"Malhar Land","type":"Property","fy24":400,"fy25":400,"fy26":900},
        {"name":"Uspaar / Sarve Land","type":"Property","fy24":340,"fy25":340,"fy26":1000},
        {"name":"NPS","type":"Retirement","fy24":370,"fy25":470,"fy26":478},
        {"name":"EPF","type":"Retirement","fy24":150,"fy25":234,"fy26":306},
        {"name":"Art & Jewellery","type":"Alternative","fy24":93,"fy25":106,"fy26":130},
    ]

    liabilities = [
        {"name":"HL1 — C51 Kalpataru (top-up) XX99508","type":"Home Loan","fy24":None,"fy25":None,"fy26":440},
        {"name":"HL2 — C34 Kalpataru (top-up) XX00382","type":"Home Loan","fy24":None,"fy25":None,"fy26":342},
        {"name":"HL3 — C34 HomeSaver","type":"Home Loan","fy24":None,"fy25":200,"fy26":None},
        {"name":"HL4 — C51 original XX42596","type":"Home Loan","fy24":200,"fy25":None,"fy26":191},
        {"name":"HL5 — ICICI 9175 HomeSaver (MaxGain)","type":"Home Loan","fy24":495,"fy25":538,"fy26":571},
        {"name":"OD — ICICI 7281","type":"Credit Line","fy24":None,"fy25":None,"fy26":5},
        {"name":"Loan against shares","type":"Pledge Loan","fy24":440,"fy25":840,"fy26":None},
    ]

    def pl_for_fy(fy, exp, fin_detail, det):
        inc = income[fy]
        total_inc = inc["salary"] + inc["esop"] + inc["dividends"] + inc["interest"] + inc["capital_gains"] + inc["other"]
        non_fin = sum(exp.get(s, 0) for s in NON_FIN)
        # Financial = Home Loan + Insurance + OD Interest (Tax handled separately below)
        home_loan = fin_detail.get("Home Loan", 0)
        insurance = fin_detail.get("Insurance", 0)
        od_interest = round(
            fin_detail.get("Financial Expense / OD Interest", 0) +
            fin_detail.get("Financial Expense", 0), 2)
        personal_loans = fin_detail.get("Personal Loans", 0)
        tax_paid = fin_detail.get("Tax", 0)                       # advance/self-assessment in ledger
        tax_das = round(max(0, tax_total[fy] - tax_paid), 2)      # TDS by employer (floor 0 for slight overpayments)
        fin_excl_tax = round(home_loan + insurance + od_interest + personal_loans, 2)
        total_tax = round(tax_paid + tax_das, 2)
        total_exp = round(non_fin + fin_excl_tax + total_tax, 2)
        return {
            "income": inc,
            "total_income": round(total_inc, 2),
            "expenses": exp,
            "expense_detail": det,
            "non_financial_total": round(non_fin, 2),
            "fin_detail": {
                "home_loan": home_loan,
                "insurance": insurance,
                "od_interest": od_interest,
                "personal_loans": personal_loans,
                "tax_paid": tax_paid,
                "tax_das": tax_das,
                "tax_total_estimated": fy == "FY26",  # flag to show (est.) in UI until Form 16 loaded
            },
            "financial_total": fin_excl_tax,
            "total_tax": total_tax,
            "total_expenditure": total_exp,
            "net_surplus": round(total_inc - total_exp, 2),
        }

    return jsonify({
        "pl": {
            "FY24": pl_for_fy("FY24", exp_fy24, fin_fy24, det_fy24),
            "FY25": pl_for_fy("FY25", exp_fy25, fin_fy25, det_fy25),
            "FY26": pl_for_fy("FY26", exp_fy26, fin_fy26, det_fy26),
        },
        "balance_sheet": balance_sheet,
        "assets": assets,
        "liabilities": liabilities,
        "super_order": SUPER_ORDER,
        "non_fin_supers": NON_FIN,
    })


def _default_asset_registry():
    return {"classes": [
        # ── ASSETS ──────────────────────────────────────────────────────────────
        {"id":"gcpl_shares","label":"Company Shares — GCPL","section":"assets","icon":"bi-building","items":[
            {"id":"gcpl_main","name":"GCPL Shares (vested pool)","date_acquired":"Various tranches FY21–FY26","purchase_value_L":None,"value_mar26_L":3500,"value_today_L":None,
             "notes":"141,732 shares vested as of Mar 26. Includes pledged shares. Excludes unvested grants.",
             "documents":["FY26/4_Financial_Investments/Sudhir Sitapati_ESGS Data.xlsx","FY26/4_Financial_Investments/Demat Account_90174124_01_04_2025-31_03_2026 (2).pdf","FY26/4_Financial_Investments/ICICI_Demat_90174124_Holdings_Mar26.pdf"],"sub_items":[]},
        ]},
        {"id":"listed_equity","label":"Listed Equity / PMS","section":"assets","icon":"bi-graph-up-arrow","items":[
            {"id":"solidarity","name":"Solidarity PMS","date_acquired":"~FY22","purchase_value_L":None,"value_mar26_L":671,"value_today_L":None,
             "notes":"Domestic long-only PMS. Managed by Solidarity Investment Managers.",
             "documents":["FY26/4_Financial_Investments/Solidaruty.pdf"],"sub_items":[]},
            {"id":"latent_pms","name":"Latent PMF","date_acquired":"~FY25","purchase_value_L":None,"value_mar26_L":187,"value_today_L":None,
             "notes":"Portfolio Management Fund (PMF). Long-only concentrated equity strategy.","documents":["FY26/4_Financial_Investments/latent.pdf"],"sub_items":[]},
        ]},
        {"id":"mutual_funds","label":"Mutual Funds","section":"assets","icon":"bi-pie-chart","items":[
            {"id":"ppfas","name":"PPFAS Flexi Cap MF","date_acquired":"~FY22","purchase_value_L":None,"value_mar26_L":165,"value_today_L":None,
             "notes":"Parag Parikh Flexi Cap Fund. Folio 15152304.","documents":["FY26/4_Financial_Investments/Sudhir Sitapati_15152304.pdf"],"sub_items":[]},
            {"id":"sbi_etf","name":"SBI ETF Nifty Next 50","date_acquired":"~FY24","purchase_value_L":None,"value_mar26_L":90,"value_today_L":None,
             "notes":"Index ETF held in ICICI Demat 90174124.","documents":[],"sub_items":[]},
        ]},
        {"id":"international_equity","label":"International Equity","section":"assets","icon":"bi-globe2","items":[
            {"id":"marcellus_ib","name":"Marcellus International Basket (USD)","date_acquired":"Feb 2026","purchase_value_L":None,"value_mar26_L":342,"value_today_L":None,
             "notes":"USD 360K × ₹95. Managed by Marcellus Investment Managers. A/c U11622321.",
             "documents":["FY26/4_Financial_Investments/Marcellus_IB_U11622321_Feb26.pdf","FY26/4_Financial_Investments/Marcellus_Summary_FY26.pdf"],"sub_items":[]},
            {"id":"stanchart_intl","name":"StanChart International Portfolio","date_acquired":"~FY23","purchase_value_L":None,"value_mar26_L":470,"value_today_L":None,
             "notes":"Standard Chartered Bank Singapore A/c 62-1-833928-2. Valuation as of 31-Mar-2026 from portfolio statement (generated 19-Jun-2026).",
             "documents":["FY26/4_Financial_Investments/Portfolio-Sudhir.pdf","FY26/1_Bank_Accounts/StanChart_6007_SGD_Statement_Apr26.pdf","FY26/1_Bank_Accounts/eStatement_Consolidated Statement_6007_SGD_Apr_2026.pdf"],
             "sub_items":[
                {"id":"sc_gdx","name":"VanEck Gold Miners ETF (GDX) — 1,043 units","value_mar26_L":91,"notes":"USD 95,716 × ₹95 = 91L. Largest position. 19.4% of portfolio. 103% unrealised gain since Mar 2025."},
                {"id":"sc_gld","name":"SPDR Gold Shares ETF (GLD) — 160 units","value_mar26_L":65,"notes":"USD 68,846 × ₹95 = 65L. 13.9% of portfolio. 54% unrealised gain."},
                {"id":"sc_slv","name":"iShares Silver Trust ETF (SLV) — 470 units","value_mar26_L":30,"notes":"USD 32,026 × ₹95 = 30L. 6.5% of portfolio. 429% total return incl realised."},
                {"id":"sc_qqq","name":"Invesco QQQ ETF (QQQ) — 65 units","value_mar26_L":36,"notes":"USD 37,517 × ₹95 = 36L. 7.6% of portfolio."},
                {"id":"sc_ilf","name":"iShares Latin America 40 ETF (ILF) — 800 units","value_mar26_L":27,"notes":"USD 28,416 × ₹95 = 27L. 5.7% of portfolio."},
                {"id":"sc_cnya","name":"iShares MSCI China A ETF (CNYA) — 800 units","value_mar26_L":26,"notes":"USD 27,344 × ₹95 = 26L. 5.5% of portfolio."},
                {"id":"sc_asea","name":"Global X FTSE SE Asia ETF (ASEA) — 1,200 units","value_mar26_L":22,"notes":"USD 23,316 × ₹95 = 22L. 4.7% of portfolio."},
                {"id":"sc_ewj","name":"iShares MSCI Japan ETF (EWJ) — 200 units","value_mar26_L":16,"notes":"USD 16,888 × ₹95 = 16L. 3.4% of portfolio."},
                {"id":"sc_qqq_hkd","name":"Invesco QQQ ETF 3455 (HKD) — 11 units","value_mar26_L":6,"notes":"HKD 48,653 = USD 6,206 × ₹95 = 6L. 1.3% of portfolio."},
                {"id":"sc_intc","name":"Intel Corp (INTC) — 2 units","value_mar26_L":0,"notes":"USD 88. Negligible position."},
                {"id":"sc_usd_cash","name":"USD Cash (A/c 6208256430)","value_mar26_L":64,"notes":"USD 67,113 × ₹95 = 64L. Separate from AVER account."},
                {"id":"sc_aver","name":"USD AVER (A/c 6218339282)","value_mar26_L":86,"notes":"USD 90,974 × ₹95 = 86L. This is the main USD savings account (Apr-2 statement balance)."},
             ]},
        ]},
        {"id":"gold","label":"Gold","section":"assets","icon":"bi-gem","items":[
            {"id":"sgb","name":"Sovereign Gold Bonds (SGB)","date_acquired":"Multiple tranches FY21–FY24","purchase_value_L":None,"value_mar26_L":314,"value_today_L":None,
             "notes":"RBI-issued SGBs. 8-year maturity. Interest 2.5% p.a. tax-free at maturity.","documents":[],"sub_items":[]},
            {"id":"gold_etf","name":"Nippon Gold ETF (GOLDBEES)","date_acquired":"~FY25","purchase_value_L":None,"value_mar26_L":68,"value_today_L":None,
             "notes":"Gold ETF held in ICICI Demat 90174124.","documents":[],"sub_items":[]},
        ]},
        {"id":"property","label":"Property","section":"assets","icon":"bi-house-door","items":[
            {"id":"c34","name":"C34 Kalpataru — Thane (primary residence)","date_acquired":"~2018","purchase_value_L":None,"value_mar26_L":1300,"value_today_L":None,
             "notes":"4BHK flat at Kalpataru Serenity, Thane. Linked to HL2 (top-up) and HL5 (ICICI HomeSaver/MaxGain).","documents":[],"sub_items":[]},
            {"id":"c51","name":"C51 Kalpataru — Thane","date_acquired":"~2021","purchase_value_L":None,"value_mar26_L":880,"value_today_L":None,
             "notes":"Flat at Kalpataru, Thane. Linked to HL1 (top-up) and HL4 (original).","documents":[],"sub_items":[]},
            {"id":"malhar","name":"Malhar Land — Kashid","date_acquired":"~FY22","purchase_value_L":None,"value_mar26_L":900,"value_today_L":None,
             "notes":"Land parcel near Kashid beach. Uspaar / Sarve entity.","documents":[],"sub_items":[]},
            {"id":"uspaar","name":"Uspaar / Sarve Land","date_acquired":"~FY22","purchase_value_L":None,"value_mar26_L":1000,"value_today_L":None,
             "notes":"Land holding via Uspaar / Sarve.","documents":[],"sub_items":[]},
        ]},
        {"id":"retirement","label":"Retirement — NPS / EPF","section":"assets","icon":"bi-shield-check","items":[
            {"id":"nps","name":"NPS (National Pension System)","date_acquired":"~FY20","purchase_value_L":None,"value_mar26_L":478,"value_today_L":None,
             "notes":"Tier I NPS account. Employer + employee contributions. Locked till retirement.","documents":[],"sub_items":[]},
            {"id":"epf","name":"EPF / PF","date_acquired":"~FY15","purchase_value_L":None,"value_mar26_L":306,"value_today_L":None,
             "notes":"Employee Provident Fund. GCPL payroll deductions.","documents":["FY26/4_Financial_Investments/SohBeneficiaryRpt - 2026-06-10T131210.333.pdf","FY26/3_Retirement/718379.pdf"],"sub_items":[]},
        ]},
        {"id":"private_eq","label":"Private Equity","section":"assets","icon":"bi-rocket-takeoff","items":[
            {"id":"gpa","name":"Godrej Pet (GPA)","date_acquired":"~FY24","purchase_value_L":164,"value_mar26_L":164,"value_today_L":None,"notes":"Unlisted. Cost basis held as value.","documents":[],"sub_items":[]},
            {"id":"lo_foods","name":"LO Foods","date_acquired":"~FY22","purchase_value_L":20,"value_mar26_L":20,"value_today_L":None,"notes":"Unlisted. Cost basis.","documents":[],"sub_items":[]},
            {"id":"localbuy","name":"LocalBuy / Superk","date_acquired":"~FY22","purchase_value_L":100,"value_mar26_L":100,"value_today_L":None,"notes":"Unlisted. Cost basis.","documents":[],"sub_items":[]},
            {"id":"licious","name":"Licious","date_acquired":"~FY22","purchase_value_L":52,"value_mar26_L":52,"value_today_L":None,"notes":"Unlisted. Cost basis.","documents":[],"sub_items":[]},
            {"id":"x_to_10x","name":"X to 10X","date_acquired":"~FY22","purchase_value_L":10,"value_mar26_L":10,"value_today_L":None,"notes":"Unlisted. Cost basis.","documents":[],"sub_items":[]},
        ]},
        {"id":"alt","label":"Art & Jewellery","section":"assets","icon":"bi-palette","items":[
            {"id":"art_jewellery","name":"Art & Jewellery","date_acquired":"Various","purchase_value_L":None,"value_mar26_L":130,"value_today_L":None,
             "notes":"Includes: Mizugami artwork (~10L), Cartier jewellery (~4.5L), and other art/jewellery (~115L).","documents":[],"sub_items":[]},
        ]},
        {"id":"bank","label":"Bank Accounts","section":"assets","icon":"bi-bank","items":[
            {"id":"icici_1331","name":"ICICI Bank 1331 (savings)","date_acquired":None,"purchase_value_L":None,"value_mar26_L":15,"value_today_L":None,
             "notes":"Primary operating account. ICICI Bank A/c ending 1331.","documents":[],"sub_items":[]},
            {"id":"sbi_4852","name":"SBI 4852 (savings)","date_acquired":None,"purchase_value_L":None,"value_mar26_L":5,"value_today_L":None,
             "notes":"SBI savings account. Approximate balance.","documents":[],"sub_items":[]},
        ]},
        {"id":"insurance","label":"Insurance — LIC Policies","section":"assets","icon":"bi-umbrella","items":[
            {"id":"ins_895456071","name":"LIC 895456071 — Sudhir (Plan 823, 15yr)","date_acquired":"24/03/2014","purchase_value_L":None,"value_mar26_L":None,"value_today_L":None,
             "notes":"LIC Plan 823, Term 15yr, Premium 15yr. Sum Assured ₹4 Cr. Premium ₹1,00,000 p.a. (incl. GST). FUP 03/27. Maturity 03/29. Nominee: Ketaki. NACH mode.",
             "documents":["data/2026 Policy Register.pdf"],"sub_items":[]},
            {"id":"ins_896050907","name":"LIC 896050907 — Sudhir (Plan 823, 20yr)","date_acquired":"28/12/2016","purchase_value_L":None,"value_mar26_L":None,"value_today_L":None,
             "notes":"LIC Plan 823, Term 20yr, Premium 20yr. Sum Assured ₹4 Cr. Premium ₹1,42,800 p.a. FUP 12/26. Maturity 12/36. Nominee: Kamala Ganesh R P. NACH mode.",
             "documents":["data/2026 Policy Register.pdf"],"sub_items":[]},
            {"id":"ins_934301314","name":"LIC 934301314 — Sudhir (Plan 855, 10yr)","date_acquired":"04/10/2022","purchase_value_L":None,"value_mar26_L":None,"value_today_L":None,
             "notes":"LIC Plan 855, Term 10yr, Premium 10yr. Sum Assured ₹4 Cr. Premium ₹1,01,320 p.a. (incl. GST). FUP 10/26. Maturity 10/32. Nominee: Ketaki. NACH mode.",
             "documents":["data/2026 Policy Register.pdf"],"sub_items":[]},
            {"id":"ins_895992110","name":"LIC 895992110 — Ketaki (Plan 815, 21yr)","date_acquired":"04/07/2016","purchase_value_L":None,"value_mar26_L":None,"value_today_L":None,
             "notes":"LIC Plan 815, Term 21yr, Premium 21yr. Sum Assured ₹75L + Accidental Benefit ₹75L. Premium ₹4,17,915 p.a. (incl. GST). FUP 07/26. Maturity 07/37. Nominee: Sudhir. NACH mode.",
             "documents":["data/2026 Policy Register.pdf"],"sub_items":[]},
        ]},
        # ── LIABILITIES ──────────────────────────────────────────────────────────
        {"id":"home_loans","label":"Home Loans","section":"liabilities","icon":"bi-house","items":[
            {"id":"hl1_c51","name":"HL1 — C51 Kalpataru top-up (XX99508)","date_acquired":None,"purchase_value_L":None,"value_mar26_L":440,"value_today_L":None,
             "notes":"Top-up home loan against C51. Outstanding ₹440L as at Mar 31 2026.","documents":["FY26/2_Home_Loans/ICICI_TopUp_XX99508_C51_Statement.pdf","FY26/2_Home_Loans/Statement of Account_TBMUM00007499508 (1).pdf"],"sub_items":[]},
            {"id":"hl2_c34","name":"HL2 — C34 Kalpataru top-up (XX00382)","date_acquired":None,"purchase_value_L":None,"value_mar26_L":342,"value_today_L":None,
             "notes":"Top-up home loan against C34. Outstanding ₹342L as at Mar 31 2026.","documents":["FY26/2_Home_Loans/ICICI_TopUp_XX00382_C34_Statement.pdf","FY26/2_Home_Loans/Statement of Account_TBMUM00007500382 (1).pdf"],"sub_items":[]},
            {"id":"hl4_c51","name":"HL4 — C51 original (XX42596)","date_acquired":None,"purchase_value_L":None,"value_mar26_L":191,"value_today_L":None,
             "notes":"Original home loan for C51.","documents":["FY26/2_Home_Loans/Statement of Account_TBMUM00006542596 (1).pdf"],"sub_items":[]},
            {"id":"hl5_maxgain","name":"HL5 — ICICI 9175 HomeSaver MaxGain","date_acquired":None,"purchase_value_L":None,"value_mar26_L":380,"value_today_L":None,
             "notes":"ICICI HomeSaver (overdraft-linked) for C34. A/c 9175.","documents":["FY26/1_Bank_Accounts/ICICI_9175_HomeSaver_Statement_FY26.pdf"],"sub_items":[]},
        ]},
        {"id":"credit_lines","label":"Credit Lines / OD","section":"liabilities","icon":"bi-credit-card","items":[
            {"id":"od_7281","name":"OD — ICICI 7281","date_acquired":None,"purchase_value_L":None,"value_mar26_L":5,"value_today_L":None,
             "notes":"Overdraft facility ICICI A/c 7281. Current utilisation ~5L.","documents":[],"sub_items":[]},
        ]},
        {"id":"loan_shares","label":"Loan Against Shares","section":"liabilities","icon":"bi-bank2","items":[]},
    ]}

@app.route("/api/asset-registry", methods=["GET"])
@login_required
def api_asset_registry_get():
    registry = db.load("asset_registry")
    if not registry:
        registry = _default_asset_registry()
    return jsonify(registry)

@app.route("/api/asset-registry", methods=["POST"])
@login_required
def api_asset_registry_save():
    data = request.get_json(force=True)
    db.save("asset_registry", data)
    return jsonify({"ok": True})

@app.route("/api/asset-registry/reset", methods=["POST"])
@login_required
def api_asset_registry_reset():
    db.save("asset_registry", _default_asset_registry())
    return jsonify({"ok": True})


@app.route("/api/data-file/<path:filepath>", methods=["GET"])
@login_required
def api_data_file(filepath):
    """Serve any file from the data/ directory."""
    directory = os.path.dirname(os.path.join(DATA_DIR, filepath))
    filename = os.path.basename(filepath)
    return send_from_directory(directory, filename)


def _assign_missing_seq(ledger):
    """Assign seq numbers to any transactions that don't have one yet."""
    max_seq = max((t.get("seq") or 0 for t in ledger), default=0)
    changed = False
    for t in ledger:
        if not t.get("seq"):
            max_seq += 1
            t["seq"] = max_seq
            changed = True
    return changed


# ── Cash Register (SBI petty cash) ──────────────────────────────────────────

def _parse_ledger_date(d):
    from datetime import datetime
    for fmt in ('%d-%b-%y', '%d-%b-%Y', '%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(d.strip(), fmt)
        except Exception:
            pass
    return None


def _sbi_cash_ledger_entries(fy=27):
    """Return FY-filtered cash ledger entries.

    Cash Given (credit side): account=cash, source=transfer, credit>0 (cashin_* entries).
    Cash Spent (debit side):  account=cash, debit>0 (approval_log or cash_register).
    """
    from datetime import datetime
    fy_start = datetime(2000 + fy - 1, 4, 1)
    fy_end   = datetime(2000 + fy,     3, 31, 23, 59, 59)

    ledger = db.load("master_ledger") or []
    out = []
    for t in ledger:
        acct = (t.get("account") or "").upper()
        if acct != "CASH":
            continue

        debit  = float(t.get("debit") or 0)
        credit = float(t.get("credit") or 0)
        src    = t.get("source") or ""

        if src == "transfer" and credit > 0:
            side   = "given"
            amount = credit
        elif (src in ("approval_log", "cash_register") or debit > 0) and debit > 0:
            side   = "spent"
            amount = debit
        else:
            continue

        dt = _parse_ledger_date(t.get("date") or "")
        if dt is None or not (fy_start <= dt <= fy_end):
            continue

        out.append({
            "id":        t.get("txn_id") or t.get("id") or "",
            "seq":       t.get("seq"),
            "date":      t.get("date") or "",
            "date_iso":  dt.strftime("%Y-%m-%d"),
            "amount":    amount,
            "narration": t.get("paid_to") or t.get("remarks") or "ATM / SBI Withdrawal",
            "heading":   (t.get("heading") or "").strip(),
            "side":      side,
            "source":    src,
            "deletable": src == "cash_register",
            "method":    t.get("method") or "",
            "bill_b64":  t.get("bill_b64") or "",
            "notes":     t.get("remarks") or "",
        })

    out.sort(key=lambda x: (x["date_iso"], x["seq"] or 0))
    return out


def _sbi_upi_to_staff(staff_list):
    """Pull UPI transfers from SBI to known staff from master ledger."""
    if not staff_list:
        return []
    # Build lookup: token → staff name
    tokens = {}
    for s in staff_list:
        name = s.get("name", "")
        for field in ("upi_id", "upi_phone"):
            val = (s.get(field) or "").lower().strip()
            if val:
                tokens[val] = name
        # Also match by first name in paid_to/narration
        if name:
            tokens[name.lower().split()[0]] = name

    ledger = db.load("master_ledger") or []
    out = []
    for t in ledger:
        acct = (t.get("account") or "").upper()
        if "SBI" not in acct:
            continue
        typ = (t.get("type") or "").lower()
        if typ != "transfer":
            continue
        amt = t.get("amount") or 0
        if amt <= 0:
            continue
        narr = (t.get("paid_to") or t.get("remarks") or "").lower()
        heading = (t.get("heading") or "").lower()
        # Skip ATM/cash withdrawals (already in withdrawals list)
        if heading in ("cash",) or "atm" in narr or "withdrawal" in narr:
            continue
        matched_staff = None
        for token, sname in tokens.items():
            if token and token in narr:
                matched_staff = sname
                break
        if not matched_staff:
            continue
        out.append({
            "id": t.get("txn_id") or t.get("id") or "",
            "date": t.get("date") or "",
            "amount": amt,
            "narration": t.get("paid_to") or t.get("remarks") or "",
            "staff": matched_staff,
            "source": "ledger",
            "type": "upi_staff",
        })
    out.sort(key=lambda x: x["date"], reverse=True)
    return out


def _detect_staff_upis(entries, staff_list):
    """Tag register entries whose narration matches a staff UPI id or phone."""
    upi_map = {}
    for s in staff_list:
        for field in ("upi_id", "upi_phone"):
            val = (s.get(field) or "").lower().strip()
            if val:
                upi_map[val] = s["name"]
    for e in entries:
        narr = (e.get("narration") or "").lower()
        for token, name in upi_map.items():
            if token and token in narr:
                e.setdefault("staff", name)
                break
    return entries


@app.route("/api/cash-register", methods=["GET"])
@login_required
def api_cash_register_get():
    staff       = (db.load("cash_staff") or {}).get("staff", [])
    cash_ledger = _sbi_cash_ledger_entries()   # includes source="cash_register" entries
    upi_staff   = _sbi_upi_to_staff(staff)
    upi_ids     = {u["id"] for u in upi_staff}
    cash_ledger = [e for e in cash_ledger if e["id"] not in upi_ids]
    return jsonify({
        "cash_ledger": cash_ledger,
        "upi_staff":   upi_staff,
        "staff":       staff,
    })


@app.route("/api/cash-register/entry", methods=["POST"])
@login_required
def api_cash_register_entry():
    """Save a cash register payment as a master ledger entry."""
    data   = request.get_json(force=True)
    ledger = db.load("master_ledger") or []

    who     = data.get("staff") or data.get("vendor") or "Cash payment"
    txn_id  = data.get("id") or ("cr_" + datetime.utcnow().strftime("%Y%m%d%H%M%S%f"))
    amount  = float(data.get("amount") or 0)

    # Update: remove old entry first
    ledger = [t for t in ledger if t.get("txn_id") != txn_id]

    entry = {
        "txn_id":          txn_id,
        "seq":             max((t.get("seq") or 0 for t in ledger), default=0) + 1,
        "date":            data.get("date") or "",
        "account":         "cash",
        "bank":            "cash",
        "type":            "expense",
        "heading":         data.get("heading") or "Misc",
        "paid_to":         who,
        "debit":           amount,
        "credit":          0.0,
        "amount":          amount,
        "remarks":         data.get("notes") or "",
        "raw_description": f"{who} – {data.get('heading') or 'Cash'}",
        "source":          "cash_register",
        "method":          data.get("method") or "cash",
        "cr_type":         data.get("type") or "vendor",
        "bill_b64":        data.get("bill_b64") or "",
        "confidence":      "manual",
        "uncertain":       False,
        "uncertain_fields":[],
    }
    ledger.append(entry)
    ledger.sort(key=lambda t: t.get("date", ""), reverse=True)
    db.save("master_ledger", ledger)
    return jsonify({"ok": True, "id": txn_id, "seq": entry["seq"]})


@app.route("/api/cash-register/entry/<eid>", methods=["DELETE"])
@login_required
def api_cash_register_delete(eid):
    """Delete a cash register entry from the master ledger."""
    ledger = db.load("master_ledger") or []
    ledger = [t for t in ledger if t.get("txn_id") != eid]
    db.save("master_ledger", ledger)
    return jsonify({"ok": True})


@app.route("/api/cash-staff", methods=["GET"])
@login_required
def api_cash_staff_get():
    return jsonify(db.load("cash_staff") or {"staff": [
        {"id":"shiloch","name":"Shiloch","upi_id":"","upi_phone":""},
        {"id":"mary","name":"Mary","upi_id":"","upi_phone":""},
        {"id":"santosh","name":"Santosh","upi_id":"","upi_phone":""},
        {"id":"mohammed","name":"Mohammed","upi_id":"","upi_phone":""},
    ]})


@app.route("/api/cash-staff", methods=["POST"])
@login_required
def api_cash_staff_save():
    db.save("cash_staff", request.get_json(force=True))
    return jsonify({"ok": True})


@app.route("/api/cash-register/ai-extract", methods=["POST"])
@login_required
def api_cash_register_ai_extract():
    """Extract amount, vendor, date from a bill image (base64) using Azure OpenAI."""
    data = request.get_json(force=True)
    image_b64 = data.get("image_b64", "")
    mime = data.get("mime", "image/jpeg")
    if not image_b64:
        return jsonify({"error": "no image"}), 400
    try:
        from src.azure_openai import client, DEPLOYMENT
        resp = client.chat.completions.create(
            model=DEPLOYMENT,
            max_completion_tokens=400,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{image_b64}"}},
                    {"type": "text", "text": (
                        "This is a bill or receipt. Extract: amount (number only, INR), "
                        "vendor/shop name, date (YYYY-MM-DD if visible). "
                        "Also suggest the best expense heading from: Groceries, Staff Salary, "
                        "Electricity & Gas, Wellness, Clothes, Gifts, Medical, Children Education, "
                        "Holiday, Eating Out, Entertainment, Malhar, Maintenance Expense, "
                        "Home office, One Time Charge, Kalpataru Maintenance, Misc. "
                        "Reply as JSON: {\"amount\": 0, \"vendor\": \"\", \"date\": \"\", \"heading\": \"\", \"notes\": \"\"}"
                    )}
                ]
            }]
        )
        import re, json as _json
        raw = resp.choices[0].message.content or ""
        m = re.search(r'\{.*\}', raw, re.DOTALL)
        result = _json.loads(m.group()) if m else {"amount": 0, "vendor": "", "date": "", "heading": "", "notes": raw}
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/cash-recon", methods=["GET"])
def api_cash_recon():
    """ATM withdrawals vs cash-authorised payments."""
    transactions = _load_json(TRANSACTIONS_PATH)
    log = _load_json(APPROVAL_LOG)

    # ATM withdrawals from ICICI transactions
    atm_txns = [t for t in transactions if "atm" in (t.get("description") or "").lower() and t.get("type") == "debit"]
    atm_total = sum(t.get("amount", 0) for t in atm_txns)

    # Cash payments from approval log (this month)
    month_prefix = datetime.now().strftime("%Y-%m")
    cash_approved = [
        e for e in log
        if e.get("payment_method") == "cash"
        and e.get("action") in ("AUTO_APPROVE", "APPROVED", "APPROVED_LOWER")
        and (e.get("timestamp") or "").startswith(month_prefix)
    ]
    cash_paid_total = sum((e.get("approved_amount") or e.get("amount", 0)) for e in cash_approved)

    # Monthly breakdown
    monthly: dict = {}
    for t in atm_txns:
        m = (t.get("date") or "")[:7]
        if m:
            monthly.setdefault(m, {"atm": 0, "cash_paid": 0})
            monthly[m]["atm"] += t.get("amount", 0)

    for e in [x for x in log if x.get("payment_method") == "cash" and x.get("action") in ("AUTO_APPROVE", "APPROVED", "APPROVED_LOWER")]:
        m = (e.get("timestamp") or "")[:7]
        if m:
            monthly.setdefault(m, {"atm": 0, "cash_paid": 0})
            monthly[m]["cash_paid"] += e.get("approved_amount") or e.get("amount", 0)

    monthly_rows = [{"month": k, **v} for k, v in sorted(monthly.items(), reverse=True)]

    return jsonify({
        "atm_total": atm_total,
        "cash_paid_total": cash_paid_total,
        "monthly_rows": monthly_rows,
    })


@app.route("/api/submit-expense", methods=["POST"])
def api_submit_expense():
    """Handle expense submission from the web form."""
    data = request.get_json()

    # Follow-up answer to a pending clarification
    if "request_id" in data and "followup_answer" in data:
        req_id = data["request_id"]
        answer = data["followup_answer"]
        req = PENDING_CLARIFICATION.get(req_id)
        if req:
            req.description = f"{req.description} [{answer}]"
            PENDING_CLARIFICATION.pop(req_id, None)

            if data.get("followup_round", 1) >= 2:
                # Max rounds hit — escalate with unclear flag
                from src.approval_engine import ApprovalDecision
                decision = ApprovalDecision(
                    request_id=req_id, action="ESCALATE",
                    reason="Unclear pricing after 2 follow-up rounds"
                )
                decision.escalation_message = engine._build_escalation_message(req, "\n⚠️ Pricing unclear after follow-up")
                send_approval_request(decision.escalation_message)
                return jsonify({"action": "ESCALATE", "request_id": req_id,
                                "vendor": req.vendor, "amount": req.amount})

            decision = engine.evaluate(req)
        else:
            return jsonify({"action": "ESCALATE", "request_id": req_id})
    else:
        # Fresh submission
        try:
            # Use submitted date if provided, preserving HH:MM:SS from now
            _expense_date = data.get("expense_date", "")
            _timestamp = None
            if _expense_date:
                from datetime import datetime as _dt
                _now = _dt.now()
                _timestamp = f"{_expense_date}T{_now.strftime('%H:%M:%S')}"
            req = ExpenseRequest(
                submitter=data["submitter"],
                vendor=data["vendor"],
                amount=float(data["amount"]),
                category=data.get("category", ""),
                description=data["description"],
                payment_method=data.get("payment_method", "upi"),
                is_post_facto=data.get("is_post_facto", False),
                heading=data.get("heading") or data.get("category", ""),
                expense_type=data.get("expense_type", "expense"),
                timestamp=_timestamp or "",
            )
        except (KeyError, ValueError) as e:
            return jsonify({"error": str(e)}), 400

        decision = engine.evaluate(req)

    if decision.action == "PENDING_CLARIFICATION":
        PENDING_CLARIFICATION[decision.request_id] = req

    if decision.action == "AUTO_APPROVE":
        sync_approved_to_history()
        try:
            send_auto_approval_notice(req.submitter, req.vendor, req.amount, decision.request_id)
        except Exception as ex:
            print(f"[whatsapp] auto-approve notify failed: {ex}")
    elif decision.action == "ESCALATE" and decision.escalation_message:
        # Generate AI commentary for Sudhir's review screen (non-blocking)
        try:
            req.ai_comment = engine.generate_ai_comment(req)
        except Exception:
            pass
        # Update log with AI comment
        if req.ai_comment:
            log = _load_json(APPROVAL_LOG)
            for e in log:
                if e.get("request_id") == req.request_id:
                    e["ai_comment"] = req.ai_comment
                    break
            _save_json(APPROVAL_LOG, log)
        try:
            send_approval_request(decision.escalation_message)
        except Exception as ex:
            print(f"[whatsapp] escalate notify failed: {ex}")

    return jsonify({
        "action": decision.action,
        "request_id": decision.request_id,
        "vendor": req.vendor,
        "amount": req.amount,
        "market_rate": decision.market_rate,
        "market_status": decision.market_status,
        "budget_alert": decision.budget_alert,
        "follow_up_question": decision.follow_up_question,
        "follow_up_options": decision.follow_up_options,
    })


@app.route("/api/decide", methods=["POST"])
def api_decide():
    """Approve, reject, or query from dashboard buttons."""
    data = request.get_json()
    request_id = data.get("request_id")
    response = data.get("response", "")
    query_text = data.get("query_text", "")

    log = _load_json(APPROVAL_LOG)
    entry = next((e for e in log if e.get("request_id") == request_id), None)
    if not entry:
        return jsonify({"error": "not found"}), 404

    engine.update_log_with_sudhir_response(request_id, response, query_text=query_text)

    resp_upper = response.strip().upper()
    try:
        if resp_upper == "Y":
            send_approval_result(entry["submitter"], entry["vendor"], entry["amount"], approved=True, request_id=request_id)
            sync_approved_to_history()
        elif resp_upper == "N":
            send_approval_result(entry["submitter"], entry["vendor"], entry["amount"], approved=False, request_id=request_id)
        elif resp_upper == "Q" and query_text:
            # Notify submitter of query via WhatsApp (send to submitter, not Sudhir)
            from src.whatsapp_handler import HOUSEHOLD_MEMBERS, send_message
            submitter = entry.get("submitter", "")
            submitter_number = HOUSEHOLD_MEMBERS.get(submitter.lower())
            msg = (f"❓ Query on your expense request\n"
                   f"Vendor: {entry.get('vendor')} — ₹{entry.get('amount'):,.0f}\n"
                   f"Query: {query_text}\n"
                   f"Ref: {request_id}")
            if submitter_number:
                send_message(submitter_number, msg)
            else:
                send_approval_request(msg)  # fallback to Sudhir if no number on file
    except Exception as ex:
        print(f"[whatsapp] decide notify failed (resp={resp_upper}): {ex}")

    return jsonify({"status": "ok"})


def _approval_to_ledger_entry(e: dict) -> dict:
    """Convert an approval log entry into a master ledger transaction."""
    import hashlib
    from src.master_ledger import _fy_info, _parse_date, classify_transaction
    paid_at = e.get("confirmed_at") or e.get("response_timestamp") or e.get("timestamp","")
    date_str = paid_at[:10]
    dt = _parse_date(date_str) or datetime.now()
    fy = _fy_info(dt)
    amount = float(e.get("approved_amount") or e.get("amount") or 0)
    raw = f"{date_str}|approval|{e.get('vendor','')}|{amount:.2f}"
    txn_id = hashlib.sha1(raw.encode()).hexdigest()[:16]

    _CANONICAL_HEADINGS = {
        "Groceries","Staff Salary","Electricity & Gas","Misc","Cash","Alcohol",
        "Wellness","Clothes","Gifts","Medical","Amma","Ketki","Children Education",
        "Charity","Uspaar","Holiday","Eating Out","Entertainment","Malhar",
        "Maintenance Expense","Home office","One Time Charge","Kalpataru Maintenance",
        "Financial Expense / OD Interest","Insurance","Home Loan","Tax",
        "Short Term Advance","Credit Card Loan","Investment",
    }
    APP_TO_HEADING = {
        "groceries":"Groceries","staff":"Staff Salary","utilities":"Electricity & Gas",
        "miscellaneous":"Misc","personal_care":"Wellness","clothing":"Clothes",
        "gifts":"Gifts","medical":"Medical","education":"Children Education",
        "dining":"Eating Out","entertainment":"Entertainment","transport":"Holiday",
        "maintenance":"Maintenance Expense","home_repair":"One Time Charge",
    }
    # Prefer explicit heading from new form; fall back to category mapping
    explicit_heading = e.get("heading", "")
    cat = e.get("category", "")
    heading = (explicit_heading if explicit_heading in _CANONICAL_HEADINGS
               else cat if cat in _CANONICAL_HEADINGS
               else APP_TO_HEADING.get(cat, "Misc"))

    _pm = (e.get("payment_method") or "cash").lower()
    _is_sbi = _pm in ("sbi", "sbi-4852", "sbi4852", "sbi3152", "sbi-3152", "sbi-3142", "sbi3142")
    _acct   = "SBI-4852prov" if _is_sbi else "cash"
    _bank   = "SBI"          if _is_sbi else "approval"

    txn = {
        "txn_id":          txn_id,
        "date":            date_str,
        "fy_month_no":     fy["fy_month_no"],
        "fy_month_name":   fy["fy_month_name"],
        "fy_year":         fy["fy_year"],
        "account":         _acct,
        "account_type":    _pm,
        "bank":            _bank,
        "raw_description": e.get("description",""),
        "paid_to":         e.get("vendor",""),
        "debit":           amount,
        "credit":          0,
        "type":            "expense",
        "heading":         heading,
        "remarks":         f"Approved by Sudhir. Ref: {e.get('request_id','')}",
        "uncertain":       False,
        "uncertain_fields":[],
        "confidence":      "approval",
        "source":          "approval_log",
        "gmail_id":        None,
        "ai_saving_tip":   None,
        "saving_agreed":   None,
        "reconciled_with": e.get("request_id"),
        "created_at":      datetime.now().isoformat(),
    }
    return txn


def _sync_approvals_to_ledger():
    """Add approved approval entries to master ledger (idempotent).

    SBI payments (payment_method=sbi*) are added immediately as SBI-4852prov
    so they appear in the ledger before the statement arrives.
    Cash/other entries are only added once confirmed_paid=True.
    """
    from src.master_ledger import load_ledger
    log    = db.load("approval_log")
    ledger = db.load("master_ledger")
    existing_ids = {t["txn_id"] for t in ledger}

    added = 0
    for e in log:
        if e.get("action") not in ("AUTO_APPROVE","APPROVED","APPROVED_LOWER"):
            continue
        pm = (e.get("payment_method") or "cash").lower()
        is_sbi = pm in ("sbi", "sbi-4852", "sbi4852", "sbi3152", "sbi-3152", "sbi-3142", "sbi3142")
        # All approved entries sync immediately (SBI as prov, others as cash)
        # Only skip if it's an old-style cash entry with no confirmed_paid
        if not is_sbi and pm == "cash" and not e.get("confirmed_paid"):
            continue
        txn = _approval_to_ledger_entry(e)
        if txn["txn_id"] in existing_ids:
            continue
        ledger.append(txn)
        existing_ids.add(txn["txn_id"])
        added += 1

    if added:
        ledger.sort(key=lambda t: t.get("date",""), reverse=True)
        _assign_missing_seq(ledger)
        db.save("master_ledger", ledger)
    return added


@app.route("/api/mark-paid", methods=["POST"])
def api_mark_paid():
    """Mark a cash expense as confirmed paid and push it into the master ledger."""
    data = request.get_json()
    request_id = data.get("request_id")
    now_iso = datetime.now().isoformat()
    _update_log_entry(request_id, {"confirmed_paid": True, "confirmed_at": now_iso})

    # Push this entry straight into the ledger
    log   = db.load("approval_log")
    entry = next((e for e in log if e.get("request_id") == request_id), None)
    if entry:
        ledger = db.load("master_ledger")
        txn = _approval_to_ledger_entry(entry)
        if not any(t["txn_id"] == txn["txn_id"] for t in ledger):
            ledger.insert(0, txn)
            _assign_missing_seq(ledger)
            db.save("master_ledger", ledger)

    return jsonify({"status": "ok"})


@app.route("/api/cancel-expense", methods=["POST"])
@login_required
def api_cancel_expense():
    """Allow a submitter to cancel their own auto-approved or pending expense."""
    data = request.get_json()
    request_id = data.get("request_id")
    caller = session.get("user", "")

    log = db.load("approval_log")
    entry = next((e for e in log if e.get("request_id") == request_id), None)
    if not entry:
        return jsonify({"error": "not found"}), 404

    # Only the submitter (or sudhir) can cancel
    submitter = entry.get("submitter", "").lower()
    if caller != submitter and caller != "sudhir":
        return jsonify({"error": "not authorised"}), 403

    if entry.get("status") == "cancelled":
        return jsonify({"error": "already cancelled"}), 400

    entry["status"] = "cancelled"
    entry["cancelled_by"] = caller
    entry["cancelled_at"] = datetime.now().isoformat()
    db.save("approval_log", log)

    # Also remove from master ledger if it was synced there
    from src.master_ledger import _fy_info, _parse_date
    import hashlib
    paid_at = entry.get("confirmed_at") or entry.get("response_timestamp") or entry.get("timestamp", "")
    date_str = paid_at[:10]
    amount = float(entry.get("approved_amount") or entry.get("amount") or 0)
    raw = f"{date_str}|approval|{entry.get('vendor','')}|{amount:.2f}"
    txn_id = hashlib.sha1(raw.encode()).hexdigest()[:16]
    ledger = db.load("master_ledger")
    before = len(ledger)
    ledger = [t for t in ledger if t.get("txn_id") != txn_id]
    if len(ledger) < before:
        db.save("master_ledger", ledger)

    return jsonify({"status": "cancelled", "removed_from_ledger": len(ledger) < before})


@app.route("/api/vincent-reply", methods=["POST"])
@login_required
def api_vincent_reply():
    """Vincent answers a query from Sudhir."""
    data = request.get_json()
    request_id = data.get("request_id")
    reply = data.get("reply", "").strip()
    if not reply:
        return jsonify({"error": "empty reply"}), 400

    log = db.load("approval_log")
    entry = next((e for e in log if e.get("request_id") == request_id), None)
    if not entry:
        return jsonify({"error": "not found"}), 404

    caller = session.get("user", "")
    if caller != entry.get("submitter", "").lower() and caller != "sudhir":
        return jsonify({"error": "not authorised"}), 403

    if entry.get("query_state") != "waiting_vincent":
        return jsonify({"error": "no pending query"}), 400

    # Answer the last unanswered query
    queries = entry.get("queries", [])
    for qi in reversed(queries):
        if qi.get("a") is None:
            qi["a"] = reply
            qi["a_time"] = datetime.now().isoformat()
            break
    entry["queries"] = queries
    entry["query_state"] = "waiting_sudhir"
    db.save("approval_log", log)

    # Notify Sudhir via WhatsApp (best-effort)
    try:
        msg = (f"💬 Reply to your query\n"
               f"Vendor: {entry.get('vendor')} — ₹{entry.get('amount'):,.0f}\n"
               f"Query: {queries[-1].get('q','')}\n"
               f"Reply: {reply}\n"
               f"Ref: {request_id}")
        send_to_sudhir(msg)
    except Exception as ex:
        print(f"[whatsapp] vincent-reply notify failed: {ex}")

    return jsonify({"status": "ok"})


@app.route("/api/my-expenses", methods=["GET"])
@login_required
def api_my_expenses():
    """Return the logged-in user's recent expense submissions."""
    caller = session.get("user", "")
    log = db.load("approval_log")
    mine = [e for e in log if e.get("submitter", "").lower() == caller]
    # Return most recent 20, newest first
    mine = sorted(mine, key=lambda e: e.get("timestamp", ""), reverse=True)[:20]
    return jsonify(mine)


@app.route("/api/sync-approvals-to-ledger", methods=["POST"])
@login_required
def api_sync_approvals_to_ledger():
    """Backfill all confirmed-paid approvals into the master ledger."""
    added = _sync_approvals_to_ledger()
    return jsonify({"status": "ok", "added": added})


@app.route("/api/ignore-unauth", methods=["POST"])
def api_ignore_unauth():
    """Mark an unmatched bank debit as known/ignored."""
    data = request.get_json()
    gmail_id = data.get("gmail_id")
    recon = _load_json(RECONCILE_LOG)
    for entry in recon:
        if entry.get("gmail_id") == gmail_id:
            entry["ignored"] = True
            entry["ignored_at"] = datetime.now().isoformat()
            break
    _save_json(RECONCILE_LOG, recon)
    return jsonify({"status": "ok"})


@app.route("/api/sync-statements", methods=["POST"])
@login_required
def api_sync_statements():
    try:
        result = fetch_and_parse_statements()
        return jsonify({"status": "ok", **result})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/transactions/<txn_id>", methods=["PATCH"])
@login_required
def api_update_transaction(txn_id):
    """Update editable fields of an ICICI transaction (paid_to, acc_type, heading, remarks)."""
    data = request.get_json()
    allowed = {"paid_to", "acc_type", "heading", "remarks"}
    updates = {k: v for k, v in data.items() if k in allowed}
    if not updates:
        return jsonify({"error": "no valid fields"}), 400
    transactions = _load_json(TRANSACTIONS_PATH)
    for t in transactions:
        if t.get("txn_id") == txn_id:
            t.update(updates)
            # Mark as manually verified once user edits classification
            if "acc_type" in updates or "heading" in updates:
                t["confidence"] = "manual"
            _save_json(TRANSACTIONS_PATH, transactions)
            return jsonify({"status": "ok"})
    return jsonify({"error": "not found"}), 404


# ── MASTER LEDGER APIs ────────────────────────────────────────────────────────

@app.route("/api/master-ledger", methods=["GET"])
@login_required
def api_master_ledger():
    """Return master ledger entries. Filters: uncertain, account, bank, type, heading, month, q, fy."""
    txns = load_ledger()

    from src.master_ledger import _parse_date as _ml_pd2
    from datetime import date as _date
    date_from = request.args.get("date_from")  # YYYY-MM-DD, overrides FY filter
    date_to   = request.args.get("date_to")
    if date_from or date_to:
        _lo = _date.fromisoformat(date_from) if date_from else _date(2000, 1, 1)
        _hi = _date.fromisoformat(date_to)   if date_to   else _date(2099, 12, 31)
        txns = [t for t in txns if (d := _ml_pd2(t.get("date",""))) and _lo <= d.date() <= _hi]
    else:
        # Fiscal year filter: FY26=Apr2025-Mar2026, FY27=Apr2026-Mar2027 (default)
        fy = request.args.get("fy", "FY27")
        _FY_RANGES = {"FY24": ("2023-04-01","2024-03-31"), "FY25": ("2024-04-01","2025-03-31"), "FY26": ("2025-04-01","2026-03-31"), "FY27": ("2026-04-01","2027-03-31")}
        if fy in _FY_RANGES:
            _fy_lo = _date.fromisoformat(_FY_RANGES[fy][0])
            _fy_hi = _date.fromisoformat(_FY_RANGES[fy][1])
            txns = [t for t in txns if (d := _ml_pd2(t.get("date",""))) and _fy_lo <= d.date() <= _fy_hi]

    # Optional filters
    only_uncertain = request.args.get("uncertain") == "1"
    account_filter = request.args.get("account")
    bank_filter    = request.args.get("bank")
    type_filter    = request.args.get("type")
    heading_filter = request.args.get("heading")
    month_filter   = request.args.get("month")   # YYYY-MM
    search         = request.args.get("q","").lower()

    if only_uncertain:
        txns = [t for t in txns if t.get("uncertain")]
    if account_filter:
        txns = [t for t in txns if t.get("account") == account_filter]
    if bank_filter:
        txns = [t for t in txns if t.get("bank") == bank_filter]
    if type_filter:
        txns = [t for t in txns if (t.get("type") or "").lower() == type_filter.lower()]
    if heading_filter == "__none__":
        txns = [t for t in txns if not (t.get("heading") or "").strip()]
    elif heading_filter:
        txns = [t for t in txns if (t.get("heading") or "") == heading_filter]
    if month_filter:
        txns = [t for t in txns if (t.get("date",""))[:7] == month_filter
                or t.get("date","").endswith(month_filter[-2:] + "/" + month_filter[:4])]
    if search:
        txns = [t for t in txns
                if search in (t.get("raw_description","")).lower()
                or search in (t.get("paid_to","")).lower()
                or search in (t.get("heading") or "").lower()]

    # Summary stats
    total_debit  = sum(t.get("debit",0)  for t in txns)
    total_credit = sum(t.get("credit",0) for t in txns)
    uncertain_ct = sum(1 for t in txns if t.get("uncertain"))

    # Accounts list for filter dropdown
    all_accounts = sorted({t.get("account","") for t in load_ledger() if t.get("account")})

    return jsonify({
        "transactions": txns,
        "total_debit":  total_debit,
        "total_credit": total_credit,
        "uncertain":    uncertain_ct,
        "accounts":     all_accounts,
        "count":        len(txns),
    })


@app.route("/api/master-ledger/sync", methods=["POST"])
@login_required
def api_ledger_sync():
    """Pull new bank alert emails from Gmail only (fast). PDF statements use /sync-statements."""
    data  = request.get_json() or {}
    days  = int(data.get("days", 90))
    force = bool(data.get("force", False))
    result = ledger_sync_gmail(days_back=days, force=force)

    # Fix account names, reclassify unknown entries, remove cross-source duplicates
    repair_pdf_descriptions()
    result["deduped"] = deduplicate_ledger()

    # Reconcile with approval log
    log = _load_json(APPROVAL_LOG)
    matched = reconcile_with_approvals(log)
    result["reconciled"] = matched
    return jsonify({"status": "ok", **result})


_stmt_sync_state = {"running": False, "result": None, "error": None, "started_at": None}


def _run_statement_sync(force: bool):
    """Heavy PDF sync — runs in background thread to avoid request timeout/OOM."""
    result = {}
    try:
        # Purge pre-FY27 pdf_import entries on force
        if force:
            _fy27_start = datetime(2026, 4, 1)
            _ledger = _ml_load_json(LEDGER_PATH)
            _before = len(_ledger)
            _ledger = [
                t for t in _ledger
                if not (
                    t.get("source") == "pdf_import"
                    and (lambda d: d is None or d < _fy27_start)(_ml_parse_date(t.get("date", "")))
                )
            ]
            if len(_ledger) < _before:
                _ml_save_json(LEDGER_PATH, _ledger)
                result["pdf_purged"] = _before - len(_ledger)

        pdf_result = fetch_and_parse_statements(force_reprocess=force)
        result["pdf_statements"] = pdf_result.get("statements", 0)
        result["pdf_new"]        = pdf_result.get("new", 0)

        from src.sbi_statement_parser import fetch_and_parse_sbi_statements
        sbi_result = fetch_and_parse_sbi_statements(force_reprocess=force)
        result["sbi_statements"] = sbi_result.get("statements", 0)
        result["sbi_new"]        = sbi_result.get("new", 0)

        result["pdf_imported"]   = import_from_icici_transactions()
        result["pdf_repaired"]   = repair_pdf_descriptions()
        result["deduped"]        = deduplicate_ledger()
        log = _load_json(APPROVAL_LOG)
        result["reconciled"]     = reconcile_with_approvals(log)
        result["merged"]         = _merge_approval_to_sbi_internal()
        result["status"] = "ok"
    except Exception as e:
        result["status"] = "error"
        result["error"]  = str(e)
    finally:
        _stmt_sync_state["result"]  = result
        _stmt_sync_state["running"] = False


@app.route("/api/master-ledger/sync-statements", methods=["POST"])
@login_required
def api_ledger_sync_statements():
    """Start PDF statement sync in background. Returns immediately — poll /sync-status."""
    if _stmt_sync_state["running"]:
        return jsonify({"status": "already_running", "message": "Sync already in progress — check /api/master-ledger/sync-status"})
    data  = request.get_json() or {}
    force = bool(data.get("force", False))
    _stmt_sync_state["running"]    = True
    _stmt_sync_state["result"]     = None
    _stmt_sync_state["error"]      = None
    _stmt_sync_state["started_at"] = datetime.now().isoformat()
    threading.Thread(target=_run_statement_sync, args=(force,), daemon=True).start()
    return jsonify({"status": "started", "message": "Sync running in background — poll /api/master-ledger/sync-status"})


@app.route("/api/master-ledger/sync-status", methods=["GET"])
@login_required
def api_ledger_sync_status():
    """Check status of background PDF statement sync."""
    return jsonify({
        "running":    _stmt_sync_state["running"],
        "started_at": _stmt_sync_state["started_at"],
        "result":     _stmt_sync_state["result"],
    })


@app.route("/api/master-ledger/fix-7281", methods=["POST"])
@login_required
def api_fix_7281():
    """Directly reclassify all non-manual ICICI-7281 entries by description keywords."""
    INTEREST_KW = ["int.coll", "int coll", "renewal", "sgst", "cgst", "interest", "bank charge", "processing fee"]
    TRANSFER_KW = ["inft", "neft", "imps", "rtgs", "upi", "transfer", "self", "inf/"]

    ledger = _ml_load_json(LEDGER_PATH)
    fixed = 0
    for txn in ledger:
        if txn.get("account") != "ICICI-7281":
            continue
        if txn.get("confidence") == "manual":
            continue
        desc = (txn.get("raw_description") or "").lower()

        # Normalise date format while we're here
        if txn.get("date") and "-" in txn.get("date", ""):
            dt = _ml_parse_date(txn["date"])
            if dt:
                txn["date"] = dt.strftime("%d/%m/%Y")

        if any(k in desc for k in INTEREST_KW):
            txn["type"] = "Expense"
            txn["heading"] = "Financial Expense"
            txn["uncertain"] = False
            fixed += 1
        elif any(k in desc for k in TRANSFER_KW) or not desc or txn.get("heading") in (None, "", "null"):
            txn["type"] = "Transfer"
            txn["heading"] = "Interbank"
            txn["uncertain"] = not desc
            fixed += 1

    if fixed:
        _ml_save_json(LEDGER_PATH, ledger)
    return jsonify({"status": "ok", "fixed": fixed})


@app.route("/api/master-ledger/bulk-classify", methods=["POST"])
@login_required
def api_bulk_classify():
    """Bulk-set type+heading for transactions matching account + date range.
    Skips transfers, income, investments, and manually-confirmed entries.
    Body: {accounts, date_from, date_to, type, heading}
    """
    data       = request.get_json() or {}
    accounts   = [a.upper() for a in data.get("accounts", [])]
    date_from  = _ml_parse_date(data.get("date_from", ""))
    date_to    = _ml_parse_date(data.get("date_to", ""))
    new_type   = data.get("type", "Expense")
    new_heading= data.get("heading", "")

    if not accounts or not date_from or not date_to or not new_heading:
        return jsonify({"error": "accounts, date_from, date_to, heading required"}), 400

    SKIP_TYPES    = {"transfer", "income", "investment", "Transfer", "Income", "Investment"}
    SKIP_HEADINGS = {"Salary", "Interbank", "Home Loan", "Loan Repayment", "Tax",
                     "Financial Expense", "Insurance", "GCPL Share Sale", "Foreign Investment",
                     "Uspaar", "Art", "Property Investment"}

    ledger = _ml_load_json(LEDGER_PATH)
    updated = 0
    for txn in ledger:
        if txn.get("account", "").upper() not in accounts:
            continue
        if txn.get("confidence") == "manual":
            continue
        if txn.get("type") in SKIP_TYPES or txn.get("heading") in SKIP_HEADINGS:
            continue
        dt = _ml_parse_date(txn.get("date", ""))
        if not dt or not (date_from <= dt <= date_to):
            continue
        txn["type"]       = new_type
        txn["heading"]    = new_heading
        txn["confidence"] = "manual"
        txn["uncertain"]  = False
        updated += 1

    if updated:
        _ml_save_json(LEDGER_PATH, ledger)
    return jsonify({"status": "ok", "updated": updated})


@app.route("/api/debug/ledger-accounts", methods=["GET"])
@login_required
def api_debug_ledger_accounts():
    """Show distinct account names and sample descriptions per account."""
    from collections import defaultdict
    ledger = _ml_load_json(LEDGER_PATH)
    by_account = defaultdict(list)
    for t in ledger:
        by_account[t.get("account", "?")].append({
            "desc": t.get("raw_description", "")[:80],
            "type": t.get("type", ""),
            "heading": t.get("heading", ""),
            "source": t.get("source", ""),
        })
    return jsonify({
        acct: {"count": len(txns), "samples": txns[:5]}
        for acct, txns in sorted(by_account.items())
    })


@app.route("/api/debug/seq-lookup", methods=["GET"])
@login_required
def api_debug_seq_lookup():
    """Look up transactions by seq numbers. ?seqs=1,2,6,8"""
    seqs = [int(s.strip()) for s in (request.args.get("seqs") or "").split(",") if s.strip().isdigit()]
    if not seqs:
        return jsonify({"error": "provide ?seqs=1,2,3"}), 400
    ledger = load_ledger()
    hits = {t["seq"]: t for t in ledger if t.get("seq") in seqs}
    return jsonify({"results": [
        {k: hits[s].get(k) for k in ("seq","txn_id","date","account","debit","credit","type","heading","source","confidence","raw_description","paid_to")}
        for s in seqs if s in hits
    ]})


@app.route("/api/debug/search-ledger", methods=["GET"])
@login_required
def api_debug_search_ledger():
    """Search ledger by keyword across description, paid_to, account."""
    q = (request.args.get("q") or "").lower()
    if not q:
        return jsonify({"error": "provide ?q=keyword"}), 400
    ledger = load_ledger()
    hits = [t for t in ledger if any(
        q in str(t.get(f) or "").lower()
        for f in ("raw_description", "paid_to", "account", "heading", "type", "description")
    )]
    hits.sort(key=lambda t: t.get("date", ""))
    return jsonify({"count": len(hits), "results": [
        {k: t.get(k) for k in ("txn_id","date","account","debit","credit","type","heading","source","confidence","raw_description","paid_to","description")}
        for t in hits
    ]})


@app.route("/api/admin/fix-7631", methods=["GET","POST"])
@login_required
def api_fix_7631():
    """Remove all 7631 entries from master ledger + icici_transactions, unmark CC emails for reprocessing."""
    from src import db as _db

    def _is_cc_acct(acct):
        return "7631" in str(acct) or "7009" in str(acct)

    # 1. Remove from master ledger (7631 or 7009 CC entries from pdf_import)
    ledger = load_ledger()
    before_ledger = len(ledger)
    ledger = [t for t in ledger if not (_is_cc_acct(t.get("account", "")) and t.get("source") in ("pdf_import", "icici_statement"))]
    from src.master_ledger import _load_json, _save_json, LEDGER_PATH
    _save_json(LEDGER_PATH, ledger)

    # 2. Remove from icici_transactions (7631 or 7009)
    icici = _db.load("icici_transactions") or []
    before_icici = len(icici)
    icici = [t for t in icici if not _is_cc_acct(t.get("account", ""))]
    _db.save("icici_transactions", icici)

    # 3. Unmark CC statement emails so they get re-parsed by Sync PDF
    CC_MSG_IDS = {
        "19ed2026b5009714",  # CCStatement_Current17-06-2026
        "19ecb0676fed14a6",  # Monthlystatement_19 Apr-18 May
        "19ecbbedc9cece89",  # Fwd: Monthlystatement_19 Apr-18 May
        "19ecb05c7c3d4452",  # Monthlystatement_19 Mar-18 Apr
    }
    processed = set(_db.load("processed_statement_ids", default=[]))
    processed -= CC_MSG_IDS
    _db.save("processed_statement_ids", list(processed))

    return jsonify({
        "ledger_removed": before_ledger - len(ledger),
        "icici_txns_removed": before_icici - len(icici),
        "emails_unmarked": len(CC_MSG_IDS),
        "message": "Done — now run Sync PDF to reimport correctly as ICICI-7009",
    })


@app.route("/api/debug/parse-cc", methods=["GET"])
@login_required
def api_debug_parse_cc():
    """Debug: re-fetch CC emails and show what each parser returns (first 5 txns + counts)."""
    from src.icici_statement_parser import (
        _get_pdf_attachments, _open_pdf,
        _parse_cc_statement_text, _parse_savings_statement_text,
        _parse_od_savings_text, _parse_from_text,
        _extract_account_from_text,
    )
    from googleapiclient.discovery import build
    from src.gmail_utils import get_credentials
    service = build("gmail", "v1", credentials=get_credentials())

    CC_MSG_IDS = [
        "19ed2026b5009714",
        "19ecb0676fed14a6",
        "19ecbbedc9cece89",
        "19ecb05c7c3d4452",
    ]
    results = []
    for msg_id in CC_MSG_IDS:
        try:
            pdfs = _get_pdf_attachments(service, msg_id)
            for pdf_bytes in pdfs:
                try:
                    with _open_pdf(pdf_bytes) as pdf:
                        full_text = ""
                        for page in pdf.pages:
                            full_text += (page.extract_text() or "") + "\n"
                    acct = _extract_account_from_text(full_text)
                    cc = _parse_cc_statement_text(full_text)
                    sv = _parse_savings_statement_text(full_text)
                    od_sv = _parse_od_savings_text(full_text)
                    od = _parse_from_text(full_text)
                    best = max([cc, sv, od_sv, od], key=len)
                    results.append({
                        "msg_id": msg_id,
                        "account_detected": acct,
                        "cc_parser": len(cc),
                        "savings_parser": len(sv),
                        "od_savings_parser": len(od_sv),
                        "od_parser": len(od),
                        "winner": "cc" if best is cc else "savings" if best is sv else "od_savings" if best is od_sv else "od",
                        "winner_count": len(best),
                        "first_5_winner": best[:5],
                        "text_sample": full_text[:500],
                    })
                except Exception as e:
                    results.append({"msg_id": msg_id, "error": str(e)})
        except Exception as e:
            results.append({"msg_id": msg_id, "fetch_error": str(e)})
    return jsonify(results)


@app.route("/api/account-status", methods=["GET"])
@login_required
def api_account_status():
    """Per-account, per-FY-month transaction counts. FY27 = Apr 2026 onwards."""
    from src.master_ledger import _parse_date as _ml_parse_date
    now = datetime.now()
    # Build list of FY27 months from Apr 2026 up to current month
    fy_months = []
    yr, mo = 2026, 4
    while (yr, mo) <= (now.year, now.month):
        fy_months.append(f"{yr}-{mo:02d}")
        mo += 1
        if mo == 13:
            mo, yr = 1, yr + 1

    ledger = load_ledger()
    accounts = {}  # acct -> {meta, months: {ym: count}, latest_dt, latest_date, latest_balance}
    for txn in ledger:
        acct = txn.get("account") or "Unknown"
        dt = _ml_parse_date(txn.get("date", ""))
        if not dt:
            continue
        ym = dt.strftime("%Y-%m")
        if acct not in accounts:
            accounts[acct] = {
                "account": acct,
                "bank": txn.get("bank", ""),
                "account_type": txn.get("account_type", ""),
                "months": {},
                "latest_dt": dt,
                "latest_date": txn.get("date", ""),
                "latest_balance": txn.get("balance"),
            }
        accounts[acct]["months"][ym] = accounts[acct]["months"].get(ym, 0) + 1
        if dt >= accounts[acct]["latest_dt"]:
            accounts[acct]["latest_dt"] = dt
            accounts[acct]["latest_date"] = txn.get("date", "")
            if txn.get("balance") is not None:
                accounts[acct]["latest_balance"] = txn.get("balance")

    result = sorted(accounts.values(), key=lambda x: x["account"])
    for r in result:
        r["fy_months"] = fy_months
        r["month_counts"] = [r["months"].get(ym, 0) for ym in fy_months]
        del r["months"]
        del r["latest_dt"]
    return jsonify({"accounts": result, "fy_months": fy_months})


@app.route("/api/debug/mis-actuals", methods=["GET"])
@login_required
def api_debug_mis_actuals():
    """Show what the MIS endpoint aggregates from the master ledger for FY27."""
    from src.master_ledger import _parse_date as _ml_parse_date
    from datetime import datetime
    _fy27_start = datetime(2026, 4, 1)
    fy27_actual = {}
    skipped = []
    for txn in load_ledger():
        dt = _ml_parse_date(txn.get("date", ""))
        if not dt or dt < _fy27_start:
            continue
        reason = None
        if (txn.get("type") or "").lower() not in ("expense", "official"):
            reason = f"type={txn.get('type')}"
        elif txn.get("uncertain"):
            reason = "uncertain=True"
        elif not float(txn.get("debit", 0) or 0):
            reason = "debit=0"
        if reason:
            skipped.append({"seq": txn.get("seq"), "date": txn.get("date"), "heading": txn.get("heading"), "type": txn.get("type"), "uncertain": txn.get("uncertain"), "debit": txn.get("debit"), "reason": reason})
        else:
            h = txn.get("heading") or "Misc"
            fy27_actual[h] = fy27_actual.get(h, 0) + float(txn.get("debit", 0))
    return jsonify({"included": fy27_actual, "skipped_count": len(skipped), "skipped_sample": skipped[:30]})


@app.route("/api/ledger/<txn_id>", methods=["DELETE"])
@login_required
def api_ledger_delete(txn_id):
    """Delete an SBI entry from the master ledger."""
    ledger = db.load("master_ledger") or []
    entry = next((t for t in ledger if t.get("txn_id") == txn_id), None)
    if not entry:
        return jsonify({"error": "not found"}), 404
    if not (entry.get("account") or "").upper().startswith("SBI"):
        return jsonify({"error": "only SBI entries can be deleted"}), 403
    ledger = [t for t in ledger if t.get("txn_id") != txn_id]
    db.save("master_ledger", ledger)
    return jsonify({"ok": True})


@app.route("/api/master-ledger/<txn_id>", methods=["PATCH"])
@login_required
def api_ledger_update(txn_id):
    """Update type, heading, paid_to, remarks, saving_agreed for a transaction."""
    data    = request.get_json() or {}
    allowed = {"account","paid_to","type","heading","remarks","saving_agreed","project"}
    updates = {k: v for k, v in data.items() if k in allowed}
    if not updates:
        return jsonify({"error": "no valid fields"}), 400
    ok = update_transaction(txn_id, updates)
    return jsonify({"status": "ok"}) if ok else (jsonify({"error": "not found"}), 404)


@app.route("/api/admin/set-sbi-password", methods=["GET"])
@login_required
def api_set_sbi_password():
    """Store SBI PDF password in DB. Usage: ?pw=YOUR_PASSWORD"""
    pw = request.args.get("pw", "")
    if not pw:
        return jsonify({"error": "provide ?pw=your_password"}), 400
    db.save("sbi_pdf_password", pw)
    return jsonify({"message": "SBI PDF password saved to DB", "length": len(pw)})


@app.route("/api/admin/reset-sbi", methods=["GET"])
@login_required
def api_reset_sbi():
    """Unmark all SBI statement emails so they get re-processed on next Sync PDF."""
    from src import db as _db
    processed = set(_db.load("processed_statement_ids", default=[]))
    # Remove any SBI transactions from icici_transactions store
    icici = _db.load("icici_transactions") or []
    before = len(icici)
    icici = [t for t in icici if t.get("bank") != "SBI" and not str(t.get("account","")).startswith("SBI")]
    _db.save("icici_transactions", icici)
    # Unmark all processed IDs that came from SBI emails by clearing the full set
    # and re-adding only non-SBI ones — we do this by removing IDs added in the last SBI run
    # Simpler: just rebuild without SBI — but we don't know which IDs are SBI.
    # So we rely on the parser NOT marking on password failure going forward.
    # For now, wipe all processed IDs so everything gets retried (safe — dedup prevents doubles).
    # Only wipe if user confirms by passing ?confirm=yes
    from flask import request as _req
    if _req.args.get("confirm") == "yes":
        # Clear only recently added IDs (risky); safer: just clear all and let dedup handle it
        _db.save("processed_statement_ids", [])
        return jsonify({"message": "All processed statement IDs cleared — Sync PDF will retry everything", "sbi_txns_removed": before - len(icici)})
    return jsonify({"message": "Pass ?confirm=yes to clear all processed statement IDs", "sbi_txns_removed": before - len(icici)})


@app.route("/api/admin/repair-sbi", methods=["GET"])
@login_required
def api_repair_sbi():
    """
    Fix debit/credit direction on existing SBI transactions using stored balance sequence.
    Also parses paid_to from description and cleans newlines.
    No re-import needed — operates on master ledger in-place.
    """
    from src.master_ledger import _load_json, _save_json, LEDGER_PATH, _parse_date
    from src.sbi_statement_parser import _parse_paid_to
    import re as _re

    _SBI_BF_KEYWORDS = (
        "brought forward", "opening bal", "closing bal", "b/f balance", "b/f",
        "brought fwd", "b/f bal", "opening balance", "closing balance",
    )

    ledger = _load_json(LEDGER_PATH)

    # Remove Brought Forward / summary entries — check all text fields
    before_len = len(ledger)
    def _is_bf(t):
        if not str(t.get("account", "")).startswith("SBI"):
            return False
        if t.get("confidence") == "manual":
            return False
        text = " ".join(str(t.get(f) or "") for f in
                        ("transaction_details", "description", "paid_to", "vendor")).lower()
        return any(kw in text for kw in _SBI_BF_KEYWORDS)
    ledger = [t for t in ledger if not _is_bf(t)]
    removed = before_len - len(ledger)

    # Group SBI transactions, sorted by date then seq
    sbi_txns = [t for t in ledger
                if str(t.get("account", "")).startswith("SBI")
                and t.get("confidence") != "manual"]
    sbi_txns.sort(key=lambda t: (
        _parse_date(t.get("date", "")) or datetime.min,
        t.get("seq", 0)
    ))

    fixed = 0
    paid_to_samples = []

    for i, txn in enumerate(sbi_txns):
        changed = False

        amount = float(txn.get("debit") or txn.get("credit") or txn.get("amount") or 0)
        if amount == 0:
            continue

        # ── Direction fix — primary: description prefix (100% reliable) ──────
        # SBI always prefixes WDL TFR (withdrawal=debit) or DEP TFR (deposit=credit)
        # ATM withdrawals are always debits regardless of prefix
        raw_for_dir = (txn.get("raw_description") or "").strip().upper()
        _ATM_KW = ["ATM", "CASH WD", "ATW ", "CASH WTHDL", "ATM WTDL", "CASH WITHDRAWAL"]
        if raw_for_dir.startswith("WDL") or any(raw_for_dir.startswith(k) for k in _ATM_KW) \
                or any(k in raw_for_dir for k in _ATM_KW):
            correct_dir = "debit"
        elif raw_for_dir.startswith("DEP"):
            correct_dir = "credit"
        else:
            # Fallback: balance delta (unreliable for same-day; only use when confident)
            correct_dir = None
            bal = txn.get("balance")
            if bal is not None:
                acct = txn.get("account")
                for j in range(i - 1, -1, -1):
                    if sbi_txns[j].get("account") == acct and sbi_txns[j].get("balance") is not None:
                        delta = float(bal) - float(sbi_txns[j]["balance"])
                        correct_dir = "debit" if delta < 0 else "credit"
                        break

        if correct_dir:
            cur_d = float(txn.get("debit") or 0)
            cur_c = float(txn.get("credit") or 0)
            is_wrong = (correct_dir == "debit"  and cur_d == 0 and cur_c > 0) or \
                       (correct_dir == "credit" and cur_c == 0 and cur_d > 0)
            if is_wrong:
                txn["debit"]  = amount if correct_dir == "debit"  else 0
                txn["credit"] = amount if correct_dir == "credit" else 0
                changed = True

        # ── Normalise invalid type values left by old repair runs ────────────
        cur_type = (txn.get("type") or "").lower()
        if cur_type in ("debit", "credit", ""):
            raw_up = (txn.get("raw_description") or "").strip().upper()
            _ATM_TYPE_KW = ["ATM", "CASH WD", "ATW ", "CASH WTHDL", "ATM WTDL", "CASH WITHDRAWAL"]
            is_atm_entry = any(k in raw_up for k in _ATM_TYPE_KW)
            if is_atm_entry:
                # ATM withdrawals are short-term advances to Vincent until he logs cash spends
                new_type = "investment"
                new_head = "Loans"
            elif raw_up.startswith("WDL"):
                new_type = "expense"
                new_head = None
            elif raw_up.startswith("DEP"):
                new_type = "transfer"
                new_head = None
            else:
                new_type = "expense" if float(txn.get("debit") or 0) > 0 else "transfer"
                new_head = None
            if new_type != cur_type:
                txn["type"] = new_type
                if new_head:
                    txn["heading"] = new_head
                changed = True

        # ── Description cleaning (raw_description is the field SBI entries use) ─
        raw_desc   = txn.get("raw_description") or txn.get("transaction_details") or txn.get("description") or ""
        clean_desc = _re.sub(r"\s+", " ", raw_desc).strip()
        if clean_desc != raw_desc:
            txn["raw_description"] = clean_desc
            changed = True

        # ── paid_to — always recompute and overwrite ──────────────────────────
        paid_to = _parse_paid_to(clean_desc) if clean_desc else None
        if paid_to:
            if paid_to != txn.get("paid_to"):
                txn["paid_to"] = paid_to
                changed = True
            if len(paid_to_samples) < 5:
                paid_to_samples.append({"desc": clean_desc[:60], "paid_to": paid_to})

        if changed:
            fixed += 1

    if fixed or removed:
        _save_json(LEDGER_PATH, ledger)

    return jsonify({
        "fixed": fixed,
        "bf_removed": removed,
        "total_sbi": len(sbi_txns),
        "paid_to_samples": paid_to_samples,
    })


def _merge_approval_to_sbi_internal() -> dict:
    """
    For every approval_log entry in the master ledger find its matching SBI-3152
    statement entry, enrich the SBI entry with the approval's vendor/heading/description,
    and remove the approval_log duplicate.  Called automatically on every SBI sync.

    Match rules:
      sbi3152 payment → SBI non-ATM debit, amount ±₹100, date ±7 days
      cash payment    → SBI ATM debit, amount ≥ 90 % of approval, date ±30 days
    Unmatched approval entries are left untouched.
    """
    from src.master_ledger import _parse_date as _ml_pd, _save_json, LEDGER_PATH, _APP_TO_HEADING
    from datetime import datetime as _dt

    _SBI_CASH_KW = ["atm", "cash withdrawal", "atw", "cash wthdl", "atm wtdl"]

    ledger = load_ledger()

    approval_entries = [t for t in ledger if t.get("source") == "approval_log"
                        and "4852prov" in (t.get("account") or "").lower()]
    sbi_entries      = [t for t in ledger if "4852" in (t.get("account") or "")
                        and t.get("source") != "approval_log"
                        and float(t.get("debit") or 0) > 0]

    to_remove_ids = set()
    merged = 0

    def _match_score(prov_amt, sbi_amt, days):
        """Lower is better. Days dominate; amount diff breaks ties."""
        return days * 10000 + abs(prov_amt - sbi_amt)

    # Build all candidate (score, appr, sbi) pairs then assign optimally
    candidates = []
    for appr in approval_entries:
        pm = (appr.get("account_type") or appr.get("payment_method") or "").lower()
        log_amt = float(appr.get("debit") or appr.get("amount") or 0)
        if log_amt == 0:
            continue
        try:
            log_date = _ml_pd(appr.get("date", "")) or _dt.fromisoformat(
                (appr.get("created_at") or "")[:10])
        except Exception:
            continue
        if not log_date:
            continue
        for sbi in sbi_entries:
            sbi_amt  = float(sbi.get("debit") or 0)
            if sbi_amt == 0:
                continue
            sbi_date = _ml_pd(sbi.get("date", ""))
            if not sbi_date:
                continue
            diff = abs(sbi_amt - log_amt)
            if diff > 200 or diff / max(sbi_amt, log_amt, 1) > 0.05:
                continue
            days_diff = abs((sbi_date - log_date).days)
            if days_diff > 7:
                continue
            raw = (sbi.get("raw_description") or sbi.get("transaction_details") or "").lower()
            if any(kw in raw for kw in _SBI_CASH_KW):
                continue
            candidates.append((_match_score(log_amt, sbi_amt, days_diff), appr, sbi))

    candidates.sort(key=lambda x: x[0])
    matched_appr_ids = set()
    matched_sbi_ids  = set()
    best_matches = []
    for _, appr, sbi in candidates:
        if appr["txn_id"] in matched_appr_ids:
            continue
        if sbi["txn_id"] in matched_sbi_ids:
            continue
        matched_appr_ids.add(appr["txn_id"])
        matched_sbi_ids.add(sbi["txn_id"])
        best_matches.append((appr, sbi))

    for appr, best in best_matches:
        # Idempotent — already merged from this exact approval
        if best.get("merged_from_approval") == appr.get("txn_id"):
            to_remove_ids.add(appr["txn_id"])
            continue

        cat     = appr.get("category") or appr.get("account_type") or "miscellaneous"
        # ── From prov entry (Vincent's data) — always wins ──────────────────
        best["paid_to"]   = appr.get("paid_to") or appr.get("vendor") or best.get("paid_to")
        best["type"]      = appr.get("type") or "personal"
        best["heading"]   = (appr.get("heading")
                             or _APP_TO_HEADING.get(cat, best.get("heading", "Misc")))
        best["description"] = appr.get("description") or appr.get("remarks") or best.get("description")
        # ── Bookkeeping fields ───────────────────────────────────────────────
        best["submitter"]            = appr.get("submitter") or best.get("submitter")
        best["request_id"]           = appr.get("request_id") or best.get("request_id")
        best["approval_vendor"]      = appr.get("paid_to") or appr.get("vendor")
        best["reconciled_with"]      = appr.get("reconciled_with") or appr.get("txn_id")
        best["merged_from_approval"] = appr.get("txn_id")
        best["uncertain"]            = False
        best["uncertain_fields"]     = []
        best["confidence"]           = "merged"
        best["source"]               = "approval_log"
        to_remove_ids.add(appr["txn_id"])
        merged += 1

    # ── Pass 2: match uncertain SBI-4852 statement entries directly against ──────
    # approval_log (no prov entry was ever created for these).
    # Uses a wider 30-day window because Vincent submits approvals weekly.
    _SBI_PMS = ("sbi","sbi-4852","sbi4852","sbi3152","sbi-3152","sbi-3142","sbi3142")
    log = db.load("approval_log")
    unreconciled_approvals = [
        e for e in log
        if e.get("action") in ("AUTO_APPROVE","APPROVED","APPROVED_LOWER")
        and (e.get("payment_method") or "").lower() in _SBI_PMS
        and not e.get("reconciled_to")
        and f"appr_{e.get('request_id','')}" not in {t.get("txn_id") for t in ledger}
    ]

    uncertain_stmt = [
        t for t in ledger
        if "4852" in (t.get("account") or "")
        and "prov" not in (t.get("account") or "").lower()
        and t.get("uncertain")
        and not t.get("merged_from_approval")
        and float(t.get("debit") or 0) > 0
    ]

    cands2 = []
    for stmt in uncertain_stmt:
        s_amt  = float(stmt.get("debit") or 0)
        s_date = _ml_pd(stmt.get("date",""))
        if not s_date: continue
        raw = (stmt.get("raw_description") or stmt.get("transaction_details") or "").lower()
        if any(kw in raw for kw in _SBI_CASH_KW): continue
        for e in unreconciled_approvals:
            e_amt  = float(e.get("approved_amount") or e.get("amount") or 0)
            e_date = _ml_pd((e.get("timestamp") or "")[:10])
            if not e_date: continue
            diff = abs(s_amt - e_amt)
            if diff > 200 or diff / max(s_amt, e_amt, 1) > 0.05: continue
            days = abs((s_date - e_date).days)
            if days > 30: continue
            cands2.append((_match_score(s_amt, e_amt, days), stmt, e))

    cands2.sort(key=lambda x: x[0])
    matched_stmt2 = set(); matched_appr2 = set()
    direct_merged = 0
    for _, stmt, e in cands2:
        if stmt["txn_id"] in matched_stmt2: continue
        if e.get("request_id","") in matched_appr2: continue
        matched_stmt2.add(stmt["txn_id"])
        matched_appr2.add(e.get("request_id",""))
        cat = e.get("category") or e.get("account_type") or "miscellaneous"
        stmt["paid_to"]              = e.get("paid_to") or e.get("vendor") or stmt.get("paid_to")
        stmt["type"]                 = e.get("type") or "personal"
        stmt["heading"]              = (e.get("heading")
                                        or _APP_TO_HEADING.get(cat, stmt.get("heading","Misc")))
        stmt["description"]          = e.get("description") or e.get("remarks") or stmt.get("description")
        stmt["submitter"]            = e.get("submitter") or stmt.get("submitter")
        stmt["request_id"]           = e.get("request_id") or stmt.get("request_id")
        stmt["approval_vendor"]      = e.get("paid_to") or e.get("vendor")
        stmt["reconciled_with"]      = e.get("request_id")
        stmt["merged_from_approval"] = f"appr_{e.get('request_id','')}"
        stmt["uncertain"]            = False
        stmt["uncertain_fields"]     = []
        stmt["confidence"]           = "merged"
        stmt["source"]               = "approval_log"
        direct_merged += 1

    if to_remove_ids or direct_merged:
        ledger = [t for t in ledger if t.get("txn_id") not in to_remove_ids]
        _save_json(LEDGER_PATH, ledger)

    return {
        "merged": merged,
        "approval_entries_removed": len(to_remove_ids),
        "direct_merged": direct_merged,
        "unmatched_approvals_kept": len(approval_entries) - merged,
    }


@app.route("/api/admin/migrate-cash-upi-to-sbi", methods=["GET"])
@login_required
def api_migrate_cash_upi_to_sbi():
    """One-time: reclassify all account='cash/upi' entries to SBI-3152."""
    from src.master_ledger import _save_json, LEDGER_PATH
    ledger  = load_ledger()
    updated = 0
    for txn in ledger:
        if (txn.get("account") or "").lower() == "cash/upi":
            txn["account"]      = "SBI-4852"
            txn["bank"]         = "SBI"
            txn["source"]       = "sbi_statement"
            txn["account_type"] = "sbi4852"
            updated += 1
    if updated:
        _save_json(LEDGER_PATH, ledger)
    return jsonify({"migrated": updated})


@app.route("/api/admin/merge-approval-to-sbi", methods=["GET"])
@login_required
def api_merge_approval_to_sbi():
    result = _merge_approval_to_sbi_internal()
    result["total_sbi"] = sum(
        1 for t in load_ledger()
        if "SBI" in (t.get("account") or "").upper() and t.get("source") == "sbi_statement"
    )
    return jsonify(result)


@app.route("/api/admin/bulk-holiday-7009", methods=["GET"])
@login_required
def api_bulk_holiday_7009():
    """Set all ICICI-7009 transactions between 23-May and 1-Jun to expense/Holiday."""
    from src.master_ledger import _parse_date as _ml_parse_date, _save_json, LEDGER_PATH
    from datetime import datetime
    date_from = datetime(2026, 5, 23)
    date_to   = datetime(2026, 6, 1)
    ledger = load_ledger()
    updated = 0
    for txn in ledger:
        if (txn.get("account") or "").endswith("7009"):
            dt = _ml_parse_date(txn.get("date", ""))
            if dt and date_from <= dt <= date_to:
                txn["type"]    = "Expense"
                txn["heading"] = "Holiday"
                txn["confidence"] = "manual"
                updated += 1
    _save_json(LEDGER_PATH, ledger)
    return jsonify({"updated": updated, "message": f"Set {updated} ICICI-7009 txns (23-May to 1-Jun) to Expense/Holiday"})


@app.route("/api/cc-balance", methods=["GET"])
@login_required
def api_cc_balance():
    """Credit card unpaid balances per card."""
    return jsonify(get_cc_balance())


@app.route("/api/approval-log-raw", methods=["GET"])
@login_required
def api_approval_log_raw():
    """Debug: return last 30 raw entries from approval log."""
    log = _load_json(APPROVAL_LOG)
    return jsonify({"total": len(log), "last_30": log[-30:]})


@app.route("/api/db-status", methods=["GET"])
@login_required
def api_db_status():
    """Debug: show DB connection health and stored keys."""
    return jsonify(db.db_status())


@app.route("/api/approvals-structured", methods=["GET"])
@login_required
def api_approvals_structured():
    """Return approval log structured into pending / month / unauthorized / tracker."""
    log   = _load_json(APPROVAL_LOG)
    recon = _load_json(RECONCILE_LOG)

    now          = datetime.now()
    # Allow caller to pass ?month=YYYY-MM; default to current month
    month_prefix = request.args.get("month") or now.strftime("%Y-%m")

    approved_actions = ("AUTO_APPROVE", "APPROVED", "APPROVED_LOWER")

    def _effective_month(e):
        """Use approval date for approved entries; submission date for auto-approvals."""
        if e.get("action") in ("APPROVED", "APPROVED_LOWER"):
            return (e.get("response_timestamp") or e.get("timestamp",""))[:7]
        return (e.get("timestamp",""))[:7]

    today_prefix = now.strftime("%Y-%m")
    # All months that have at least one approved entry (for the month picker), not in the future
    months_with_data = sorted({
        _effective_month(e)
        for e in log
        if (e.get("action") in approved_actions
            or (e.get("action") == "ESCALATE" and "sudhir_response" in e))
        and len(_effective_month(e)) == 7
        and _effective_month(e) <= today_prefix
    }, reverse=True)

    def _is_pending(e):
        """Sudhir needs to act: fresh ESCALATE, OR Vincent has answered a query."""
        if e.get("status") == "cancelled":
            return False
        qs = e.get("query_state")
        if qs == "waiting_vincent":
            return False   # ball is with Vincent
        if qs == "waiting_sudhir":
            return True    # Vincent replied, Sudhir must act
        # No active query: pending only if ESCALATE with no final response
        action = e.get("action", "")
        if action == "ESCALATE" and "sudhir_response" not in e:
            return True
        return False

    def _is_approved(e):
        if e.get("status") == "cancelled":
            return False
        if e.get("query_state") in ("waiting_vincent", "waiting_sudhir"):
            return False   # in query limbo — show in pending
        return e.get("action") in approved_actions

    pending      = [e for e in log if _is_pending(e)]

    # Approved: last 90 days
    from datetime import timedelta
    cutoff = (now - timedelta(days=90)).isoformat()
    approved = sorted(
        [e for e in log if _is_approved(e) and (e.get("timestamp","") >= cutoff or e.get("response_timestamp","") >= cutoff)],
        key=lambda e: e.get("response_timestamp") or e.get("timestamp",""),
        reverse=True
    )

    this_month   = [e for e in log
                    if _effective_month(e) == month_prefix
                    and e.get("status") != "cancelled"
                    and (e.get("action") in approved_actions
                         or (e.get("action") == "ESCALATE" and "sudhir_response" in e))]
    unauthorized = [e for e in recon if not e.get("matched") and not e.get("is_recurring") and not e.get("ignored")]
    tracker      = [e for e in log
                    if _is_approved(e)
                    and not e.get("confirmed_paid")]

    return jsonify({
        "pending":           pending,
        "approved":          approved,
        "this_month":        this_month,
        "selected_month":    month_prefix,
        "months_with_data":  months_with_data,
        "unauthorized":      unauthorized,
        "tracker":           tracker,
    })


@app.route("/api/insights", methods=["GET"])
@login_required
def api_insights():
    """AI-generated spending insights from master ledger + approval log."""
    try:
        from openai import AzureOpenAI
        client = AzureOpenAI(
            api_key=os.getenv("AZURE_OPENAI_KEY"),
            azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),
            api_version=os.getenv("AZURE_OPENAI_API_VERSION","2024-12-01-preview"),
        )
        deployment = os.getenv("AZURE_OPENAI_DEPLOYMENT","gpt-5.5")
    except Exception:
        return jsonify({"insights":[]})

    ledger = load_ledger()
    now    = datetime.now()
    month_prefix = now.strftime("%Y-%m")

    # Last 3 months summary by heading
    from collections import defaultdict
    by_heading: dict = defaultdict(float)
    for t in ledger:
        if t.get("debit") and t.get("heading"):
            by_heading[t["heading"]] += t["debit"]

    # Budget from file
    budget_path = os.path.join(CONFIG_DIR, "budget_fy27.json")
    with open(budget_path) as f:
        budget_annual = json.load(f).get("annual",{})

    summary = [
        {"heading": h, "spend": round(v), "budget": budget_annual.get(h,0)}
        for h, v in sorted(by_heading.items(), key=lambda x: -x[1])[:15]
    ]

    prompt = f"""You are a personal finance advisor for an Indian household (Mumbai, upper-income).
Analyse this spending data and give 5 concise, actionable insights to reduce expenses.
Focus on high-spend categories vs budget, patterns worth questioning, and concrete alternatives.

Spending summary (from bank transactions): {json.dumps(summary)}

Reply as JSON array of 5 objects: [{{"title":"...", "detail":"...", "category":"...", "potential_saving":"₹X,XXX/month"}}]
Be specific, not generic. No obvious tips."""

    try:
        resp = client.chat.completions.create(
            model=deployment, max_completion_tokens=800,
            messages=[{"role":"system","content":"Reply only with JSON."},
                      {"role":"user","content":prompt}]
        )
        text = resp.choices[0].message.content.strip()
        if text.startswith("```"):
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
        insights = json.loads(text)
    except Exception as e:
        insights = [{"title":"Unable to generate insights","detail":str(e),"category":"","potential_saving":""}]

    return jsonify({"insights": insights})


@app.route("/api/insights-chat", methods=["POST"])
@login_required
def api_insights_chat():
    """Free-form chat about expense data using Azure OpenAI."""
    try:
        from openai import AzureOpenAI
        client = AzureOpenAI(
            api_key=os.getenv("AZURE_OPENAI_KEY"),
            azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),
            api_version=os.getenv("AZURE_OPENAI_API_VERSION","2024-12-01-preview"),
        )
        deployment = os.getenv("AZURE_OPENAI_DEPLOYMENT","gpt-5.5")
    except Exception as e:
        return jsonify({"reply": f"AI not configured: {e}"}), 500

    data = request.get_json()
    question = data.get("question","").strip()
    history  = data.get("history", [])  # [{role, content}, ...]
    if not question:
        return jsonify({"reply": ""}), 400

    # Build data context
    ledger = load_ledger()
    from collections import defaultdict

    # account × heading × YYYY-MM → spend
    by_heading: dict                          = defaultdict(float)
    by_month_heading: dict                    = defaultdict(lambda: defaultdict(float))
    vendors: dict                             = defaultdict(float)
    by_account: dict                          = defaultdict(float)
    by_account_heading: dict                  = defaultdict(lambda: defaultdict(float))
    by_account_month: dict                    = defaultdict(lambda: defaultdict(float))
    by_account_heading_month: dict            = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    # FY buckets: "FY2627" etc
    by_fy: dict                               = defaultdict(float)
    by_account_fy: dict                       = defaultdict(lambda: defaultdict(float))
    by_account_fy_heading: dict               = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

    _MON = {"jan":"01","feb":"02","mar":"03","apr":"04","may":"05","jun":"06",
            "jul":"07","aug":"08","sep":"09","oct":"10","nov":"11","dec":"12"}

    def _to_ym(raw: str) -> str:
        """Normalise any date string to YYYY-MM, return '' on failure."""
        if not raw:
            return ""
        try:
            if raw[2:3] == "/" and raw[5:6] == "/":   # DD/MM/YYYY
                parts = raw.split("/")
                return f"{parts[2][:4]}-{parts[1].zfill(2)}"
            if raw[2:3] == "-" and not raw[3:4].isdigit():  # DD-Mon-YY or DD-Mon-YYYY
                parts = raw.split("-")
                mon = _MON.get(parts[1][:3].lower(), "")
                yr  = parts[2][:4] if len(parts[2]) == 4 else "20" + parts[2][:2]
                return f"{yr}-{mon}" if mon else ""
            if raw[4:5] == "-":                        # YYYY-MM-DD
                return raw[:7]
        except Exception:
            pass
        return ""

    def _fy(mo: str) -> str:
        """YYYY-MM → 'FY2526' style label."""
        if not mo or len(mo) < 7:
            return "unknown"
        try:
            y, m = int(mo[:4]), int(mo[5:7])
        except ValueError:
            return "unknown"
        if m >= 4:
            return f"FY{str(y)[2:]}{str(y+1)[2:]}"
        return f"FY{str(y-1)[2:]}{str(y)[2:]}"

    for t in ledger:
        raw_date = t.get("date") or t.get("timestamp","")
        mo   = _to_ym(raw_date)
        acct = t.get("account","unknown")
        fy   = _fy(mo)
        hdg  = (t.get("heading") or "").strip().title()  # normalise case variants

        if t.get("debit"):
            amt = float(t["debit"])
            by_account[acct] += amt
            by_fy[fy]         += amt
            by_account_fy[acct][fy] += amt
            if mo:
                by_account_month[acct][mo] += amt
            if hdg:
                by_heading[hdg] += amt
                by_account_heading[acct][hdg] += amt
                by_account_fy_heading[acct][fy][hdg] += amt
                if mo:
                    by_month_heading[mo][hdg]                   += amt
                    by_account_heading_month[acct][hdg][mo]     += amt
        if t.get("debit") and t.get("paid_to"):
            vendors[t["paid_to"]] += float(t["debit"])

    # Recent 12 months (wider window for FY questions)
    all_months = sorted(by_month_heading.keys())
    recent_months = all_months[-12:]
    monthly_data = {m: {h: round(v) for h, v in by_month_heading[m].items()} for m in recent_months}
    top_vendors = dict(sorted(vendors.items(), key=lambda x: -x[1])[:20])

    # Per-account rich context
    accounts_context = {}
    for acct, total in sorted(by_account.items(), key=lambda x: -x[1]):
        top_cats = {h: round(v) for h, v in sorted(by_account_heading[acct].items(), key=lambda x: -x[1])[:12]}
        # monthly totals — last 12 months
        recent_mo = {m: round(by_account_month[acct][m]) for m in sorted(by_account_month[acct])[-12:]}
        # per-FY totals for this account
        fy_totals = {fy: round(v) for fy, v in sorted(by_account_fy[acct].items())}
        # per-FY per-heading for this account (top 10 headings per FY)
        fy_heading = {}
        for fy, hdg_map in by_account_fy_heading[acct].items():
            fy_heading[fy] = {h: round(v) for h, v in sorted(hdg_map.items(), key=lambda x: -x[1])[:10]}
        accounts_context[acct] = {
            "all_time_total": round(total),
            "by_heading_alltime": top_cats,
            "monthly_last_12": recent_mo,
            "by_fy": fy_totals,
            "by_fy_and_heading": fy_heading,
        }

    # Approval log summary
    log = db.load("approval_log")
    pending_count = sum(1 for e in log if not e.get("status") == "cancelled" and
                        not any(a.get("action") in ("APPROVED","AUTO_APPROVE")
                                for a in e.get("approved_actions",[])))

    today = datetime.now()
    current_fy = _fy(today.strftime("%Y-%m"))

    system_prompt = f"""You are a personal finance assistant for Sudhir, an Indian executive household (Mumbai).
You have full access to their transaction ledger. Answer questions concisely and specifically.

Today: {today.strftime("%d %b %Y")}. Current financial year: {current_fy} (April–March).
FY label format: FY2627 = April 2026 – March 2027.

Data available:
1. All-time spend by heading (all accounts combined): {json.dumps({h: round(v) for h, v in sorted(by_heading.items(), key=lambda x: -x[1])[:20]})}
2. Monthly spend by heading — last 12 months (all accounts): {json.dumps(monthly_data)}
3. Top 20 vendors by all-time spend: {json.dumps({k: round(v) for k, v in top_vendors.items()})}
4. Per-account breakdown (all-time total, top headings, monthly last 12m, FY totals, FY×heading): {json.dumps(accounts_context)}
5. Pending expense approvals: {pending_count}

All amounts in ₹. Use Indian number format (lakhs/crores) for large numbers.
Keep answers short and direct. Use bullet points. If data is insufficient, say exactly what's missing."""

    messages = [{"role":"system","content":system_prompt}]
    for h in history[-10:]:  # keep last 10 turns for context
        messages.append({"role": h["role"], "content": h["content"]})
    messages.append({"role":"user","content":question})

    try:
        resp = client.chat.completions.create(
            model=deployment, max_completion_tokens=600,
            messages=messages
        )
        reply = resp.choices[0].message.content.strip()
    except Exception as e:
        reply = f"Error: {e}"

    return jsonify({"reply": reply})


@app.route("/api/saved-insights", methods=["GET"])
@login_required
def api_saved_insights():
    return jsonify({"items": db.load("saved_insights", [])})


@app.route("/api/save-insight", methods=["POST"])
@login_required
def api_save_insight():
    import uuid
    data = request.get_json()
    items = db.load("saved_insights", [])
    items.insert(0, {
        "id": str(uuid.uuid4()),
        "question": data.get("question",""),
        "reply": data.get("reply",""),
        "saved_at": datetime.now().isoformat(),
    })
    db.save("saved_insights", items)
    return jsonify({"ok": True})


@app.route("/api/delete-insight", methods=["POST"])
@login_required
def api_delete_insight():
    data = request.get_json()
    items = db.load("saved_insights", [])
    items = [i for i in items if i.get("id") != data.get("id")]
    db.save("saved_insights", items)
    return jsonify({"ok": True})


# ── Projects ──────────────────────────────────────────────────────────────────

@app.route("/api/projects", methods=["GET"])
@login_required
def api_projects_list():
    import uuid
    projects = db.load("projects", [])
    ledger   = load_ledger()
    # compute spent + txn count per project (match by id or name for legacy)
    for p in projects:
        matched = [t for t in ledger if t.get("project") in (p["id"], p["name"])]
        p["spent"]     = round(sum(float(t.get("debit") or 0) for t in matched))
        p["txn_count"] = len(matched)
    def _clean(p):
        out = dict(p)
        out["attachments"] = [{k:v for k,v in a.items() if k != "data_b64"} for a in p.get("attachments",[])]
        return out
    return jsonify({"projects": [_clean(p) for p in projects]})


@app.route("/api/projects", methods=["POST"])
@login_required
def api_projects_save():
    import uuid
    data     = request.get_json() or {}
    projects = db.load("projects", [])
    pid      = data.get("id")
    if pid:
        for p in projects:
            if p["id"] == pid:
                p.update({k: data[k] for k in ("name","budget","desc","status","contractor") if k in data})
                break
    else:
        projects.append({
            "id":          str(uuid.uuid4()),
            "name":        data.get("name","Unnamed"),
            "contractor":  data.get("contractor",""),
            "budget":      data.get("budget", 0),
            "desc":        data.get("desc",""),
            "status":      data.get("status","open"),
            "items":       [],
            "attachments": [],
            "created_at":  datetime.now().isoformat(),
        })
    db.save("projects", projects)
    return jsonify({"ok": True})


@app.route("/api/projects/<project_id>", methods=["DELETE"])
@login_required
def api_projects_delete(project_id):
    projects = db.load("projects", [])
    projects = [p for p in projects if p["id"] != project_id]
    db.save("projects", projects)
    return jsonify({"ok": True})


@app.route("/api/projects/<project_id>/transactions", methods=["GET"])
@login_required
def api_project_transactions(project_id):
    projects = db.load("projects", [])
    proj     = next((p for p in projects if p["id"] == project_id), None)
    if not proj:
        return jsonify({"error": "not found"}), 404
    ledger = load_ledger()
    txns   = [t for t in ledger if t.get("project") in (project_id, proj["name"])]
    txns.sort(key=lambda t: (t.get("date") or t.get("timestamp","") or ""), reverse=True)
    return jsonify({"transactions": txns})


@app.route("/api/projects/<project_id>/items", methods=["POST"])
@login_required
def api_project_items_save(project_id):
    projects = db.load("projects", [])
    proj = next((p for p in projects if p["id"] == project_id), None)
    if not proj:
        return jsonify({"error": "not found"}), 404
    data = request.get_json() or {}
    proj["items"] = data.get("items", [])
    db.save("projects", projects)
    return jsonify({"ok": True})


@app.route("/api/projects/<project_id>/items/<item_id>/ai-comment", methods=["POST"])
@login_required
def api_project_item_ai(project_id, item_id):
    try:
        from openai import AzureOpenAI
        client = AzureOpenAI(
            api_key=os.getenv("AZURE_OPENAI_KEY"),
            azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),
            api_version=os.getenv("AZURE_OPENAI_API_VERSION","2024-12-01-preview"),
        )
        deployment = os.getenv("AZURE_OPENAI_DEPLOYMENT","gpt-5.5")
    except Exception as e:
        return jsonify({"comment": f"AI unavailable: {e}"})

    data = request.get_json() or {}
    material   = data.get("material","")
    qty        = data.get("qty", 0)
    unit_price = data.get("unit_price", 0)
    value      = data.get("value", 0)

    prompt = f"""You are evaluating a construction/project line item for a residential property in Mumbai, India.
Item: {material}
Quantity: {qty}
Price per unit: ₹{unit_price}
Total value: ₹{value}

Give a 1-sentence assessment: is this price reasonable for Mumbai in 2025-26? Flag if it seems high or low. Be specific."""

    try:
        resp = client.chat.completions.create(
            model=deployment,
            max_completion_tokens=120,
            messages=[{"role":"user","content":prompt}]
        )
        comment = resp.choices[0].message.content.strip()
    except Exception as e:
        comment = f"Error: {e}"

    # save comment back to item
    projects = db.load("projects", [])
    proj = next((p for p in projects if p["id"] == project_id), None)
    if proj:
        for item in proj.get("items", []):
            if item.get("id") == item_id:
                item["ai_comment"] = comment
                break
        db.save("projects", projects)
    return jsonify({"comment": comment})


@app.route("/api/projects/<project_id>/attachments", methods=["POST"])
@login_required
def api_project_attach_upload(project_id):
    import uuid, base64
    projects = db.load("projects", [])
    proj = next((p for p in projects if p["id"] == project_id), None)
    if not proj:
        return jsonify({"error": "not found"}), 404
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "no file"}), 400
    attach_type = request.form.get("type","other")
    data_b64 = base64.b64encode(file.read()).decode()
    attachment = {
        "id":           str(uuid.uuid4()),
        "name":         file.filename,
        "type":         attach_type,
        "content_type": file.content_type,
        "data_b64":     data_b64,
        "uploaded_at":  datetime.now().isoformat(),
    }
    if "attachments" not in proj:
        proj["attachments"] = []
    proj["attachments"].append(attachment)
    db.save("projects", projects)
    return jsonify({"ok": True})


@app.route("/api/projects/<project_id>/attachments/<attach_id>", methods=["GET"])
@login_required
def api_project_attach_download(project_id, attach_id):
    import base64
    from flask import Response
    projects = db.load("projects", [])
    proj = next((p for p in projects if p["id"] == project_id), None)
    if not proj:
        return jsonify({"error": "not found"}), 404
    att = next((a for a in proj.get("attachments",[]) if a["id"] == attach_id), None)
    if not att:
        return jsonify({"error": "not found"}), 404
    data = base64.b64decode(att["data_b64"])
    return Response(data, content_type=att.get("content_type","application/octet-stream"),
                    headers={"Content-Disposition": f'inline; filename="{att["name"]}"'})


@app.route("/api/projects/<project_id>/attachments/<attach_id>", methods=["DELETE"])
@login_required
def api_project_attach_delete(project_id, attach_id):
    projects = db.load("projects", [])
    proj = next((p for p in projects if p["id"] == project_id), None)
    if not proj:
        return jsonify({"error": "not found"}), 404
    proj["attachments"] = [a for a in proj.get("attachments",[]) if a["id"] != attach_id]
    db.save("projects", projects)
    return jsonify({"ok": True})


@app.route("/export", methods=["GET"])
def export():
    month_str = request.args.get("month")
    year, month = (int(x) for x in month_str.split("-")) if month_str else (None, None)
    path = export_monthly_excel(year=year, month=month)
    return jsonify({"status": "exported", "file": path})


@app.route("/api/recent-expenses", methods=["GET"])
def api_recent_expenses():
    """Return last 20 approval log entries for the photo upload dropdown."""
    try:
        log = _db.load("approval_log") or []
        recent = sorted(log, key=lambda e: e.get("timestamp", ""), reverse=True)[:20]
        items = [
            {
                "id": e.get("request_id", ""),
                "vendor": e.get("vendor", ""),
                "amount": e.get("amount", 0),
                "date": (e.get("timestamp", "")[:10]),
                "category": e.get("category", ""),
            }
            for e in recent
        ]
        return jsonify({"expenses": items})
    except Exception as e:
        return jsonify({"expenses": [], "error": str(e)})


@app.route("/api/photos/upload", methods=["POST"])
def api_photos_upload():
    """Store estimate and bill photos (base64) for an expense."""
    data = request.get_json()
    expense_id = data.get("expense_id", "").strip()
    if not expense_id:
        return jsonify({"error": "expense_id required"}), 400
    estimates = data.get("estimates", [])
    bills = data.get("bills", [])
    existing = _db.load(f"photos_{expense_id}") or {"estimates": [], "bills": []}
    existing["estimates"] = existing.get("estimates", []) + estimates
    existing["bills"] = existing.get("bills", []) + bills
    _db.save(f"photos_{expense_id}", existing)
    return jsonify({
        "ok": True,
        "expense_id": expense_id,
        "estimates": len(existing["estimates"]),
        "bills": len(existing["bills"]),
    })


@app.route("/api/transfer-recon", methods=["GET"])
@login_required
def api_transfer_recon():
    """
    Two sections:
    1. All transactions typed 'transfer' — listed as debits and credits for manual review.
    2. Suspected transfers — non-transfer transactions that look like interbank moves:
       - Description contains NEFT/RTGS/IMPS/TFR/transfer keywords, OR
       - Large round amount (≥₹10,000, multiple of 1000) that exactly matches
         a debit or credit on a DIFFERENT account within 7 days.
    """
    from src.master_ledger import _parse_date as _ml_pd
    from datetime import date as _date2

    ledger = load_ledger()

    # Fiscal year filter
    _fy_tr = request.args.get("fy", "FY27")
    _FY_TR = {"FY24": ("2023-04-01","2024-03-31"), "FY25": ("2024-04-01","2025-03-31"), "FY26": ("2025-04-01","2026-03-31"), "FY27": ("2026-04-01","2027-03-31")}
    if _fy_tr in _FY_TR:
        _lo = _date2.fromisoformat(_FY_TR[_fy_tr][0])
        _hi = _date2.fromisoformat(_FY_TR[_fy_tr][1])
        def _in_fy_tr(t):
            d = _ml_pd(t.get("date",""))
            return d and _lo <= d.date() <= _hi
        ledger = [t for t in ledger if _in_fy_tr(t)]

    def _row(t):
        return {
            "txn_id":      t["txn_id"],
            "seq":         t.get("seq"),
            "date":        t.get("date"),
            "account":     t.get("account"),
            "direction":   "debit" if float(t.get("debit") or 0) > 0 else "credit",
            "amount":      float(t.get("debit") or t.get("credit") or 0),
            "description": (t.get("raw_description") or t.get("transaction_details") or t.get("description") or "").strip(),
            "paid_to":     t.get("paid_to") or "",
            "type":        t.get("type") or "",
            "heading":     t.get("heading") or "",
        }

    # ── Section 1: match transfer debits ↔ credits ───────────────────────────
    # At least one side must be typed "transfer"; counterpart can be any type.
    # Tolerances: amount ±max(₹1000, 1%), date ±14 days.
    transfers = [t for t in ledger if (t.get("type") or "").lower() == "transfer"]
    t_debits  = [t for t in transfers if float(t.get("debit")  or 0) > 0]
    t_credits = [t for t in transfers if float(t.get("credit") or 0) > 0]

    # Full-ledger pools for counterpart search
    all_debits_pool  = [t for t in ledger if float(t.get("debit")  or 0) > 0]
    all_credits_pool = [t for t in ledger if float(t.get("credit") or 0) > 0]

    matched_d = set()
    matched_c = set()
    pairs = []

    for d in sorted(t_debits, key=lambda x: x.get("date", "")):
        if d["txn_id"] in matched_d:
            continue
        d_amt  = float(d.get("debit", 0))
        d_date = _ml_pd(d.get("date", ""))
        if not d_date:
            continue
        best_c, best_score = None, (999, 999)
        for c in all_credits_pool:
            if c["txn_id"] in matched_c:
                continue
            if c["txn_id"] == d["txn_id"]:
                continue
            c_amt  = float(c.get("credit", 0))
            c_date = _ml_pd(c.get("date", ""))
            if not c_date:
                continue
            if abs(d_amt - c_amt) > max(1000, d_amt * 0.01):
                continue
            days = abs((d_date - c_date).days)
            if days > 14:
                continue
            # Prefer seq-adjacent entries as tiebreaker (same-batch imports)
            seq_gap = abs((d.get("seq") or 0) - (c.get("seq") or 0))
            score = (days, min(seq_gap, 999))
            if score < best_score:
                best_c, best_score = c, score
        if best_c:
            matched_d.add(d["txn_id"])
            matched_c.add(best_c["txn_id"])
            pairs.append({
                "debit_seq":    d.get("seq"),
                "credit_seq":   best_c.get("seq"),
                "debit_date":   d.get("date"),
                "credit_date":  best_c.get("date"),
                "days_gap":     best_days,
                "amount":       d_amt,
                "from_account": d.get("account"),
                "to_account":   best_c.get("account"),
                "from_desc":    (d.get("raw_description") or d.get("transaction_details") or "").strip(),
                "to_desc":      (best_c.get("raw_description") or best_c.get("transaction_details") or "").strip(),
            })

    # Also try credit-typed transfers whose counterpart debit may be any type
    for c in sorted(t_credits, key=lambda x: x.get("date", "")):
        if c["txn_id"] in matched_c:
            continue
        c_amt  = float(c.get("credit", 0))
        c_date = _ml_pd(c.get("date", ""))
        if not c_date:
            continue
        best_d, best_score = None, (999, 999)
        for d in all_debits_pool:
            if d["txn_id"] in matched_d:
                continue
            if d["txn_id"] == c["txn_id"]:
                continue
            d_amt  = float(d.get("debit", 0))
            d_date = _ml_pd(d.get("date", ""))
            if not d_date:
                continue
            if abs(c_amt - d_amt) > max(1000, c_amt * 0.01):
                continue
            days = abs((c_date - d_date).days)
            if days > 14:
                continue
            seq_gap = abs((c.get("seq") or 0) - (d.get("seq") or 0))
            score = (days, min(seq_gap, 999))
            if score < best_score:
                best_d, best_score = d, score
        if best_d:
            matched_c.add(c["txn_id"])
            matched_d.add(best_d["txn_id"])
            pairs.append({
                "debit_seq":    best_d.get("seq"),
                "credit_seq":   c.get("seq"),
                "debit_date":   best_d.get("date"),
                "credit_date":  c.get("date"),
                "days_gap":     best_days,
                "amount":       c_amt,
                "from_account": best_d.get("account"),
                "to_account":   c.get("account"),
                "from_desc":    (best_d.get("raw_description") or best_d.get("transaction_details") or "").strip(),
                "to_desc":      (c.get("raw_description") or c.get("transaction_details") or "").strip(),
            })

    transfer_debits  = [_row(t) for t in t_debits  if t["txn_id"] not in matched_d]
    transfer_credits = [_row(t) for t in t_credits if t["txn_id"] not in matched_c]

    # ── Section 2: suspected transfers ────────────────────────────────────────
    # Only flag a non-transfer entry if it has the same amount as an UNMATCHED
    # transfer debit or credit, on any account, within 7 days — i.e. it looks
    # like the missing counterpart of an unreconciled transfer.
    unmatched_transfers = (
        [t for t in t_debits  if t["txn_id"] not in matched_d] +
        [t for t in t_credits if t["txn_id"] not in matched_c]
    )

    suspected = []
    seen_suspected = set()
    for ut in unmatched_transfers:
        ut_amt  = float(ut.get("debit") or ut.get("credit") or 0)
        ut_date = _ml_pd(ut.get("date", ""))
        ut_is_debit = float(ut.get("debit") or 0) > 0
        if not ut_date:
            continue
        for t in ledger:
            if (t.get("type") or "").lower() == "transfer":
                continue
            if t["txn_id"] in seen_suspected:
                continue
            t_amt  = float(t.get("debit") or t.get("credit") or 0)
            t_date = _ml_pd(t.get("date", ""))
            if not t_date:
                continue
            if abs(ut_amt - t_amt) > max(500, ut_amt * 0.005):
                continue
            if abs((ut_date - t_date).days) > 7:
                continue
            # Should be opposite direction to be a counterpart
            t_is_debit = float(t.get("debit") or 0) > 0
            if t_is_debit == ut_is_debit:
                continue
            row = _row(t)
            row["reason"] = [f"matches unreconciled transfer #{ut.get('seq')} (₹{ut_amt:,.0f}) on {ut.get('date')}"]
            row["matched_transfer_seq"] = ut.get("seq")
            suspected.append(row)
            seen_suspected.add(t["txn_id"])

    suspected.sort(key=lambda x: x.get("date", ""), reverse=True)

    # ── Section 3: manual pairs ───────────────────────────────────────────────
    manual_pairs_raw = db.load("manual_transfer_pairs") or []
    seq_to_txn = {t.get("seq"): t for t in ledger if t.get("seq") is not None}
    manual_pairs_out = []
    for mp in manual_pairs_raw:
        t1 = seq_to_txn.get(mp["seq1"])
        t2 = seq_to_txn.get(mp["seq2"])
        if not t1 or not t2:
            continue
        # Ensure d=debit side, c=credit side
        if float(t1.get("debit") or 0) > 0:
            d, c = t1, t2
        else:
            d, c = t2, t1
        d_amt = float(d.get("debit") or 0)
        c_amt = float(c.get("credit") or 0)
        d_date = _ml_pd(d.get("date", ""))
        c_date = _ml_pd(c.get("date", ""))
        days = abs((d_date - c_date).days) if d_date and c_date else 0
        manual_pairs_out.append({
            "debit_seq":    d.get("seq"),
            "credit_seq":   c.get("seq"),
            "debit_date":   d.get("date"),
            "credit_date":  c.get("date"),
            "days_gap":     days,
            "amount":       d_amt or c_amt,
            "from_account": d.get("account"),
            "to_account":   c.get("account"),
            "from_desc":    (d.get("raw_description") or d.get("transaction_details") or "").strip(),
            "to_desc":      (c.get("raw_description") or c.get("transaction_details") or "").strip(),
            "manual":       True,
        })
        # Remove from unmatched lists if present
        transfer_debits  = [r for r in transfer_debits  if r.get("seq") not in (d.get("seq"), c.get("seq"))]
        transfer_credits = [r for r in transfer_credits if r.get("seq") not in (d.get("seq"), c.get("seq"))]
        suspected        = [r for r in suspected        if r.get("seq") not in (d.get("seq"), c.get("seq"))]

    return jsonify({
        "pairs":            pairs + manual_pairs_out,
        "transfer_debits":  transfer_debits,
        "transfer_credits": transfer_credits,
        "suspected":        suspected,
    })


@app.route("/api/transfer-recon/manual-pair", methods=["POST"])
@login_required
def api_transfer_recon_manual_pair():
    data = request.get_json() or {}
    seq1 = data.get("seq1")
    seq2 = data.get("seq2")
    if seq1 is None or seq2 is None:
        return jsonify({"error": "seq1 and seq2 required"}), 400
    if seq1 == seq2:
        return jsonify({"error": "seq1 and seq2 must be different"}), 400

    ledger = load_ledger()
    seq_to_txn = {t.get("seq"): t for t in ledger if t.get("seq") is not None}
    if seq1 not in seq_to_txn:
        return jsonify({"error": f"Seq #{seq1} not found in ledger"}), 404
    if seq2 not in seq_to_txn:
        return jsonify({"error": f"Seq #{seq2} not found in ledger"}), 404

    pairs = db.load("manual_transfer_pairs") or []
    # Remove any existing link that involves either seq — old links broken
    pairs = [p for p in pairs if seq1 not in (p["seq1"], p["seq2"]) and seq2 not in (p["seq1"], p["seq2"])]
    pairs.append({"seq1": seq1, "seq2": seq2})
    db.save("manual_transfer_pairs", pairs)
    return jsonify({"ok": True, "linked": [seq1, seq2]})


def _parse_acc27_excel(file_obj):
    """Extract SBI entries from ACC27 SudhirExpenses sheet."""
    import openpyxl, io
    from datetime import datetime as _dt2

    HEADING_FIX = {
        'Staff salary': 'Staff Salary', 'Grocries': 'Groceries',
        'Internet': 'Misc', 'Miscellaneous': 'Misc',
    }

    # FY27 month_num → expected calendar year (month_num stored in col 0)
    FY_MONTH_YEAR = {1: 2026, 2: 2026, 3: 2026, 4: 2026, 5: 2026,
                     6: 2026, 7: 2026, 8: 2026, 9: 2026, 10: 2027,
                     11: 2027, 12: 2027}  # Apr=1…Mar=12

    def _fix_year(d, mo_num):
        """Correct year typos using the FY month number."""
        if d is None or mo_num is None: return d
        expected = FY_MONTH_YEAR.get(int(mo_num))
        if expected and d.year != expected:
            try: return d.replace(year=expected)
            except ValueError: pass
        return d

    def _pd(cell, mo_num=None):
        v = cell.value
        if isinstance(v, _dt2):
            fmt = cell.number_format or ''
            if 'mm-dd' in fmt:
                try: v = _dt2(v.year, v.day, v.month)
                except ValueError: pass
            return _fix_year(v, mo_num)
        if isinstance(v, str):
            s = v.strip()
            # Fix missing hyphen: '06-062026' → '06-06-2026'
            import re
            s = re.sub(r'^(\d{2})-(\d{2})(\d{4})$', r'\1-\2-\3', s)
            for sfmt in ('%d-%m-%Y','%d/%m/%Y','%Y-%m-%d','%d-%b-%Y'):
                try: return _fix_year(_dt2.strptime(s, sfmt), mo_num)
                except: pass
        return None

    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb['SudhirExpenses']
    out = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        if row[0].value is None: continue
        acct = str(row[2].value or '').lower().replace(' ','')
        if 'sbi' not in acct: continue
        d = _pd(row[3], row[0].value)
        if not d: continue
        debit  = float(row[6].value) if row[6].value else 0
        credit = float(row[7].value) if row[7].value else 0
        amt = abs(debit or credit)
        if amt == 0: continue
        heading = str(row[9].value or '').strip()
        heading = HEADING_FIX.get(heading, heading)
        out.append({
            'date': d.strftime('%d/%m/%Y'),
            'account': 'SBI-4852',
            'debit': debit, 'credit': credit, 'amount': amt,
            'description': str(row[4].value or '').strip(),
            'paid_to': str(row[5].value or '').strip() or None,
            'type': str(row[8].value or '').strip().lower(),
            'heading': heading,
            'notes': str(row[10].value or '').strip() or None,
        })
    return out


_ATM_KEYWORDS = ["atm", "atw", "cash withdrawal", "cash wthdl", "atm wtdl", "cash wtdl"]

def _run_acc27_match(xl_list, apply_it=False, limit=5):
    import hashlib
    from src.master_ledger import _parse_date as _ml_pd, _assign_seq
    from collections import defaultdict

    ledger = load_ledger()

    # Index by rounded amount only — account name may differ between Excel and ledger
    by_amt = defaultdict(list)
    for t in ledger:
        for field in ("debit","credit"):
            v = round(float(t.get(field) or 0))
            if v > 0:
                by_amt[v].append(t)

    def _is_sbi(t):
        acct = (t.get("account") or "").upper()
        return "SBI" in acct

    existing_ids = {t["txn_id"] for t in ledger}

    results = []
    applied = 0
    created = 0
    already_matched = set()

    for xl in xl_list:
        xl_amt   = round(float(xl.get("amount", 0)))
        xl_date  = _ml_pd(xl.get("date",""))
        candidates = by_amt.get(xl_amt, [])
        best, best_days = None, 999
        for t in candidates:
            if t.get("txn_id") in already_matched: continue
            t_date = _ml_pd(t.get("date",""))
            if not t_date or not xl_date:
                if best is None and _is_sbi(t): best, best_days = t, 99
                continue
            days = abs((xl_date - t_date).days)
            if days > 30: continue
            if days < best_days or (days == best_days and _is_sbi(t) and not _is_sbi(best)):
                best, best_days = t, days

        if best is None:
            results.append({"xl": xl, "ledger": None, "match": "none"})
            if apply_it:
                # Create a new ledger entry from the Excel row
                raw = f"acc27|{xl.get('date','')}|{xl.get('description','')}|{xl.get('amount',0)}"
                new_id = hashlib.sha1(raw.encode()).hexdigest()[:16]
                if new_id not in existing_ids:
                    # June entries have confirmed type/heading from Excel — not uncertain
                    # Apr/May unmatched are likely cash payments — flag for Vincent
                    is_june = (xl.get("date","")[3:5] == "06")
                    new_entry = {
                        "txn_id":          new_id,
                        "date":            xl.get("date",""),
                        "account":         "SBI-4852",
                        "debit":           xl.get("debit") or 0,
                        "credit":          xl.get("credit") or 0,
                        "raw_description": xl.get("description",""),
                        "description":     xl.get("description",""),
                        "type":            xl.get("type") or "expense",
                        "heading":         xl.get("heading",""),
                        "paid_to":         xl.get("paid_to") or None,
                        "source":          "acc27_excel",
                    }
                    if not is_june:
                        new_entry["uncertain"]        = True
                        new_entry["uncertain_fields"] = ["heading"]
                    if xl.get("notes"): new_entry["remarks"] = xl["notes"]
                    ledger.append(new_entry)
                    existing_ids.add(new_id)
                    created += 1
            continue

        already_matched.add(best.get("txn_id"))
        confidence = "exact" if best_days <= 3 else "good" if best_days <= 14 else "weak"
        results.append({
            "xl": xl,
            "ledger": {k: best.get(k) for k in
                       ["txn_id","seq","date","account","debit","credit",
                        "raw_description","type","heading","paid_to"]},
            "match": confidence, "days_gap": best_days,
        })

        if apply_it:
            if confidence == "exact":
                best["type"]    = xl.get("type") or best.get("type")
                best["heading"] = xl.get("heading") or best.get("heading")
                if xl.get("paid_to"): best["paid_to"] = xl["paid_to"]
                if xl.get("notes"):   best["remarks"]  = xl["notes"]
                best.pop("uncertain", None)
                best.pop("uncertain_fields", None)
                applied += 1
            else:
                best["uncertain"]        = True
                best["uncertain_fields"] = ["heading"]

        if not apply_it and len([r for r in results if r["match"] != "none"]) >= limit:
            break

    if apply_it:
        # Rule: all SBI ATM withdrawals → type=transfer, heading=Cash + paired cash-in credit
        from src.master_ledger import _ensure_cash_in, _is_sbi_cash_transfer
        atm_fixed = 0
        for t in ledger:
            if not _is_sbi(t): continue
            desc = (t.get("raw_description") or "").lower()
            if any(kw in desc for kw in _ATM_KEYWORDS) and t.get("type") != "transfer":
                t["type"]    = "transfer"
                t["heading"] = t.get("heading") or "Cash"
                atm_fixed += 1
        # Ensure paired cash-in credits exist for all SBI→Cash transfers
        for t in ledger:
            if _is_sbi_cash_transfer(t):
                _ensure_cash_in(ledger, t)
        _assign_seq(ledger)
        _save_json(LEDGER_PATH, ledger)
        sbi_total    = sum(1 for t in ledger if "SBI" in (t.get("account") or "").upper())
        sbi_uncertain = sum(1 for t in ledger if t.get("uncertain") and "SBI" in (t.get("account") or "").upper())
        total_uncertain = sum(1 for t in ledger if t.get("uncertain"))
        return {"applied": applied, "created": created, "atm_fixed": atm_fixed,
                "total": len(xl_list), "ledger_total": len(ledger),
                "sbi_total": sbi_total, "sbi_uncertain": sbi_uncertain,
                "total_uncertain": total_uncertain}

    matched   = [r for r in results if r["match"] != "none"]
    unmatched = [r for r in results if r["match"] == "none"]
    return {"results": matched, "unmatched": unmatched,
            "unmatched_count": len(unmatched), "total_xl": len(xl_list)}


@app.route("/api/admin/acc27-debug")
@login_required
def api_acc27_debug():
    """Show SBI ledger entries and first 20 unmatched Excel entries for diagnosis."""
    from src.master_ledger import _parse_date as _ml_pd
    import io
    ledger  = load_ledger()
    sbi_txns = [t for t in ledger if "SBI" in (t.get("account") or "").upper()]
    # unique accounts in ledger
    all_accounts = sorted(set(t.get("account","") for t in ledger))
    xl_list = []
    if request.files.get("file"):
        xl_list = _parse_acc27_excel(io.BytesIO(request.files["file"].read()))

    # For each xl entry, show nearest SBI ledger entry by date (ignoring amount)
    samples = []
    for xl in xl_list[:30]:
        xl_date = _ml_pd(xl.get("date",""))
        nearest, nearest_days = None, 9999
        for t in sbi_txns:
            td = _ml_pd(t.get("date",""))
            if td and xl_date:
                d = abs((xl_date - td).days)
                if d < nearest_days:
                    nearest, nearest_days = t, d
        samples.append({
            "xl_date": xl.get("date"), "xl_amt": xl.get("amount"),
            "xl_desc": xl.get("description","")[:60],
            "nearest_date": nearest.get("date") if nearest else None,
            "nearest_amt": (nearest.get("debit") or nearest.get("credit")) if nearest else None,
            "nearest_acct": nearest.get("account") if nearest else None,
            "nearest_seq": nearest.get("seq") if nearest else None,
            "days_gap": nearest_days if nearest else None,
        })

    return jsonify({
        "total_ledger": len(ledger),
        "sbi_count": len(sbi_txns),
        "all_accounts_sample": all_accounts[:50],
        "sbi_sample": [{
            "seq": t.get("seq"), "date": t.get("date"),
            "account": t.get("account"),
            "debit": t.get("debit"), "credit": t.get("credit"),
            "raw_description": (t.get("raw_description","") or "")[:80]
        } for t in sbi_txns[:30]],
        "xl_nearest_sbi": samples,
    })


@app.route("/admin/acc27")
@login_required
def admin_acc27_page():
    return render_template("acc27_admin.html")


@app.route("/api/admin/acc27-match-preview", methods=["POST"])
@login_required
def api_acc27_match_preview():
    """Accept multipart Excel upload OR JSON entries list. Returns match preview."""
    import io
    if request.files.get("file"):
        xl_list  = _parse_acc27_excel(io.BytesIO(request.files["file"].read()))
        limit    = int(request.form.get("limit", 5))
        apply_it = request.form.get("apply","") == "1"
    else:
        data     = request.get_json() or {}
        xl_list  = data.get("entries", [])
        limit    = int(data.get("limit", 5))
        apply_it = data.get("apply", False)

    result = _run_acc27_match(xl_list, apply_it=apply_it, limit=limit)
    return jsonify(result)


@app.route("/api/year-summary", methods=["GET"])
@login_required
def api_year_summary():
    import os, calendar
    from datetime import date as _date
    fy = request.args.get("fy", "FY27")

    GROUPS = [
        ("HOUSEHOLD",   ["Misc","Cash","Electricity & Gas","Groceries","Staff Salary"]),
        ("PERSONAL",    ["Alcohol","Wellness"]),
        ("FAMILY",      ["Clothes","Gifts","Medical","Amma","Ketki","Children Education"]),
        ("GIVING",      ["Charity","Uspaar"]),
        ("LIFESTYLE",   ["Holiday","Eating Out","Entertainment"]),
        ("PROPERTY",    ["Malhar","Maintenance Expense","Home office","One Time Charge","Kalpataru Maintenance"]),
        ("FINANCIAL",   ["Financial Expense / OD Interest","Insurance","Home Loan","Tax"]),
    ]
    MONTHS = ["Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb","Mar"]

    # Multi-year annual totals from ExpenseSummary sheet (FY23-FY26)
    multi_year = {}
    try:
        import openpyxl
        _xl = os.path.join(os.path.dirname(__file__), "data", "ACC26ver5_MASTER.xlsx")
        wb  = openpyxl.load_workbook(_xl, data_only=True)
        ws  = wb["ExpenseSummary"]
        FY_COLS = {"FY23": 1, "FY24": 2, "FY25": 3, "FY26": 4}
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            cat = row[0]
            if not cat or not isinstance(cat, str):
                continue
            cat = cat.strip()
            vals = {}
            for fyname, col in FY_COLS.items():
                v = row[col]
                vals[fyname] = round(float(v) * 100000) if isinstance(v, (int, float)) and v == v else 0
            multi_year[cat] = vals
    except Exception:
        pass

    if fy == "FY26":
        monthly = _FY26_MONTHLY
        ann     = _FY26_ACTUALS
        categories = {}
        for cat, mv in monthly.items():
            annual = ann.get(cat) or sum(mv.values())
            categories[cat] = {"monthly": mv, "annual": annual,
                                "prior": {k: v.get(cat, 0) for k, v in
                                          {f: {c: multi_year.get(c, {}).get(f, 0) for c in multi_year}
                                           for f in ["FY23","FY24","FY25"]}.items()}}
    elif fy == "FY27":
        # Derive from live master ledger (Apr 2026–Mar 2027)
        from src.master_ledger import _parse_date as _ml_pd
        ledger  = load_ledger()
        FY_START = _date(2026, 4, 1)
        FY_END   = _date(2027, 3, 31)
        monthly_raw = {m: {} for m in MONTHS}
        MONTH_MAP = {4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",
                     10:"Oct",11:"Nov",12:"Dec",1:"Jan",2:"Feb",3:"Mar"}
        for txn in ledger:
            if (txn.get("type") or "").lower() not in ("expense","official"):
                continue
            d = _ml_pd(txn.get("date",""))
            if not d or not (FY_START <= d <= FY_END):
                continue
            heading = txn.get("heading") or "Misc"
            amt = float(txn.get("debit") or 0)
            mname = MONTH_MAP.get(d.month, "")
            if mname:
                monthly_raw[mname][heading] = monthly_raw[mname].get(heading, 0) + amt
        # Transpose to {heading: {month: amt}}
        all_cats = set(h for m in monthly_raw.values() for h in m)
        categories = {}
        for cat in all_cats:
            mv = {m: monthly_raw[m].get(cat, 0) for m in MONTHS}
            categories[cat] = {"monthly": mv, "annual": sum(mv.values()),
                                "prior": {f: multi_year.get(cat, {}).get(f, 0) for f in ["FY24","FY25","FY26"]}}
    else:
        categories = {}

    return jsonify({"fy": fy, "groups": GROUPS, "months": MONTHS,
                    "categories": categories, "multi_year": multi_year})


@app.route("/admin/acc26")
@login_required
def admin_acc26_page():
    return render_template("acc26_admin.html")


def _parse_acc26_sudhir(file_bytes):
    """Parse SudhirExpenses sheet from ACC26 xlsx. Returns list of normalised ledger-ready dicts."""
    import io, openpyxl
    from datetime import datetime as _dt

    _ACCT_MAP = {
        "sbi4852":"SBI-4852","sbi":"SBI-4852",
        "icic0018":"ICICI-0018","ici0018":"ICICI-0018","icici0018":"ICICI-0018","ICI0018":"ICICI-0018","ICIC0018":"ICICI-0018",
        "icic7281":"ICICI-7281","ici7281":"ICICI-7281","icici7281":"ICICI-7281","ICIC7281":"ICICI-7281",
        "icic1331":"ICICI-1331","ici1331":"ICICI-1331","icici1331":"ICICI-1331","icci1331":"ICICI-1331",
        "ICIC1331":"ICICI-1331","Icici1331":"ICICI-1331",
        "icic9175":"ICICI-9175","ici9175":"ICICI-9175","icici9175":"ICICI-9175","ICIC9175":"ICICI-9175",
        "cridit card":"ICICI-CC","credit card":"ICICI-CC",
    }
    _TYPE_MAP = {
        "expense":"expense","official":"official","transfer":"transfer","transfer ":"transfer",
        "income":"income","tax":"expense","investment":"investment","error":"error",
    }
    _HEADING_NORM_ACC26 = {
        "alchol":"Alcohol","alcohol":"Alcohol",
        "amma":"Amma","AMMA":"Amma",
        "birthday parties":"Entertainment","books":"Entertainment","entertaiment":"Entertainment","entertainment":"Entertainment",
        "cash":"Cash",
        "charity":"Charity","donation":"Charity",
        "children education":"Children Education","children education ":"Children Education",
        "clothes":"Clothes","clothing":"Clothes","dry cleaning":"Clothes",
        "e&g":"Electricity & Gas","electricity & gas":"Electricity & Gas","electricity":"Electricity & Gas",
        "eating out":"Eating Out","eating out ":"Eating Out",
        "financial":"Financial Expense / OD Interest","financial expense":"Financial Expense / OD Interest",
        "gift":"Gifts","gifts":"Gifts",
        "groceries":"Groceries","grociries":"Groceries","grocries":"Groceries",
        "hoilday":"Holiday","holiday":"Holiday",
        "home office":"Home office","home office ":"Home office",
        "insurance":"Insurance",
        "ketki":"Ketki",
        "maintenance":"Maintenance Expense","maintenance expense":"Maintenance Expense",
        "malhar":"Malhar","malhar renovation":"Malhar",
        "medical":"Medical",
        "misc":"Misc","miscellaneous":"Misc",
        "one time charge":"One Time Charge","one time":"One Time Charge",
        "staff salary":"Staff Salary","staff slary":"Staff Salary","salary":"Staff Salary",
        "tax":"Tax","home loan":"Home Loan",
        "uspaar":"Uspaar",
        "wellness":"Wellness",
        "kalpataru":"Kalpataru Maintenance","kalpataru maintenance":"Kalpataru Maintenance",
    }

    def _parse_date_acc26(v):
        if isinstance(v, _dt):
            return v.strftime("%d/%m/%Y")
        if isinstance(v, str):
            v = v.strip()
            for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d-%m-%y", "%Y-%m-%d"):
                try:
                    return _dt.strptime(v, fmt).strftime("%d/%m/%Y")
                except ValueError:
                    pass
        return None

    def _norm_heading_acc26(h):
        if not h:
            return ""
        return _HEADING_NORM_ACC26.get(str(h).strip().lower(), str(h).strip())

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb["SudhirExpenses"]
    rows = list(ws.iter_rows(values_only=True))

    FY26_START = _dt(2025, 4, 1)
    FY26_END   = _dt(2026, 3, 31, 23, 59, 59)

    entries = []
    for i, row in enumerate(rows[1:], start=2):  # row 1 is header
        raw_date    = row[3]
        raw_acct    = str(row[2] or "").strip()
        raw_debit   = row[6]
        raw_credit  = row[7]
        raw_type    = str(row[8] or "").strip()
        raw_heading = str(row[9] or "").strip()
        raw_desc    = str(row[4] or "").strip()
        raw_paid    = str(row[5] or "").strip()
        raw_note    = str(row[10] or "").strip()

        date_str = _parse_date_acc26(raw_date)
        if not date_str:
            continue

        # FY26 filter
        try:
            dt = _dt.strptime(date_str, "%d/%m/%Y")
        except ValueError:
            continue
        if not (FY26_START <= dt <= FY26_END):
            continue

        typ = _TYPE_MAP.get(raw_type.lower(), "expense")
        if typ == "error":
            continue   # skip data-entry errors

        debit  = float(raw_debit)  if raw_debit  else 0.0
        credit = float(raw_credit) if raw_credit else 0.0
        # Negative debits in this sheet mean credits/returns
        if debit < 0:
            credit = abs(debit)
            debit  = 0.0

        acct = _ACCT_MAP.get(raw_acct, _ACCT_MAP.get(raw_acct.lower(), raw_acct))
        heading = _norm_heading_acc26(raw_heading) if raw_heading else ""

        entries.append({
            "date":            date_str,
            "account":         acct,
            "raw_description": raw_desc,
            "paid_to":         raw_paid if raw_paid != "None" else "",
            "debit":           debit,
            "credit":          credit,
            "type":            typ,
            "heading":         heading,
            "note":            raw_note if raw_note != "None" else "",
            "source":          "acc26_import",
        })

    return entries


_ACC26_LOCAL = os.path.join(BASE_DIR, "data", "fy26", "ACC26ver5_MASTER.xlsx")
_acc26_job = {}   # in-memory job state: {status, applied, total_parsed, ledger_total, error}


def _run_acc26_import(file_bytes):
    """Background thread: parse + apply ACC26 entries to master ledger."""
    from src.master_ledger import _assign_seq
    global _acc26_job
    try:
        _acc26_job = {"status": "running", "step": "parsing"}
        entries = _parse_acc26_sudhir(file_bytes)
        _acc26_job["step"] = "loading ledger"
        ledger = load_ledger()
        existing_ids = {t["txn_id"] for t in ledger}
        added = 0
        _acc26_job["step"] = "merging"
        for e in entries:
            txn_id = f"acc26-{e['date'].replace('/','')}-{e['account']}-{int(e['debit'] or e['credit'])}"
            if txn_id in existing_ids:
                continue
            existing_ids.add(txn_id)
            ledger.append({
                "txn_id":          txn_id,
                "date":            e["date"],
                "account":         e["account"],
                "raw_description": e["raw_description"],
                "paid_to":         e["paid_to"],
                "debit":           e["debit"],
                "credit":          e["credit"],
                "type":            e["type"],
                "heading":         e["heading"],
                "note":            e["note"],
                "source":          "acc26_import",
                "uncertain":       False,
            })
            added += 1
        _acc26_job["step"] = "saving"
        _assign_seq(ledger)
        _save_json(LEDGER_PATH, ledger)
        _acc26_job = {"status": "done", "applied": added,
                      "total_parsed": len(entries), "ledger_total": len(ledger)}
    except Exception as ex:
        _acc26_job = {"status": "error", "error": str(ex)}


@app.route("/api/admin/acc26-preview", methods=["POST"])
@login_required
def api_acc26_preview():
    """Parse uploaded ACC26 xlsx (or local file if present), return entries for preview."""
    import threading
    f = request.files.get("file")
    if f:
        file_bytes = f.read()
    elif os.path.exists(_ACC26_LOCAL):
        with open(_ACC26_LOCAL, "rb") as fh:
            file_bytes = fh.read()
    else:
        return jsonify({"error": "no file uploaded and local ACC26ver5_MASTER.xlsx not found"}), 400

    apply_it = request.form.get("apply", "") == "1"

    if apply_it:
        # Fire background thread, return immediately
        global _acc26_job
        _acc26_job = {"status": "running", "step": "starting"}
        t = threading.Thread(target=_run_acc26_import, args=(file_bytes,), daemon=True)
        t.start()
        return jsonify({"status": "running"})

    # Preview only — parse now (fast, no DB write)
    entries = _parse_acc26_sudhir(file_bytes)
    by_type = {}
    for e in entries:
        by_type[e["type"]] = by_type.get(e["type"], 0) + 1
    by_heading = {}
    for e in entries:
        h = e["heading"] or "(none)"
        by_heading[h] = by_heading.get(h, 0) + 1

    return jsonify({
        "total": len(entries),
        "by_type": by_type,
        "by_heading": dict(sorted(by_heading.items())),
        "sample": entries[:20],
    })


@app.route("/api/admin/acc26-status", methods=["GET"])
@login_required
def api_acc26_status():
    """Poll import job status."""
    return jsonify(_acc26_job)


_ACC25_LOCAL = os.path.join(BASE_DIR, "data", "FY25", "ACC 25.xlsx")
_acc25_job = {}


def _parse_acc25_sudhir(file_bytes):
    """Parse SudhirExpenses sheet from ACC25 xlsx. Returns list of normalised ledger-ready dicts."""
    import io, openpyxl
    from datetime import datetime as _dt

    _ACCT_MAP = {
        "sbi4852":"SBI-4852","sbi":"SBI-4852",
        "icic0018":"ICICI-0018","ici0018":"ICICI-0018","icici0018":"ICICI-0018","ICI0018":"ICICI-0018","ICIC0018":"ICICI-0018",
        "icic7281":"ICICI-7281","ici7281":"ICICI-7281","icici7281":"ICICI-7281","ICIC7281":"ICICI-7281",
        "icic1331":"ICICI-1331","ici1331":"ICICI-1331","icici1331":"ICICI-1331","icci1331":"ICICI-1331",
        "ICIC1331":"ICICI-1331","Icici1331":"ICICI-1331",
        "icic9175":"ICICI-9175","ici9175":"ICICI-9175","icici9175":"ICICI-9175","ICIC9175":"ICICI-9175",
        "cridit card":"ICICI-CC","credit card":"ICICI-CC",
    }
    _TYPE_MAP = {
        "expense":"expense","official":"official","transfer":"transfer","transfer ":"transfer",
        "income":"income","tax":"expense","investment":"investment","error":"error",
    }
    _HEADING_NORM_ACC25 = {
        "alchol":"Alcohol","alcohol":"Alcohol",
        "amma":"Amma","AMMA":"Amma",
        "birthday parties":"Entertainment","books":"Entertainment","entertaiment":"Entertainment","entertainment":"Entertainment",
        "cash":"Cash",
        "charity":"Charity","donation":"Charity",
        "children education":"Children Education","children education ":"Children Education",
        "clothes":"Clothes","clothing":"Clothes","dry cleaning":"Clothes",
        "e&g":"Electricity & Gas","electricity & gas":"Electricity & Gas","electricity":"Electricity & Gas",
        "eating out":"Eating Out","eating out ":"Eating Out",
        "financial":"Financial Expense / OD Interest","financial expense":"Financial Expense / OD Interest",
        "gift":"Gifts","gifts":"Gifts",
        "groceries":"Groceries","grociries":"Groceries","grocries":"Groceries",
        "hoilday":"Holiday","holiday":"Holiday",
        "home office":"Home office","home office ":"Home office",
        "insurance":"Insurance",
        "ketki":"Ketki",
        "maintenance":"Maintenance Expense","maintenance expense":"Maintenance Expense",
        "malhar":"Malhar","malhar renovation":"Malhar",
        "medical":"Medical",
        "misc":"Misc","miscellaneous":"Misc",
        "one time charge":"One Time Charge","one time":"One Time Charge",
        "staff salary":"Staff Salary","staff slary":"Staff Salary","salary":"Staff Salary",
        "tax":"Tax","home loan":"Home Loan",
        "uspaar":"Uspaar",
        "wellness":"Wellness",
        "kalpataru":"Kalpataru Maintenance","kalpataru maintenance":"Kalpataru Maintenance",
    }

    def _parse_date_acc25(v):
        if isinstance(v, _dt):
            return v.strftime("%d/%m/%Y")
        if isinstance(v, str):
            v = v.strip()
            for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d-%m-%y", "%Y-%m-%d"):
                try:
                    return _dt.strptime(v, fmt).strftime("%d/%m/%Y")
                except ValueError:
                    pass
        return None

    def _norm_heading_acc25(h):
        if not h:
            return ""
        return _HEADING_NORM_ACC25.get(str(h).strip().lower(), str(h).strip())

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb["SudhirExpenses"]
    rows = list(ws.iter_rows(values_only=True))

    FY25_START = _dt(2024, 4, 1)
    FY25_END   = _dt(2025, 3, 31, 23, 59, 59)

    entries = []
    for i, row in enumerate(rows[1:], start=2):
        raw_date    = row[3]
        raw_acct    = str(row[2] or "").strip()
        raw_debit   = row[6]
        raw_credit  = row[7]
        raw_type    = str(row[8] or "").strip()
        raw_desc    = str(row[4] or "").strip()
        raw_paid    = str(row[5] or "").strip()
        raw_heading = str(row[10] or "").strip()   # col index 10 (not 9)

        date_str = _parse_date_acc25(raw_date)
        if not date_str:
            continue

        try:
            dt = _dt.strptime(date_str, "%d/%m/%Y")
        except ValueError:
            continue
        if not (FY25_START <= dt <= FY25_END):
            continue

        typ = _TYPE_MAP.get(raw_type.lower(), "expense")
        if typ == "error":
            continue

        debit  = float(raw_debit)  if raw_debit  else 0.0
        credit = float(raw_credit) if raw_credit else 0.0
        if debit < 0:
            credit = abs(debit)
            debit  = 0.0

        acct = _ACCT_MAP.get(raw_acct, _ACCT_MAP.get(raw_acct.lower(), raw_acct))
        heading = _norm_heading_acc25(raw_heading) if raw_heading else ""

        entries.append({
            "date":            date_str,
            "account":         acct,
            "raw_description": raw_desc,
            "paid_to":         raw_paid if raw_paid != "None" else "",
            "debit":           debit,
            "credit":          credit,
            "type":            typ,
            "heading":         heading,
            "note":            "",
            "source":          "acc25_import",
        })

    return entries


def _run_acc25_import(file_bytes):
    """Background thread: parse + apply ACC25 entries to master ledger."""
    from src.master_ledger import _assign_seq
    global _acc25_job
    try:
        _acc25_job = {"status": "running", "step": "parsing"}
        entries = _parse_acc25_sudhir(file_bytes)
        _acc25_job["step"] = "loading ledger"
        ledger = load_ledger()
        existing_ids = {t["txn_id"] for t in ledger}
        added = 0
        _acc25_job["step"] = "merging"
        for e in entries:
            txn_id = f"acc25-{e['date'].replace('/','')}-{e['account']}-{int(e['debit'] or e['credit'])}"
            if txn_id in existing_ids:
                continue
            existing_ids.add(txn_id)
            ledger.append({
                "txn_id":          txn_id,
                "date":            e["date"],
                "account":         e["account"],
                "raw_description": e["raw_description"],
                "paid_to":         e["paid_to"],
                "debit":           e["debit"],
                "credit":          e["credit"],
                "type":            e["type"],
                "heading":         e["heading"],
                "note":            e["note"],
                "source":          "acc25_import",
                "uncertain":       False,
            })
            added += 1
        _acc25_job["step"] = "saving"
        _assign_seq(ledger)
        _save_json(LEDGER_PATH, ledger)
        _acc25_job = {"status": "done", "applied": added,
                      "total_parsed": len(entries), "ledger_total": len(ledger)}
    except Exception as ex:
        _acc25_job = {"status": "error", "error": str(ex)}


@app.route("/admin/acc25")
@login_required
def admin_acc25_page():
    return render_template("acc25_admin.html")


@app.route("/api/admin/acc25-preview", methods=["POST"])
@login_required
def api_acc25_preview():
    """Parse uploaded ACC25 xlsx (or local file if present), return entries for preview."""
    import threading
    f = request.files.get("file")
    if f:
        file_bytes = f.read()
    elif os.path.exists(_ACC25_LOCAL):
        with open(_ACC25_LOCAL, "rb") as fh:
            file_bytes = fh.read()
    else:
        return jsonify({"error": "no file uploaded and local ACC 25.xlsx not found"}), 400

    apply_it = request.form.get("apply", "") == "1"

    if apply_it:
        global _acc25_job
        _acc25_job = {"status": "running", "step": "starting"}
        t = threading.Thread(target=_run_acc25_import, args=(file_bytes,), daemon=True)
        t.start()
        return jsonify({"status": "running"})

    entries = _parse_acc25_sudhir(file_bytes)
    by_type = {}
    for e in entries:
        by_type[e["type"]] = by_type.get(e["type"], 0) + 1
    by_heading = {}
    for e in entries:
        h = e["heading"] or "(none)"
        by_heading[h] = by_heading.get(h, 0) + 1

    return jsonify({
        "total": len(entries),
        "by_type": by_type,
        "by_heading": dict(sorted(by_heading.items())),
        "sample": entries[:20],
    })


@app.route("/api/admin/acc25-status", methods=["GET"])
@login_required
def api_acc25_status():
    """Poll import job status."""
    return jsonify(_acc25_job)


_ACC24_LOCAL = os.path.join(BASE_DIR, "data", "acc24.xlsx")
_acc24_job = {}


def _parse_acc24_sudhir(file_bytes):
    """Parse SudhirExpenses sheet from ACC24 xlsx. Returns list of normalised ledger-ready dicts."""
    import io, openpyxl
    from datetime import datetime as _dt

    _ACCT_MAP = {
        "sbi4852":"SBI-4852","sbi":"SBI-4852",
        "icic0018":"ICICI-0018","ici0018":"ICICI-0018","icici0018":"ICICI-0018","ICI0018":"ICICI-0018","ICIC0018":"ICICI-0018",
        "icic7281":"ICICI-7281","ici7281":"ICICI-7281","icici7281":"ICICI-7281","ICIC7281":"ICICI-7281",
        "icic1331":"ICICI-1331","ici1331":"ICICI-1331","icici1331":"ICICI-1331","icci1331":"ICICI-1331",
        "ICIC1331":"ICICI-1331","Icici1331":"ICICI-1331",
        "icic9175":"ICICI-9175","ici9175":"ICICI-9175","icici9175":"ICICI-9175","ICIC9175":"ICICI-9175",
        "cridit card":"ICICI-CC","credit card":"ICICI-CC",
    }
    _TYPE_MAP = {
        "expense":"expense","official":"official","transfer":"transfer","transfer ":"transfer",
        "income":"income","tax":"expense","investment":"investment","error":"error",
    }
    _HEADING_NORM_ACC24 = {
        "alchol":"Alcohol","alcohol":"Alcohol",
        "amma":"Amma","AMMA":"Amma",
        "birthday parties":"Entertainment","books":"Entertainment","entertaiment":"Entertainment","entertainment":"Entertainment",
        "cash":"Cash",
        "charity":"Charity","donation":"Charity",
        "children education":"Children Education","children education ":"Children Education",
        "clothes":"Clothes","clothing":"Clothes","dry cleaning":"Clothes",
        "e&g":"Electricity & Gas","electricity & gas":"Electricity & Gas","electricity":"Electricity & Gas",
        "eating out":"Eating Out","eating out ":"Eating Out",
        "financial":"Financial Expense / OD Interest","financial expense":"Financial Expense / OD Interest",
        "gift":"Gifts","gifts":"Gifts",
        "groceries":"Groceries","grociries":"Groceries","grocries":"Groceries",
        "hoilday":"Holiday","holiday":"Holiday",
        "home office":"Home office","home office ":"Home office",
        "insurance":"Insurance",
        "ketki":"Ketki",
        "maintenance":"Maintenance Expense","maintenance expense":"Maintenance Expense",
        "malhar":"Malhar","malhar renovation":"Malhar",
        "medical":"Medical",
        "misc":"Misc","miscellaneous":"Misc",
        "one time charge":"One Time Charge","one time":"One Time Charge",
        "staff salary":"Staff Salary","staff slary":"Staff Salary","salary":"Staff Salary",
        "tax":"Tax","home loan":"Home Loan",
        "uspaar":"Uspaar",
        "wellness":"Wellness",
        "kalpataru":"Kalpataru Maintenance","kalpataru maintenance":"Kalpataru Maintenance",
    }

    def _parse_date_acc24(v):
        FY24_START = _dt(2023, 4, 1)
        FY24_END   = _dt(2024, 3, 31, 23, 59, 59)
        if isinstance(v, _dt):
            # Try original first
            if FY24_START <= v <= FY24_END:
                return v.strftime("%d/%m/%Y")
            # Try swapping month and day (Excel MM/DD vs DD/MM confusion)
            try:
                swapped = v.replace(month=v.day, day=v.month)
                if FY24_START <= swapped <= FY24_END:
                    return swapped.strftime("%d/%m/%Y")
            except ValueError:
                pass
            return v.strftime("%d/%m/%Y")  # return as-is, range filter will drop it
        if isinstance(v, str):
            v = v.strip()
            for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d-%m-%y", "%Y-%m-%d"):
                try:
                    return _dt.strptime(v, fmt).strftime("%d/%m/%Y")
                except ValueError:
                    pass
        return None

    def _norm_heading_acc24(h):
        if not h:
            return ""
        return _HEADING_NORM_ACC24.get(str(h).strip().lower(), str(h).strip())

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb["SudhirExpenses"]
    rows = list(ws.iter_rows(values_only=True))

    FY24_START = _dt(2023, 4, 1)
    FY24_END   = _dt(2024, 3, 31, 23, 59, 59)

    entries = []
    for i, row in enumerate(rows[1:], start=2):
        raw_date    = row[3]
        raw_acct    = str(row[2] or "").strip()
        raw_debit   = row[6]
        raw_credit  = row[7]
        raw_type    = str(row[9] or "").strip()   # type at col[9] for ACC24
        raw_desc    = str(row[4] or "").strip()
        raw_paid    = str(row[5] or "").strip()
        raw_heading = str(row[10] or "").strip()  # heading at col[10]

        date_str = _parse_date_acc24(raw_date)
        if not date_str:
            continue

        try:
            dt = _dt.strptime(date_str, "%d/%m/%Y")
        except ValueError:
            continue
        if not (FY24_START <= dt <= FY24_END):
            continue

        typ = _TYPE_MAP.get(raw_type.lower(), "expense")
        if typ == "error":
            continue

        debit  = float(raw_debit)  if raw_debit  else 0.0
        credit = float(raw_credit) if raw_credit else 0.0
        if debit < 0:
            credit = abs(debit)
            debit  = 0.0

        acct = _ACCT_MAP.get(raw_acct, _ACCT_MAP.get(raw_acct.lower(), raw_acct))
        heading = _norm_heading_acc24(raw_heading) if raw_heading else ""

        entries.append({
            "date":            date_str,
            "account":         acct,
            "raw_description": raw_desc,
            "paid_to":         raw_paid if raw_paid != "None" else "",
            "debit":           debit,
            "credit":          credit,
            "type":            typ,
            "heading":         heading,
            "note":            "",
            "source":          "acc24_import",
        })

    return entries


def _run_acc24_import(file_bytes):
    """Background thread: parse + apply ACC24 entries to master ledger."""
    from src.master_ledger import _assign_seq
    global _acc24_job
    try:
        _acc24_job = {"status": "running", "step": "parsing"}
        entries = _parse_acc24_sudhir(file_bytes)
        _acc24_job["step"] = "loading ledger"
        ledger = load_ledger()
        existing_ids = {t["txn_id"] for t in ledger}
        added = 0
        _acc24_job["step"] = "merging"
        for e in entries:
            txn_id = f"acc24-{e['date'].replace('/','')}-{e['account']}-{int(e['debit'] or e['credit'])}"
            if txn_id in existing_ids:
                continue
            existing_ids.add(txn_id)
            ledger.append({
                "txn_id":          txn_id,
                "date":            e["date"],
                "account":         e["account"],
                "raw_description": e["raw_description"],
                "paid_to":         e["paid_to"],
                "debit":           e["debit"],
                "credit":          e["credit"],
                "type":            e["type"],
                "heading":         e["heading"],
                "note":            e["note"],
                "source":          "acc24_import",
                "uncertain":       False,
            })
            added += 1
        _acc24_job["step"] = "saving"
        _assign_seq(ledger)
        _save_json(LEDGER_PATH, ledger)
        _acc24_job = {"status": "done", "applied": added,
                      "total_parsed": len(entries), "ledger_total": len(ledger)}
    except Exception as ex:
        _acc24_job = {"status": "error", "error": str(ex)}


@app.route("/admin/acc24")
@login_required
def admin_acc24_page():
    return render_template("acc24_admin.html")


@app.route("/api/admin/acc24-preview", methods=["POST"])
@login_required
def api_acc24_preview():
    """Parse uploaded ACC24 xlsx (or local file if present), return entries for preview."""
    import threading
    f = request.files.get("file")
    if f:
        file_bytes = f.read()
    elif os.path.exists(_ACC24_LOCAL):
        with open(_ACC24_LOCAL, "rb") as fh:
            file_bytes = fh.read()
    else:
        return jsonify({"error": "no file uploaded and local acc24.xlsx not found"}), 400

    apply_it = request.form.get("apply", "") == "1"

    if apply_it:
        global _acc24_job
        _acc24_job = {"status": "running", "step": "starting"}
        t = threading.Thread(target=_run_acc24_import, args=(file_bytes,), daemon=True)
        t.start()
        return jsonify({"status": "running"})

    entries = _parse_acc24_sudhir(file_bytes)
    by_type = {}
    for e in entries:
        by_type[e["type"]] = by_type.get(e["type"], 0) + 1
    by_heading = {}
    for e in entries:
        h = e["heading"] or "(none)"
        by_heading[h] = by_heading.get(h, 0) + 1

    return jsonify({
        "total": len(entries),
        "by_type": by_type,
        "by_heading": dict(sorted(by_heading.items())),
        "sample": entries[:20],
    })


@app.route("/api/admin/acc24-status", methods=["GET"])
@login_required
def api_acc24_status():
    """Poll import job status."""
    return jsonify(_acc24_job)


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


# ── WHATSAPP WEBHOOK ──────────────────────────────────────────────────────────

def _handle_sudhir_response(body: str):
    parts = body.strip().split()
    command = parts[0].upper()

    log = _load_json(APPROVAL_LOG)
    pending = [e for e in log if e.get("action") == "ESCALATE" and "sudhir_response" not in e]
    if not pending:
        return build_twiml_reply("No pending approvals found.")

    latest = pending[-1]
    request_id = latest["request_id"]
    submitter = latest["submitter"]
    vendor = latest["vendor"]
    amount = latest["amount"]

    engine.update_log_with_sudhir_response(request_id, body.strip())

    if command == "Y":
        send_approval_result(submitter, vendor, amount, approved=True, request_id=request_id)
        sync_approved_to_history()
        return build_twiml_reply(f"✅ Approved. {submitter.title()} notified.\nRef: {request_id}")
    elif command == "N":
        send_approval_result(submitter, vendor, amount, approved=False, request_id=request_id)
        return build_twiml_reply(f"❌ Rejected. {submitter.title()} notified.\nRef: {request_id}")
    elif command == "L" and len(parts) > 1:
        try:
            lower_amount = float(parts[1].replace(",", ""))
            send_approval_result(submitter, vendor, amount, approved=True,
                                 request_id=request_id, approved_amount=lower_amount)
            sync_approved_to_history()
            return build_twiml_reply(f"✅ Approved at Rs {lower_amount:,.0f}.\nRef: {request_id}")
        except ValueError:
            return build_twiml_reply("Invalid amount. Use: L 5000")

    return build_twiml_reply("Use Y / N / L <amount> to respond.")


def _handle_member_message(sender: str, body: str):
    submitter = NUMBER_TO_NAME.get(sender, "unknown")

    # Clarification reply?
    pending_id = next((rid for rid, req in PENDING_CLARIFICATION.items()
                       if req.submitter.lower() == submitter.lower()), None)
    if pending_id:
        req = PENDING_CLARIFICATION.pop(pending_id)
        req.description = f"{req.description} [{body.strip()}]"
        decision = engine.evaluate(req)
        if decision.action == "AUTO_APPROVE":
            send_auto_approval_notice(submitter, req.vendor, req.amount, decision.request_id)
            sync_approved_to_history()
            return build_twiml_reply(f"✅ Auto-approved.\nRef: {decision.request_id}")
        elif decision.action == "ESCALATE":
            send_approval_request(decision.escalation_message)
            return build_twiml_reply(f"📤 Sent to Sudhir.\nRef: {decision.request_id}")

    parts = [p.strip() for p in body.split(",")]
    if len(parts) < 5:
        return build_twiml_reply(
            "Format: Vendor, Amount, Category, Description, cash/upi\n"
            "Example: Swiggy, 850, dining, Dinner order, upi"
        )

    try:
        vendor = parts[0]
        amount = float(parts[1].replace("Rs", "").replace("rs", "").replace(",", "").strip())
        category = parts[2].lower().strip()
        description = parts[3]
        payment = parts[4].lower().strip()
        is_post_facto = len(parts) > 5 and "post" in parts[5].lower()
    except (ValueError, IndexError):
        return build_twiml_reply("Couldn't parse. Check format and try again.")

    req = ExpenseRequest(submitter=submitter, vendor=vendor, amount=amount,
                         category=category, description=description,
                         payment_method=payment, is_post_facto=is_post_facto)
    decision = engine.evaluate(req)

    if decision.action == "AUTO_APPROVE":
        send_auto_approval_notice(submitter, vendor, amount, decision.request_id)
        sync_approved_to_history()
        reply = f"✅ Auto-approved!\nRef: {decision.request_id}"
        if decision.budget_alert:
            reply += "\n⚠️ Category nearing monthly budget."
        return build_twiml_reply(reply)

    elif decision.action == "ESCALATE":
        send_approval_request(decision.escalation_message)
        reply = f"📤 Sent to Sudhir for approval.\nRef: {decision.request_id}"
        return build_twiml_reply(reply)

    elif decision.action == "PENDING_CLARIFICATION":
        PENDING_CLARIFICATION[decision.request_id] = req
        send_clarification_request(submitter, decision.follow_up_question,
                                   decision.follow_up_options, decision.request_id)
        return build_twiml_reply(f"❓ One question sent to clarify.\nRef: {decision.request_id}")

    return build_twiml_reply("Something went wrong. Please try again.")


@app.route("/webhook", methods=["POST"])
def webhook():
    incoming = parse_incoming(request.form)
    sender = incoming["from"]
    body = incoming["body"]
    if not body:
        return build_twiml_reply(""), 200
    if sender == SUDHIR:
        return app.response_class(_handle_sudhir_response(body), mimetype="text/xml")
    if sender in NUMBER_TO_NAME:
        return app.response_class(_handle_member_message(sender, body), mimetype="text/xml")
    return build_twiml_reply(""), 200


# ── SCHEDULER ────────────────────────────────────────────────────────────────

def _scheduled_reconciliation():
    try:
        run_reconciliation(notify_sudhir=True)
    except Exception as e:
        print(f"Reconciliation error: {e}")


if __name__ == "__main__":
    scheduler = BackgroundScheduler()
    scheduler.add_job(_scheduled_reconciliation, "interval", minutes=15)
    scheduler.start()
    port = int(os.getenv("PORT", 5000))
    print(f"Starting expense app on port {port}...")
    app.run(host="0.0.0.0", port=port, debug=False)
