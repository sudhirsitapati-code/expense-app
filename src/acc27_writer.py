"""
acc27_writer.py
Writes approved expenses to the Acc27 history file for future market-rate comparisons,
and exports monthly summaries to Excel.
"""

import json
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")

APPROVAL_LOG_PATH = os.path.join(DATA_DIR, "approval_log.json")
ACC27_HISTORY_PATH = os.path.join(DATA_DIR, "acc26_history.json")
EXPORTS_DIR = os.path.join(DATA_DIR, "exports")


def _load_json(path: str) -> list:
    if not os.path.exists(path):
        return []
    with open(path) as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            return []


def _save_json(path: str, data: list):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w") as f:
        json.dump(data, f, indent=2)


def sync_approved_to_history():
    """Append newly approved expenses to acc26_history.json for market-rate learning."""
    approval_log = _load_json(APPROVAL_LOG_PATH)
    history = _load_json(ACC27_HISTORY_PATH)

    existing_ids = {e.get("request_id") for e in history}
    added = 0

    for entry in approval_log:
        if entry.get("action") not in ("AUTO_APPROVE", "APPROVED", "APPROVED_LOWER"):
            continue
        if entry.get("request_id") in existing_ids:
            continue

        history.append({
            "request_id": entry["request_id"],
            "timestamp": entry["timestamp"],
            "vendor": entry["vendor"],
            "amount": entry.get("approved_amount") or entry["amount"],
            "category": entry["category"],
            "description": entry["description"],
            "payment_method": entry.get("payment_method", ""),
        })
        existing_ids.add(entry["request_id"])
        added += 1

    _save_json(ACC27_HISTORY_PATH, history)
    print(f"Synced {added} new approved expense(s) to history.")
    return added


def export_monthly_excel(year: int = None, month: int = None) -> str:
    """Export approved expenses for a given month to an Excel file."""
    now = datetime.now()
    year = year or now.year
    month = month or now.month
    month_prefix = f"{year}-{month:02d}"

    approval_log = _load_json(APPROVAL_LOG_PATH)
    rows = [
        e for e in approval_log
        if e.get("timestamp", "").startswith(month_prefix)
        and e.get("action") in ("AUTO_APPROVE", "APPROVED", "APPROVED_LOWER")
    ]

    os.makedirs(EXPORTS_DIR, exist_ok=True)
    filename = f"expenses_{month_prefix}.xlsx"
    filepath = os.path.join(EXPORTS_DIR, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Expenses {month_prefix}"

    headers = ["Date", "Submitter", "Vendor", "Category", "Description",
               "Amount (Rs)", "Payment", "Status", "Ref ID"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, entry in enumerate(rows, 2):
        date_str = entry.get("timestamp", "")[:10]
        amount = entry.get("approved_amount") or entry.get("amount", 0)
        ws.append([
            date_str,
            entry.get("submitter", "").title(),
            entry.get("vendor", ""),
            entry.get("category", ""),
            entry.get("description", ""),
            amount,
            entry.get("payment_method", ""),
            entry.get("action", ""),
            entry.get("request_id", ""),
        ])

    # Column widths
    col_widths = [12, 12, 25, 15, 35, 15, 12, 15, 22]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Summary row
    if rows:
        total_row = len(rows) + 2
        ws.cell(row=total_row, column=5, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=sum(
            (e.get("approved_amount") or e.get("amount", 0)) for e in rows
        )).font = Font(bold=True)

    wb.save(filepath)
    print(f"Exported {len(rows)} expense(s) to {filepath}")
    return filepath


if __name__ == "__main__":
    sync_approved_to_history()
    path = export_monthly_excel()
    print(f"Excel report: {path}")
